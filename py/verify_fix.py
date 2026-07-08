import sys; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl; wb=openpyxl.load_workbook('E:/workspace/github/film-tvs/res/data_new.xlsx')
for sn in ['电视剧资源','电影资源','动漫资源']:
    ws=wb[sn]
    issues = 0
    for r in range(2, ws.max_row+1):
        t=str(ws.cell(r,3).value or '')[:20]
        c8=str(ws.cell(r,8).value or '')
        c11=str(ws.cell(r,11).value or '')
        # Check: col8 has old-style path or "其他" while col11 has quark/baidu key
        if ('quark_' in c11 or 'baidu_' in c11) and (c8.startswith('其他') or c8.startswith('../') or 'baidu-menus' in c8 or 'res/' in c8):
            print(f'MISMATCH: {sn} row{r}: {t} | C8="{c8}" | C11="{c11}"')
            issues += 1
    if not issues:
        print(f'{sn}: All OK!')
print('\nVerification complete')
