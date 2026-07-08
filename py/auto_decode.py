"""
Auto-decode garbled names based on "every even char → pinyin initial" pattern.
Uses pypinyin to reverse-lookup characters by pinyin initial.
"""
import sys, os, json, re
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from py.name_map_tool import make_session, get_stoken, list_all, extract_pwd
from py.name_map import NameMapper
import openpyxl
import pypinyin

# Build reverse index: letter → list of (character, frequency_order)
# First get all common Chinese characters
COMMON_CHARS = list("的一是在不了有和人这中大为上个国我以要他时来用们生到作地于出就分对成会可主发年动同工也能下过子说产种面而方后多定行学法所民得经十三之进着等部度家电力里如水化高自二理起小物现实加量都两体制机当使点从业本去把性好应开它合还因由其些然前外天政四日那社义事平形相全表间样与关各重新线内数正心反你明看原又么利比或但质气第向道命此变条只没结解问意建月公无系军很情者最立代想已通并提直题党程展五果料象员革位入常文总次品式活设及管特件长求老头基资边流路级少图山统接知较将组见计别她手角期根论运农指几九区强放决西被干做必战先回则任取据处队南给色光门即保治北造百规热领七海口东导器压志世金增争济阶油思术极交受联什认六共权收证改清己美再采转更单风切打白教速花带安场身车例真务具万每目至达走积示议声报斗完类八离华名确才科张信马节话米整空元况今集温传土许步群广石记需段研界拉林律叫且究观越织装影算低持音众书布复容儿须际商非验连断深难近矿千周委素技备半办青省列习响约支般史感劳便团往酸历市克何除消构府称太准精值号率族维划选标写存候毛亲快效斯院查江型眼王按格养易置派层片始却专状育厂京识适属圆包火住调满县局照参红细引听该铁价严龙飞")

# Build pinyin lookup
PINYIN_INITIALS = {}
for ch in COMMON_CHARS:
    initials = set()
    for p in pypinyin.lazy_pinyin(ch):
        if p:
            initials.add(p[0].upper())
    for init in initials:
        if init not in PINYIN_INITIALS:
            PINYIN_INITIALS[init] = []
        PINYIN_INITIALS[init].append(ch)

def try_garbled_match(garbled, candidates_dict):
    """Check if a specific replacement makes a valid Chinese word/phrase"""
    # Simple check: try all replacements and see which one makes sense
    best_score = -1
    best_replacement = None
    
    # Find letter positions
    segments = re.split(r'([A-Z])', garbled)
    
    # Try all combinations (combinatorial explosion, limit to simple cases)
    # For each letter, we have candidates from PINYIN_INITIALS
    letters = [c for c in garbled if c.isupper()]
    positions = [i for i, c in enumerate(garbled) if c.isupper()]
    
    if not letters:
        return None
    
    # Simple approach: for each letter position, pick the most common character
    result = list(garbled)
    decoded = {}
    for letter, pos in zip(letters, positions):
        candidates = PINYIN_INITIALS.get(letter, [])
        if not candidates:
            continue
        # Pick the character that makes the most sense in context
        # For now, pick the most common one (first in frequency list)
        decoded[pos] = candidates[0]
        result[pos] = candidates[0]
    
    return ''.join(result)

def auto_decode_name(garbled_name):
    """
    Auto-decode a garbled file name.
    Pattern: Chinese characters at odd positions are kept,
    at even positions are replaced by pinyin initial letter.
    """
    # Extract the base name (remove path, extension)
    name = garbled_name
    
    # Split into segments: Chinese chars and letters
    # Pattern: Chinese[-]Letter pattern keeps alternating
    # We need to handle: [Chinese] [-Letter] [-Chinese] [Letter] etc.
    
    # Simple approach: find each uppercase letter between Chinese chars
    # and replace it with the most likely character
    
    letters_found = []
    
    # Find all uppercase letters that are part of garbling
    # They appear between Chinese characters (possibly with hyphens)
    for m in re.finditer(r'(?<=[\u4e00-\u9fff])-?([A-Z])-?(?=[\u4e00-\u9fff])', name):
        letters_found.append(m)
    
    result = list(name)
    for m in reversed(letters_found):
        letter = m.group(1)
        start, end = m.start(1), m.end(1)
        candidates = PINYIN_INITIALS.get(letter, [letter])
        if candidates:
            result[start:end] = candidates[0]  # Use most common
    
    return ''.join(result)

def batch_decode_entries(pwd_id):
    """Fetch entries from API, auto-decode, and set name mappings"""
    EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'res', 'data_new.xlsx')
    
    # Find link
    wb = openpyxl.load_workbook(EXCEL_PATH)
    link = ''
    for sn in [n for n in wb.sheetnames if n != 'index']:
        ws = wb[sn]
        for r in range(2, ws.max_row+1):
            lk = str(ws.cell(r, 10).value or '')
            if pwd_id in lk:
                link = lk; break
        if link: break
    
    if not link:
        print(f'Cannot find link for {pwd_id}')
        return
    
    _, passcode = extract_pwd(link)
    s = make_session()
    stoken = get_stoken(s, pwd_id, passcode)
    if not stoken:
        print(f'STOKEN FAIL for {pwd_id}')
        return
    
    dirs, files = list_all(s, pwd_id, stoken)
    entries = dirs + files
    entries.sort(key=lambda x: (0 if x['is_dir'] else 1, x['name'].lower()))
    
    nm = NameMapper()
    existing = nm.get_all(pwd_id)
    
    auto_mapped = 0
    for e in entries:
        fid = e['fid']
        orig = e['name']
        if fid in existing:
            continue  # Already mapped
        # Auto-decode
        decoded = auto_decode_name(orig)
        if decoded and decoded != orig and orig != '007.mp4' and '新用户' not in orig:
            print(f'  {orig[:60]}')
            print(f'  → {decoded[:60]}')
            nm.set(pwd_id, fid, decoded)
            auto_mapped += 1
    
    print(f'\nAuto-decoded and mapped {auto_mapped} entries for {pwd_id}')
    
    # Save
    nm.save()
    return entries

if __name__ == '__main__':
    if len(sys.argv) > 1:
        pwd_id = sys.argv[1]
        batch_decode_entries(pwd_id)
    else:
        print('Usage: python py/auto_decode.py <pwd_id>')
