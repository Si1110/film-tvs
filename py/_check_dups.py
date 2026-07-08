#!/usr/bin/env python3
"""Check for duplicate cards in generated HTML and data"""
import re, os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__)))
from collections import Counter
from openpyxl import load_workbook

# Check HTML
with open('sections/section-01.html', 'r', encoding='utf-8') as f:
    content = f.read()

card_titles = re.findall(r'<h5[^>]*class="card-title[^"]*"[^>]*>\s*([^<]+?)\s*<', content)
print(f'Total cards in HTML: {len(card_titles)}')

title_counts = Counter(card_titles)
dups = {t: c for t, c in title_counts.items() if c > 1}
print(f'Duplicate titles: {len(dups)}')

sorted_dups = sorted(dups.items(), key=lambda x: -x[1])
print('\nTop 30 duplicates (with series info):')
for title, count in sorted_dups[:30]:
    print(f'  [{count}x] {title}')

# Check data for duplicates
wb = load_workbook('res/data_new.xlsx')
ws = wb['电视剧资源']
headers = [str(c.value) for c in ws[1]]
ti = headers.index('主标题') + 1
si = headers.index('所属系列') + 1
li = headers.index('下载链接') + 1

data_titles = []
for r in range(2, ws.max_row + 1):
    title = str(ws.cell(r, ti).value or '').strip()
    series = str(ws.cell(r, si).value or '').strip()
    link = str(ws.cell(r, li).value or '').strip()
    data_titles.append((title, series, link, r))

data_counts = Counter(t for t, _, _, _ in data_titles)
data_dups = {t: c for t, c in data_counts.items() if c > 1}
print(f'\nDuplicate titles in Excel: {len(data_dups)}')

# Show some duplicates with their details
for title, count in sorted(data_dups.items(), key=lambda x: -x[1])[:15]:
    entries = [(s, l[:30], r) for t, s, l, r in data_titles if t == title]
    print(f'\n  [{count}x] {title}:')
    for s, l, r in entries:
        print(f'    Row {r}: series="{s[:30]}", link={l}...')

# Check for same-title, same-link duplicates
print('\n\n=== Checking same title + same link (likely true duplicates) ===')
title_link = Counter((t, l) for t, _, l, _ in data_titles if t and l)
for (t, l), c in title_link.items():
    if c > 1:
        rows = [r for tt, _, ll, r in data_titles if tt == t and ll == l]
        print(f'  [{c}x] {t} | {l[:40]} | rows: {rows}')

wb.close()
