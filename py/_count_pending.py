import sys, re, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

quark_ids = set()
baidu_ids = set()
wb = load_workbook('res/data_new.xlsx')
for sn in wb.sheetnames:
    if sn == 'index': continue
    ws = wb[sn]
    for row in range(2, ws.max_row+1):
        c10 = str(ws.cell(row, 10).value or '')
        m = re.search(r'pan\.quark\.cn/s/([a-f0-9]+)', c10)
        if m: quark_ids.add(m.group(1))
        m = re.search(r'pan\.baidu\.com/s/([a-zA-Z0-9_-]+)', c10)
        if m: baidu_ids.add(m.group(1))

existing_dirs = {f.split('.')[0] for f in os.listdir('res/dirs') if f.endswith('.html')}
quark_missing = {x for x in quark_ids if f'quark_{x}' not in existing_dirs}
baidu_missing = {x for x in baidu_ids if f'baidu_{x}' not in existing_dirs}

print(f'Unique Quark pwd_ids: {len(quark_ids)}, missing files: {len(quark_missing)}')
print(f'Unique Baidu surls: {len(baidu_ids)}, missing files: {len(baidu_missing)}')
