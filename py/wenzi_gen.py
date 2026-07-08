import sys, json, urllib.parse, os, re
sys.stdout.reconfigure(encoding='utf-8')
import requests

cookie = os.environ.get('QUARK_COOKIE', '')
s = requests.Session()
for part in cookie.split(';'):
    if '=' in part.strip(): k,v=part.strip().split('=',1); s.cookies.set(k,v)

def get_stoken(pwd_id):
    r = s.post('https://drive-h.quark.cn/1/clouddrive/share/sharepage/token?pr=ucpro&fr=pc',
               json={'pwd_id':pwd_id,'passcode':'','support_visit_limit_private_share':True},
               headers={'User-Agent':'Mozilla/5.0','Content-Type':'application/json','Referer':'https://pan.quark.cn/'}, timeout=15)
    return r.json().get('data',{}).get('stoken','')

def list_all(pwd_id, stoken, pdir_fid='0', depth=0):
    dirs, files = [], []
    page = 1
    while True:
        url = ('https://drive-h.quark.cn/1/clouddrive/share/sharepage/detail?pr=ucpro&fr=pc&ver=2'
               '&pwd_id=' + pwd_id + '&stoken=' + urllib.parse.quote(stoken) +
               '&pdir_fid=' + str(pdir_fid) + '&force=0&_page=' + str(page) + '&_size=200&_fetch_total=1')
        try:
            r = s.get(url, headers={'User-Agent':'Mozilla/5.0','Referer':'https://pan.quark.cn/'}, timeout=15)
            d = r.json()
            if d.get('status')!=200: break
            items = d.get('data',{}).get('list',[])
            if not items: break
            for item in items:
                entry={'fid':item.get('fid',''),'name':item.get('file_name',''),'is_dir':item.get('dir',False),'depth':depth}
                if entry['is_dir']:
                    dirs.append(entry)
                    sd, sf = list_all(pwd_id, stoken, entry['fid'], depth+1)
                    dirs.extend(sd); files.extend(sf)
                else:
                    files.append(entry)
            total = d.get('data',{}).get('total',0)
            if page*200 >= total: break
            page += 1
        except: break
    return dirs, files

pwd_id = '7617bff6bde0'
stoken = get_stoken(pwd_id)
if not stoken:
    print('STOKEN FAIL')
    sys.exit(1)

dirs, files = list_all(pwd_id, stoken)
entries = dirs + files
entries.sort(key=lambda x: (0 if x['is_dir'] else 1, x['name'].lower()))

# Generate HTML
promo_words = ['限时', '新用户', '转存', '领取', '1TB', '签到', '永久容量', '图文指引', '每日签到', '手机端']
lines = []
for e in entries:
    name = e['name']
    skip = any(w in name for w in promo_words)
    if skip: continue
    indent_px = e['depth'] * 24
    icon = '📁' if e['is_dir'] else '📄'
    color = '#ffd700' if e['is_dir'] else '#aaa'
    lines.append(f'<div style="margin-left:{indent_px}px;color:{color}">{icon} {name}</div>')

content = '\n'.join(lines)
html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><title>温子仁系列 - 目录</title>
<style>
body {{ font-family: 'Microsoft YaHei', sans-serif; background: #1a1a2e; padding: 20px; font-size: 14px; line-height: 1.8; }}
a {{ color: #4fc3f7; text-decoration: none; }}
a:hover {{ text-decoration: underline; }}
</style>
</head><body>
{content}
</body></html>'''

dir_path = r'E:\workspace\github\film-tvs\res\dirs\温子仁系列.html'
with open(dir_path, 'w', encoding='utf-8') as f:
    f.write(html)
print(f'Generated {dir_path} with {len(lines)} entries')

# Update Excel: row85 col8 = 温子仁系列
import openpyxl
wb = openpyxl.load_workbook(r'E:\workspace\github\film-tvs\res\data_new.xlsx')
ws = wb['电影资源']
for r in range(2, ws.max_row+1):
    t = str(ws.cell(r,3).value or '')
    if '温子仁' in t and not str(ws.cell(r,8).value or ''):
        ws.cell(r,8).value = '温子仁系列'
        print(f'Updated row{r} col8 -> 温子仁系列')
wb.save(r'E:\workspace\github\film-tvs\res\data_new.xlsx')
print('Excel saved')
