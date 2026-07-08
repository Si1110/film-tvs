import os, sys, json, urllib.parse, time, re
sys.stdout.reconfigure(encoding='utf-8')
import requests
import openpyxl

COOKIE_STR = r"b-user-id=47000940-681b-3d39-3ad3-ffd51e0bef3a; _UP_A4A_11_=wba2d199ca6d4156bbf065be9709fec0; _c_WBKFRo=geWYEMxjjBz1P0w0V5kjZMCOIKZTvcVAWZmcfw6a; _UP_335_2B_=1; __sdid=AASjLyqoKdwFTpUnYB15zKlS3acNl/8PW/P0YAW4kN8dfIgFKJhVplRuuO3OK1LOpOiG8RyVyKLGVP5N19E2DjX5rJk/Y6tnlcNmc+YuFhk1lw==; _UP_D_=pc; _UP_30C_6A_=sta2e6201d15vu1n67adq3lv00werwuk; _UP_TS_=sg1ace6b77f3ab536d7cbec3937c8f84f8e; _UP_E37_B7_=sg1ace6b77f3ab536d7cbec3937c8f84f8e; _UP_TG_=sta2e6201d15vu1n67adq3lv00werwuk; tfstk=g26ivNjmF1Rs76gUEqJ_tqnJD7Fpfd9XRZHvkKL4Te8BklsA0KkVkip9gPsA-E8ebl6T5h8miaxNBV8aSwqcWiMv7FBvnI7JQI19kFLDni_uvze8eGs6hQz8y8BZ1lMJhEyv3WpU4gCt1ze8eGPwfKIUysK21uxv0K-wghuUxnKqgK-2ukYem3Gq7ZJ4xkx2cAk2_fyULhtM3E723kjeRnJ2bZJ4xM8B0jFb_F6VL9zSfXp9Rs1cKhAMzGVS3xXneCYPbeD4z9xghUSw-xk2PQmzrMjTSYspCt7k2NeEIajVVsRhn2zkl9jFQsSISkJlx6CJs64ZneB9mQJ2tmkVxpYBw9pzU2xCTGCVCNmm3h6OeIYWtokXMpWJa_7icuseQn7XNtUSCFSVVT1JER0Wj6SkKgRiTb-jIxtUD9ljGC-BxUnM7RQnxS6UdkqnaoOwAhT8xkcjGC-BxUE3xbyX_Ht6y; isg=BJ6eJiY2eP8mPqyBlRZBSGaP7zTgX2LZsPI_N0gnCuHcaz5FsO-y6cQIY3fn01rx; __pus=cfef8172d4b5c92d7d734315fd53ee47AAQxWjik0hG2opNiwmrnKLuL47LP6ewPS/HbECKHyA2QMswU/GcB81es/VV8rSN+K5ezgGipMCnRILw7RqLHx6Kx; __kp=571f8620-6c60-11f1-9146-55a822775a5f; __kps=AASj3UTlfAF8TNsc+2TeHr0r; __ktd=rdKtNeSsloEmLNyAxt/hVQ==; __uid=AASj3UTlfAF8TNsc+2TeHr0r; Video-Auth=Qsm/8v6XH0LY1llR1w94JDjXETtFT5Yb91LlXTFhu0U1g73EE5gIpCX9KFhEriYAxH/DNBCuvjKZ/9NWc6RaWqzIFGFEEipHavgsEa+AlJhMQXrLRBUD/4KffSw04p7xvD/9QepLa0cjitR4BBOviA==; __puus=21d2a6f8f99315e2a1ea9521fd809f8eAAQ271TZHk/+1L76Ex31OOfLj3KZKtUgJMlLiCpobHGg73k7AughQFofRl8LLS3amjWnzZz64CA22Zt6U38AIIBq0zEVGilSR1Kk31Ki894HTYwNcCkYgW4edu09CTf8m2CGtHSZbRXmUOhaFGoUnuP4qBw6r88eiXgCZvVPfv2vUuDdS7jylOPnOXcFeNcWLE7/W3u6bGRofRQTAhbEIvNh"

s = requests.Session()
for part in COOKIE_STR.split(';'):
    if '=' in part.strip():
        k, v = part.strip().split('=', 1)
        s.cookies.set(k, v)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Referer': 'https://pan.quark.cn/',
    'Content-Type': 'application/json',
}

def get_stoken(pwd_id, passcode=''):
    try:
        r = s.post('https://drive-h.quark.cn/1/clouddrive/share/sharepage/token?pr=ucpro&fr=pc',
                   json={'pwd_id': pwd_id, 'passcode': passcode, 'support_visit_limit_private_share': True},
                   headers=HEADERS, timeout=10)
        d = r.json()
        return d['data']['stoken'] if d.get('status') == 200 else None
    except:
        return None

def list_folder_recursive(pwd_id, stoken, pdir_fid='0', depth=0):
    dirs, files = [], []
    page = 1
    while True:
        try:
            url = ('https://drive-h.quark.cn/1/clouddrive/share/sharepage/detail'
                   '?pr=ucpro&fr=pc&ver=2&pwd_id=' + pwd_id +
                   '&stoken=' + urllib.parse.quote(stoken) +
                   '&pdir_fid=' + str(pdir_fid) +
                   '&force=0&_page=' + str(page) + '&_size=200&_fetch_total=1')
            r = s.get(url, headers=HEADERS, timeout=15)
            d = r.json()
            if d.get('status') != 200: break
            items = d.get('data', {}).get('list', [])
            if not items: break
            for item in items:
                entry = {'fid': item.get('fid', ''), 'name': item.get('file_name', ''),
                         'is_dir': item.get('dir', False), 'depth': depth, 'size': item.get('size', 0)}
                if entry['is_dir']:
                    dirs.append(entry)
                    sd, sf = list_folder_recursive(pwd_id, stoken, entry['fid'], depth + 1)
                    dirs.extend(sd); files.extend(sf)
                else:
                    files.append(entry)
            total = d.get('data', {}).get('total', 0)
            if page * 200 >= total: break
            page += 1
        except: break
    return dirs, files

def generate_quark_html(entries):
    entries.sort(key=lambda x: (0 if x['is_dir'] else 1, x['name'].lower()))
    lines = ['<div class="dir-list">',
             '  <div class="dir-section mb-3">',
             '    <h6 style="color:#ffd700;border-bottom:1px solid rgba(255,215,0,0.2);padding-bottom:8px;margin-bottom:12px;">',
             '      <i class="bi bi-folder2-open"></i> 夸克网盘目录',
             '    </h6>']
    for e in entries:
        indent = e['depth'] * 24
        icon = "📁" if e['is_dir'] else "📄"
        color = "#ffd700" if e['is_dir'] else "#aaa"
        lines.append(f'    <div style="margin-left:{indent}px;color:{color};font-size:0.9rem;" class="mb-1">{icon} {e["name"]}</div>')
    lines.append('  </div>\n</div>')
    return '\n'.join(lines)

def generate_baidu_placeholder(surl, title):
    return f'<div class="text-center py-4"><p class="text-muted mb-3">该资源目录请前往百度网盘查看：</p><a href="https://pan.baidu.com/s/{surl}" target="_blank" class="btn btn-warning px-4 py-2" style="font-weight:600;">打开百度网盘 <i class="bi bi-box-arrow-up-right"></i></a><p class="text-muted mt-3 small">点击后在新标签页中浏览文件列表</p></div>'

DIRS_DIR = 'res/dirs'
os.makedirs(DIRS_DIR, exist_ok=True)
existing_dirs = {f.split('.')[0] for f in os.listdir(DIRS_DIR) if f.endswith('.html')}

wb = openpyxl.load_workbook('res/data_new.xlsx')

# Collect what's still needed
quark_pending = {}  # pwd_id -> passcode
baidu_pending = {}  # surl -> first title

for sn in [n for n in wb.sheetnames if n != 'index']:
    ws = wb[sn]
    for row in range(2, ws.max_row + 1):
        c10 = str(ws.cell(row, 10).value or '')
        m = re.search(r'pan\.quark\.cn/s/([a-f0-9]+)', c10)
        if m:
            pwd_id = m.group(1)
            if f'quark_{pwd_id}' not in existing_dirs and pwd_id not in quark_pending:
                passcode = re.search(r'[?&]pwd=([^&]+)', c10)
                quark_pending[pwd_id] = passcode.group(1) if passcode else ''
        m = re.search(r'pan\.baidu\.com/s/([a-zA-Z0-9_-]+)', c10)
        if m:
            surl = m.group(1)
            if f'baidu_{surl}' not in existing_dirs:
                title = str(ws.cell(row, 3).value or '')
                baidu_pending.setdefault(surl, title)

print(f'Remaining: {len(quark_pending)} Quark, {len(baidu_pending)} Baidu')

# 1. Baidu placeholders (fast)
for surl, title in baidu_pending.items():
    with open(os.path.join(DIRS_DIR, f'baidu_{surl}.html'), 'w', encoding='utf-8') as f:
        f.write(generate_baidu_placeholder(surl, title))
print(f'Baidu placeholders: {len(baidu_pending)}')

# 2. Remaining Quark dirs
ok = fail = 0
for pwd_id, passcode in quark_pending.items():
    print(f'  Q [{pwd_id}] ', end='', flush=True)
    stoken = get_stoken(pwd_id, passcode)
    if not stoken:
        print('FAIL')
        fail += 1
        continue
    all_dirs, all_files = list_folder_recursive(pwd_id, stoken)
    html = generate_quark_html(all_dirs + all_files)
    with open(os.path.join(DIRS_DIR, f'quark_{pwd_id}.html'), 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'{len(all_dirs)+len(all_files)} ✓')
    ok += 1
print(f'Quark done: {ok} ok, {fail} fail')

# 3. Update Excel column 8 (目录路径) for ALL entries
print('\nUpdating Excel column 8 (目录路径)...')
updates = 0
for sn in [n for n in wb.sheetnames if n != 'index']:
    ws = wb[sn]
    for row in range(2, ws.max_row + 1):
        c8 = str(ws.cell(row, 8).value or '').strip()
        c10 = str(ws.cell(row, 10).value or '')
        new_key = ''
        m = re.search(r'pan\.quark\.cn/s/([a-f0-9]+)', c10)
        if m: new_key = f'quark_{m.group(1)}'
        else:
            m = re.search(r'pan\.baidu\.com/s/([a-zA-Z0-9_-]+)', c10)
            if m: new_key = f'baidu_{m.group(1)}'
        if new_key and c8 != new_key:
            ws.cell(row, 8).value = new_key
            updates += 1

wb.save('res/data_new.xlsx')
wb.close()
print(f'Updated {updates} Excel entries')
print('All done!')
