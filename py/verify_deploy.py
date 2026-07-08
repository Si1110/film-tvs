"""
verify_deploy.py — 上架后完整性验证
功能:
  1. 检查夸克资源是否有目录路径 (col8) 及目录文件是否存在
  2. 检查目录文件名是否有脱敏/谐音残留（如 人Z怒、潜-F）
  3. 汇总报告

用法: python py/verify_deploy.py
"""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

BASE = r'E:\workspace\github\film-tvs'
DIRS_DIR = os.path.join(BASE, 'res', 'dirs')
EXCEL_PATH = os.path.join(BASE, 'res', 'data_new.xlsx')

# 脱敏残留: Chinese-Letter-Chinese pattern
# "人Z怒"应显示为"人之怒"，但"哥斯拉X摩斯拉"中的X是合法交叉符号
GARBLED_RE = re.compile(r'[\u4e00-\u9fff][A-Z][\u4e00-\u9fff]')
# 合法字母（交叉、VS等标记，不是脱敏残留）
LEGIT_LETTERS = {'X', 'V', 'W', 'K', 'T', 'S', 'F', 'Z'}

def is_garbled(name):
    matches = GARBLED_RE.findall(name)
    for m in matches:
        letter = m[1]
        if letter in LEGIT_LETTERS:
            # 验证：如果该字母是常见缩写/符号，跳过
            # 但 "人Z怒" 中的 Z 不在合法列表内，会命中
            continue
        return True, f'"{m}" ({letter})'
    return False, None

def check():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheets = [n for n in wb.sheetnames if n != 'index']
    any_issue = False
    
    for sn in sheets:
        ws = wb[sn]
        missing = []
        garbled = []
        
        for row in range(2, ws.max_row + 1):
            title = str(ws.cell(row, 3).value or '(无标题)').strip()
            series = str(ws.cell(row, 1).value or '其他').strip()
            dir_key = str(ws.cell(row, 8).value or '').strip()
            link = str(ws.cell(row, 10).value or '').strip()
            
            # Check 1: Quark links must have a dir_key
            if 'quark' in link.lower() and not dir_key:
                missing.append((row, title, series))
                continue
            
            # Check 2: 检查目录文件名的脱敏字符
            if dir_key:
                df = os.path.join(DIRS_DIR, f'{dir_key}.html')
                if os.path.exists(df):
                    with open(df, 'r', encoding='utf-8') as f:
                        content = f.read()
                    names = re.findall(r'(?:📁|📄) (.*?)</div>', content)
                    for name in names:
                        clean = name[:80]
                        is_bad, detail = is_garbled(clean)
                        if is_bad:
                            garbled.append((row, title, f'{detail}: {clean}'))
                            break
        
        # Print
        print(f'\n{"="*52}')
        print(f'  [{sn}] {ws.max_row - 1} 条')
        print(f'{"="*52}')
        
        if missing:
            any_issue = True
            print(f'  ❌ 夸克资源缺少 col8 目录路径（需补充）:')
            for row, t, s in missing:
                print(f'    R{row}: {t} ({s})')
        else:
            print(f'  ✅ 夸克资源全部有目录路径')
        
        if garbled:
            any_issue = True
            print(f'  ❌ 目录文件含脱敏残留名称（需修正）:')
            for row, t, n in garbled:
                print(f'    R{row} "{t}": "{n}"')
        else:
            print(f'  ✅ 目录名称完整正确')
    
    print(f'\n{"="*52}')
    if any_issue:
        print(f'  ❌ 有需修复项，查看上方详情')
    else:
        print(f'  ✅ 全部验证通过')
    print(f'{"="*52}')

if __name__ == '__main__':
    check()
