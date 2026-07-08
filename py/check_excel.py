import sys; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl; wb=openpyxl.load_workbook('E:/workspace/github/film-tvs/res/data_new.xlsx')
for sn in ['电视剧资源','电影资源','动漫资源']:
    ws=wb[sn]
    print(f'\n=== {sn} headers ===')
    for c in range(1, ws.max_column+1):
        print(f'  Col {c}: "{ws.cell(1,c).value}"')
