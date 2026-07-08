#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
一键上架脚本 - film-tvs 影视网站新增资源工作流

用法:
    python py/listing.py 上架 --name "S 射雕英雄传【7个版本大合集】" [--sheet 电视剧]
    python py/listing.py 上架 --from-source  # 交互式从 0607 源表上架
    python py/listing.py 上架 --id 36        # 从源表指定行号上架

功能:
    1. 从 0607 源表读取新增资源
    2. 自动补全字段并写入 data_new.xlsx
    3. 从夸克 API 读取文件夹结构 → 生成目录 HTML (res/dirs/{key}.html)
    4. 从豆瓣 API 下载封面图
    5. 重新生成网站 HTML
    6. 自动 git commit + push
"""

import os
import re
import sys
import json
import time
import argparse
import urllib.parse
import subprocess
from datetime import datetime

# ============================================================
# 配置
# ============================================================
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_EXCEL = os.environ.get('FILM_TVS_SOURCE', r'F:\1、自媒体\3、网站\影视\影视动漫新增（0607）.xlsx')
DATA_EXCEL = os.path.join(PROJECT_ROOT, 'res', 'data_new.xlsx')
COVERS_DIR = os.path.join(PROJECT_ROOT, 'res', 'covers')
DIRS_DIR = os.path.join(PROJECT_ROOT, 'res', 'dirs')
GENERATE_SCRIPT = os.path.join(PROJECT_ROOT, 'py', 'generate_html.py')

# 源表 sheet → 目标 sheet 映射
SHEET_MAP = {
    '动漫': '动漫资源',
    '电视剧': '影视资源',
    '电影': '影视资源',
}

# ============================================================
# 夸克 API
# ============================================================
class QuarkAPI:
    """夸克网盘 API 封装（使用 Cookie 认证）"""
    
    API_TOKEN = 'https://drive-h.quark.cn/1/clouddrive/share/sharepage/token?pr=ucpro&fr=pc'
    API_DETAIL = 'https://drive-h.quark.cn/1/clouddrive/share/sharepage/detail'
    
    def __init__(self, cookie=None):
        self.cookie = cookie or os.environ.get('QUARK_COOKIE', '')
        self._session = None
        self._stoken_cache = {}
    
    @property
    def session(self):
        if self._session is None:
            import requests
            s = requests.Session()
            if self.cookie:
                for part in self.cookie.split(';'):
                    part = part.strip()
                    if '=' in part:
                        k, v = part.split('=', 1)
                        s.cookies[k.strip()] = v.strip()
            self._session = s
        return self._session
    
    def get_stoken(self, pwd_id):
        """获取分享 stoken"""
        if pwd_id in self._stoken_cache:
            return self._stoken_cache[pwd_id]
        r = self.session.post(self.API_TOKEN, json={
            'pwd_id': pwd_id, 'passcode': '',
            'support_visit_limit_private_share': True,
        }, headers=self._headers(), timeout=15)
        data = r.json()
        stoken = data['data']['stoken']
        self._stoken_cache[pwd_id] = stoken
        return stoken
    
    def list_folder(self, pwd_id, pdir_fid='0'):
        """列出分享文件夹内容"""
        stoken = self.get_stoken(pwd_id)
        url = (f'{self.API_DETAIL}?pr=ucpro&fr=pc&ver=2'
               f'&pwd_id={pwd_id}&stoken={urllib.parse.quote(stoken)}'
               f'&pdir_fid={pdir_fid}&force=0&_page=1&_size=200&_fetch_total=1')
        r = self.session.get(url, headers=self._headers(), timeout=15)
        d = r.json()
        if d.get('status') == 200 and 'data' in d:
            return d['data']['list']
        return []
    
    def build_tree(self, pwd_id, pdir_fid='0', depth=0):
        """递归构建目录树"""
        files = self.list_folder(pwd_id, pdir_fid)
        tree = []
        for f in files:
            entry = {
                'name': f['file_name'],
                'dir': f.get('dir', False),
                'fid': f['fid'],
                'size': f.get('size', 0),
            }
            if entry['dir']:
                entry['children'] = self.build_tree(pwd_id, entry['fid'], depth + 1)
                entry['count'] = f.get('include_items', len(entry.get('children', [])))
            tree.append(entry)
        # Sort: dirs first, then by name
        tree.sort(key=lambda x: (not x['dir'], x['name']))
        return tree
    
    def _headers(self):
        return {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Referer': 'https://pan.quark.cn/',
            'Origin': 'https://pan.quark.cn',
            'Content-Type': 'application/json',
        }


# ============================================================
# 豆瓣 API
# ============================================================
class DoubanAPI:
    """豆瓣 API 封装"""
    
    SEARCH_URL = 'https://movie.douban.com/j/subject_suggest'
    
    def search(self, query):
        """搜索电影/电视剧，返回 [{id, title, year, img, subtype}]"""
        import requests
        url = f'{self.SEARCH_URL}?q={urllib.parse.quote(query)}'
        r = requests.get(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        }, timeout=15)
        return r.json() if r.status_code == 200 else []
    
    def download_poster(self, img_url, save_path):
        """下载封面图并转为 WebP"""
        import requests
        from PIL import Image
        import io
        
        # Try large size first, then fallback
        for size in ['l', 'm']:
            url = re.sub(r'/ [a-z]/', f'/{size}/', img_url)
            r = requests.get(url, headers={
                'User-Agent': 'Mozilla/5.0',
                'Referer': 'https://movie.douban.com/',
            }, timeout=15)
            if len(r.content) > 5000:
                img = Image.open(io.BytesIO(r.content))
                webp_path = save_path.replace('.png', '.webp').replace('.jpg', '.webp')
                img.save(webp_path, 'WEBP', quality=90)
                # Also save JPG for compatibility
                jpg_path = webp_path.replace('.webp', '.jpg')
                img.save(jpg_path, 'JPEG', quality=90)
                return webp_path
        return None


# ============================================================
# 目录 HTML 生成
# ============================================================
def generate_dir_html(tree, title='目录'):
    """从夸克目录树生成目录 HTML（支持多级文件夹嵌套）"""
    def render_items(items, depth=0):
        html = ''
        ml = depth * 24
        for item in items:
            if item['dir']:
                count = item.get('count', len(item.get('children', [])))
                badge = f'<span class="badge bg-secondary ms-2" style="font-size:0.7rem;">{count} 项</span>'
                html += f'<div style="margin-left:{ml}px" class="mb-1">\n'
                html += f'  <div class="d-flex align-items-center mb-1">\n'
                html += f'    <span style="color:#ffd700;font-weight:bold;">&#x1f4c1; {item["name"]}</span>\n'
                html += f'    {badge}\n'
                html += f'  </div>\n'
                if item.get('children'):
                    html += render_items(item['children'], depth + 1)
                html += f'</div>\n'
            else:
                html += f'<div style="margin-left:{ml}px;color:#aaa;font-size:0.9rem;" class="mb-1">&#x1f4c4; {item["name"]}</div>\n'
        return html
    
    body = render_items(tree)
    
    return f'''<div class="dir-list">
  <div class="dir-section mb-3">
    <h6 style="color:#ffd700;border-bottom:1px solid rgba(255,215,0,0.2);padding-bottom:8px;margin-bottom:12px;">
      <i class="bi bi-folder2-open"></i> 夸克网盘目录
    </h6>
    {body}
  </div>
  <div class="text-muted small mt-2" style="border-top:1px solid rgba(255,255,255,0.06);padding-top:10px;">
    <i class="bi bi-info-circle"></i> 以上目录仅供参考，具体以夸克网盘实际内容为准
  </div>
</div>'''


# ============================================================
# Excel 操作
# ============================================================
class ExcelManager:
    """Excel 读写封装"""
    
    def __init__(self, path):
        self.path = path
        self.wb = None
    
    def open(self):
        from openpyxl import load_workbook
        self.wb = load_workbook(self.path)
        return self
    
    def close(self):
        if self.wb:
            self.wb.close()
    
    def save(self):
        self.wb.save(self.path)
    
    def get_headers(self, sheet_name):
        ws = self.wb[sheet_name]
        return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    
    def add_row(self, sheet_name, data):
        """添加一行数据到指定 sheet"""
        ws = self.wb[sheet_name]
        headers = self.get_headers(sheet_name)
        new_row = ws.max_row + 1
        for col_name, value in data.items():
            if col_name in headers:
                ws.cell(new_row, headers[col_name], value)
        return new_row
    
    def get_all_series(self, sheet_name):
        """获取指定 sheet 中所有已经存在的系列名"""
        ws = self.wb[sheet_name]
        headers = self.get_headers(sheet_name)
        col = headers.get('所属系列')
        if not col:
            return set()
        series = set()
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, col).value
            if v:
                series.add(str(v))
        return series


# ============================================================
# 源表读取
# ============================================================
def read_source_excel(path=None):
    """读取 0607 源表，返回 {sheet_name: [{名称, 来源, 地址链接}]}"""
    path = path or SOURCE_EXCEL
    if not os.path.exists(path):
        print(f'❌ 找不到源表: {path}')
        return None
    
    from openpyxl import load_workbook
    wb = load_workbook(path)
    result = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        data = []
        for r in range(2, ws.max_row + 1):
            name = str(ws.cell(r, 1).value or '').strip()
            source = str(ws.cell(r, 2).value or '').strip()
            link = str(ws.cell(r, 3).value or '').strip()
            if name:
                data.append({'名称': name, '来源': source, '地址链接': link})
        result[sn] = data
    wb.close()
    return result


def parse_source_name(name):
    """解析源表名称，返回 {title, subtitle, series, year, versions}
    
    例如: 'S 射雕英雄传【7个版本大合集】' → {prefix: 'S', title: '射雕英雄传', ...}
    """
    item = {'raw': name, 'prefix': '', 'title': name, 'subtitle': '', 'versions': ''}
    
    # Extract prefix (single letter + space)
    m = re.match(r'^([A-Z])\s+(.*)', name)
    if m:
        item['prefix'] = m.group(1)
        item['title'] = m.group(2)
    
    title = item['title']
    
    # Extract 大合集 with version count
    m = re.search(r'【(\d+)个版本大合集】', title)
    if m:
        item['versions'] = f'{m.group(1)}个版本大合集'
        item['subtitle'] = item['versions']
        item['title'] = title.split('【')[0]
    
    # Extract 合集
    m = re.search(r'(.+?)合集', title)
    if m:
        if not item['versions']:
            item['subtitle'] = '合集'
            item['title'] = m.group(1)
    
    return item


def auto_generate_data(source_item, target_sheet):
    """从源表条目自动生成 data_new.xlsx 所需的所有字段"""
    parsed = parse_source_name(source_item['名称'])
    link = source_item.get('地址链接', '')
    source_name = source_item.get('来源', '夸克网盘')
    
    # 提取 pwd_id
    pwd_id = ''
    m = re.search(r'/s/([a-zA-Z0-9]+)', link)
    if m:
        pwd_id = m.group(1)
    
    series_name = parsed['title'] + '系列'
    title = parsed['title']
    subtitle = parsed['subtitle']
    
    # 语言默认值
    if target_sheet == '动漫资源':
        lang = '日语'
        sub = '中文字幕'
    else:
        lang = '汉语普通话'
        sub = '中文字幕'
    
    # 封面路径
    cover_key = re.sub(r'[^\w\u4e00-\u9fff]+', '_', title)
    cover_path = f'../res/covers/{series_name}/{cover_key}.webp'
    
    # 目录路径（使用系列名作为 key）
    menu_path = series_name
    
    # 判断 Sheet
    if target_sheet == '动漫资源':
        foot = ''
        formats = 'mp4'
    else:
        foot = ''
        formats = 'mp4/mkv'
    
    return {
        '所属系列': series_name,
        '封面图片路径': cover_path,
        '主标题': title,
        '副标题': subtitle,
        '概要': f'{title}合集，高清资源下载。',
        '语言': lang,
        '字幕': sub,
        '目录路径': menu_path,
        '网盘名称': source_name,
        '下载链接': link,
        '解压密码': '',
        '支持格式': formats,
        '卡片页脚': foot,
        '_pwd_id': pwd_id,  # internal use
        '_cover_key': cover_key,
        '_series_name': series_name,
    }


def interactive_select(items, prompt='请选择:'):
    """交互式选择"""
    for i, item in enumerate(items, 1):
        print(f'  [{i}] {item}')
    while True:
        try:
            choice = input(f'{prompt} ')
            idx = int(choice) - 1
            if 0 <= idx < len(items):
                return idx
        except ValueError:
            pass
        print('输入无效，请重新选择')


# ============================================================
# 图片工具
# ============================================================
def ensure_cover_dir(series_name):
    """确保封面目录存在"""
    d = os.path.join(COVERS_DIR, series_name)
    os.makedirs(d, exist_ok=True)
    return d


# ============================================================
# 目录 HTML
# ============================================================
def ensure_dirs_dir():
    os.makedirs(DIRS_DIR, exist_ok=True)


# ============================================================
# Git 操作
# ============================================================
def git_commit_push(message):
    """自动 commit 并 push"""
    try:
        subprocess.run(['git', 'add', '-A'], cwd=PROJECT_ROOT, check=True, capture_output=True)
        subprocess.run(['git', 'commit', '-m', message], cwd=PROJECT_ROOT, check=True, capture_output=True)
        subprocess.run(['git', 'push'], cwd=PROJECT_ROOT, check=True, capture_output=True)
        print('✅ 已自动 commit + push')
    except subprocess.CalledProcessError as e:
        print(f'⚠️ Git 操作失败: {e.stderr.decode() if e.stderr else e}')


# ============================================================
# 主命令: 上架
# ============================================================
def cmd_list(args):
    """上架命令"""
    print('=' * 60)
    print('  📦 film-tvs 一键上架工具')
    print('=' * 60)
    
    # 1. 读取源表
    source_data = read_source_excel()
    if not source_data:
        return
    
    # 2. 选择源表 sheet
    sheet_names = [s for s in source_data if source_data[s]]
    print(f'\n📋 源表包含 {len(sheet_names)} 个分类:')
    for i, sn in enumerate(sheet_names, 1):
        print(f'  [{i}] {sn} ({len(source_data[sn])} 条)')
    
    if args.id is not None:
        # 按行号在所有 sheet 中查找
        for sn in sheet_names:
            for item in source_data[sn]:
                sid = source_data[sn].index(item) + 1
                if sid == args.id:
                    selected_sheet = sn
                    selected_item = item
                    break
            else:
                continue
            break
        else:
            print(f'❌ 未找到行号 {args.id}')
            return
    elif args.name:
        # 按名称搜索
        for sn in sheet_names:
            for item in source_data[sn]:
                if args.name in item['名称']:
                    selected_sheet = sn
                    selected_item = item
                    break
            else:
                continue
            break
        else:
            print(f'❌ 未找到 "{args.name}"')
            return
    else:
        # 交互式选择
        sheet_idx = interactive_select(sheet_names, '选择分类编号:')
        selected_sheet = sheet_names[sheet_idx]
        
        print(f'\n📝 {selected_sheet} 条目列表:')
        items = source_data[selected_sheet]
        for i, item in enumerate(items, 1):
            print(f'  [{i}] {item["名称"]}  ({item["来源"]})')
        
        item_idx = interactive_select([f'{item["名称"]} ({item["来源"]})' for item in items], '选择条目编号:')
        selected_item = items[item_idx]
    
    print(f'\n✅ 选中: {selected_sheet} → {selected_item["名称"]}')
    
    # 3. 解析并生成数据
    target_sheet = SHEET_MAP.get(selected_sheet, '影视资源')
    data = auto_generate_data(selected_item, target_sheet)
    
    print(f'  目标 Sheet: {target_sheet}')
    print(f'  系列名: {data["所属系列"]}')
    print(f'  标题: {data["主标题"]}')
    print(f'  链接: {data["下载链接"]}')
    
    pwd_id = data.pop('_pwd_id')
    cover_key = data.pop('_cover_key')
    series_name = data.pop('_series_name')
    
    # 4. 尝试获取夸克目录
    dir_tree = []
    if pwd_id:
        print('\n🔍 正在读取夸克文件夹...')
        try:
            quark = QuarkAPI()
            folder_fid = '0'
            # First get root listing to find the real folder fid
            root_files = quark.list_folder(pwd_id)
            for f in root_files:
                if f.get('dir') and selected_item['名称'] in f['file_name']:
                    folder_fid = f['fid']
                    break
                elif f.get('dir') and not folder_fid != '0':
                    folder_fid = f['fid']  # First dir found
            dir_tree = quark.build_tree(pwd_id, folder_fid)
            
            # Generate dir HTML
            ensure_dirs_dir()
            dir_html = generate_dir_html(dir_tree, data['主标题'])
            dir_path = os.path.join(DIRS_DIR, f'{series_name}.html')
            with open(dir_path, 'w', encoding='utf-8') as f:
                f.write(dir_html)
            print(f'✅ 已生成目录: res/dirs/{series_name}.html')
            
            # Update description with version info
            version_parts = []
            for item in dir_tree:
                if item['dir']:
                    version_parts.append(item['name'])
            if version_parts:
                data['概要'] = f'收录了{", ".join(version_parts[:5])}等版本的{data["主标题"]}电视剧大全集。'
                if len(version_parts) > 5:
                    data['概要'] = f'收录了{", ".join(version_parts[:5])}等{len(version_parts)}个版本的{data["主标题"]}大全集。'
                # Update card footer
                total_eps = sum(
                    len(item.get('children', [])) for item in dir_tree if item['dir']
                )
                if total_eps:
                    data['卡片页脚'] = f'{len(version_parts)}个版本 共约{total_eps}集'
        except Exception as e:
            print(f'⚠️ 夸克读取失败: {e}')
            print('  将手动创建简单目录')
    
    # 5. 尝试获取豆瓣封面
    cover_dir = ensure_cover_dir(series_name)
    cover_found = False
    search_queries = [data['主标题'], data['主标题'] + ' ' + data.get('副标题', '')]
    
    for query in search_queries:
        if not query.strip():
            continue
        print(f'\n🔍 正在搜索豆瓣封面: "{query}"...')
        try:
            douban = DoubanAPI()
            results = douban.search(query)
            if results:
                # Pick the best match (prefer TV series)
                best = results[0]
                for r in results:
                    if r.get('subtype') == 'tv':
                        best = r
                        break
                print(f'  找到: {best["title"]} ({best.get("year", "?")})')
                poster_url = best.get('img', '')
                if poster_url:
                    save_path = os.path.join(cover_dir, cover_key + '.jpg')
                    result = douban.download_poster(poster_url, save_path)
                    if result:
                        print(f'✅ 封面已保存: {result}')
                        cover_found = True
                        break
        except Exception as e:
            print(f'⚠️ 豆瓣搜索失败: {e}')
    
    if not cover_found:
        print('⚠️ 未找到豆瓣封面，使用占位图')
        # Generate placeholder
        from PIL import Image, ImageDraw
        img = Image.new('RGB', (500, 700), (30, 30, 50))
        draw = ImageDraw.Draw(img)
        try:
            from PIL import ImageFont
            font = ImageFont.truetype('C:/Windows/Fonts/msyhbd.ttc', 36)
        except:
            font = ImageFont.load_default()
        draw.text((50, 300), data['主标题'], fill='#ffd700', font=font)
        webp_path = os.path.join(cover_dir, cover_key + '.webp')
        img.save(webp_path, 'WEBP', quality=85)
        print(f'⚠️ 已生成占位封面: {webp_path}')
    
    # 6. 写入 Excel
    print('\n📝 正在写入 Excel...')
    try:
        excel = ExcelManager(DATA_EXCEL).open()
        # Check for duplicates
        existing = excel.get_all_series(target_sheet)
        if data['所属系列'] in existing:
            yn = input(f'⚠️ 系列 "{data["所属系列"]}" 已存在，是否仍然添加？(y/N): ')
            if yn.lower() != 'y':
                print('❌ 已取消')
                excel.close()
                return
        
        row = excel.add_row(target_sheet, data)
        excel.save()
        excel.close()
        print(f'✅ 已写入 {target_sheet} 第 {row} 行')
    except Exception as e:
        print(f'❌ Excel 写入失败: {e}')
        return
    
    # 7. 重新生成 HTML
    print('\n⚙️  正在重新生成网站 HTML...')
    result = subprocess.run([sys.executable, GENERATE_SCRIPT], 
                          cwd=PROJECT_ROOT, capture_output=True, text=True)
    if result.returncode == 0:
        print('✅ HTML 生成完成')
    else:
        print(f'❌ HTML 生成失败:\n{result.stdout[-500:]}\n{result.stderr[-500:]}')
        return
    
    # 8. 自动 commit + push
    if not args.no_push:
        commit_msg = f'上架 {data["主标题"]}'
        if data.get('副标题'):
            commit_msg += f' {data["副标题"]}'
        print(f'\n📤 正在推送: {commit_msg}')
        git_commit_push(commit_msg)
    
    print('\n' + '=' * 60)
    print('  ✅ 上架完成！')
    print('=' * 60)


# ============================================================
# 主入口
# ============================================================
def main():
    parser = argparse.ArgumentParser(description='film-tvs 一键上架工具')
    sub = parser.add_subparsers(dest='command')
    
    # 上架命令
    p = sub.add_parser('上架', help='一键上架新资源')
    p.add_argument('--name', help='资源名称（支持模糊匹配）')
    p.add_argument('--id', type=int, help='源表行号')
    p.add_argument('--from-source', action='store_true', help='交互式从源表选择')
    p.add_argument('--no-push', action='store_true', help='不上传 Git')
    
    # 查看命令
    p2 = sub.add_parser('查看', help='查看源表内容')
    p2.add_argument('--sheet', help='分类名称')
    p2.add_argument('--all', action='store_true', help='显示所有')
    
    # 配置命令
    p3 = sub.add_parser('配置', help='查看/设置配置')
    p3.add_argument('--show', action='store_true', help='显示当前配置')
    
    args = parser.parse_args()
    
    if args.command == '上架':
        if args.from_source or not (args.name or args.id is not None):
            args.name = None
            args.id = None
        cmd_list(args)
    elif args.command == '查看':
        cmd_show(args)
    elif args.command == '配置':
        print(f'源表路径: {SOURCE_EXCEL}')
        print(f'数据路径: {DATA_EXCEL}')
        print(f'夸克 Cookie: {"已设置" if os.environ.get("QUARK_COOKIE") else "未设置（使用环境变量 QUARK_COOKIE）"}')
    else:
        parser.print_help()


def cmd_show(args):
    data = read_source_excel()
    if not data:
        return
    for sn, items in data.items():
        if args.sheet and sn != args.sheet:
            continue
        print(f'\n=== {sn} ({len(items)} 条) ===')
        for i, item in enumerate(items, 1):
            print(f'  [{i}] {item["名称"]}  →  {item["来源"]}')


if __name__ == '__main__':
    main()
