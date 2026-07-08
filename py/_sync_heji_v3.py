import os, shutil, re, sys, io
from openpyxl import load_workbook

# Redirect stdout to UTF-8 file
log_path = "E:\\workspace\\github\\film-tvs\\res\\heji_sync_log.txt"
sys.stdout = io.open(log_path, 'w', encoding='utf-8')

missing_dir = r"F:\1、自媒体\3、网站\影视\缺失封面"
all_files = []
for cat in ["电影", "电视剧", "动漫"]:
    d = os.path.join(missing_dir, cat)
    if not os.path.isdir(d): continue
    for f in os.listdir(d):
        fp = os.path.join(d, f)
        if os.path.isfile(fp):
            all_files.append((cat, f, fp))

print(f"Total files in 缺失封面: {len(all_files)}")

def extract_core(title):
    m = re.match(r'^([^【]+)', title)
    return m.group(1).strip() if m else title.strip()

results = []
wb = load_workbook("res/data_new.xlsx")
for sn in wb.sheetnames:
    ws = wb[sn]
    h = [c.value for c in ws[1]]
    ti = next((i for i, hv in enumerate(h) if hv and "主标题" in str(hv)), None)
    ci = next((i for i, hv in enumerate(h) if hv and "封面图片路径" in str(hv)), None)
    if ti is None or ci is None: continue
    
    for row in ws.iter_rows(min_row=2):
        title = str(row[ti].value) if row[ti].value else ""
        if not title or "【" not in title or "】" not in title:
            continue
        
        core = extract_core(title)
        current = str(row[ci].value) if row[ci].value else ""
        current_is_subitem = bool(current and "【" not in current and "】" not in current)
        
        # Priority: already has a dedicated cover in 合集封面 -> skip
        if current and "合集封面" in current:
            print(f"  [{sn:8s}] DEDICATED OK [{title:40s}] {current}")
            results.append((title, "DEDICATED_OK", current, ""))
            continue
        
        match = None
        for ext in [".webp", ".jpg", ".jpeg", ".png"]:
            cand = title + ext
            for cat, fname, fp in all_files:
                if fname == cand:
                    match = (cat, fname, fp)
                    break
            if match: break
        
        if not match:
            for cat, fname, fp in all_files:
                base = os.path.splitext(fname)[0]
                if base.startswith(core) and "【" in base and "】" in base:
                    match = (cat, fname, fp)
                    break
        
        if not match:
            for cat, fname, fp in all_files:
                base = os.path.splitext(fname)[0]
                if core in base and "【" in base and "】" in base:
                    match = (cat, fname, fp)
                    break
        
        if match:
            cat, fname, src_path = match
            target_map = {"电影资源": "res/film/合集封面", "电视剧资源": "res/tv/合集封面", "动漫资源": "res/anime/合集封面"}
            target_dir = target_map.get(sn, "res/covers")
            os.makedirs(target_dir, exist_ok=True)
            
            dst = os.path.join(target_dir, fname)
            shutil.copy2(src_path, dst)
            rel = os.path.relpath(dst, "res").replace("\\", "/")
            new_cover = f"../res/{rel}"
            
            if current != new_cover:
                row[ci].value = new_cover
                flag = "SUBITEM→DEDICATED" if current_is_subitem else "UPDATED"
                print(f"  [{sn:8s}] {flag:18s} [{title:40s}] → {new_cover}")
                results.append((title, flag, current, new_cover))
            else:
                print(f"  [{sn:8s}] SKIP(SAME)   [{title:40s}] {current}")
                results.append((title, "SAME", current, ""))
        else:
            print(f"  [{sn:8s}] NO MATCH      [{title:40s}] current=[{current[:80]}]")
            results.append((title, "NO_MATCH", current, ""))

# wb.save("res/data_new.xlsx")
wb.close()

updated = [r for r in results if r[1] in ("UPDATED", "SUBITEM→DEDICATED")]
nomatch = [r for r in results if r[1] == "NO_MATCH"]
dedicated_ok = [r for r in results if r[1] == "DEDICATED_OK"]
same = [r for r in results if r[1] == "SAME"]
print(f"\n{'='*60}")
print(f"Total合集: {len(results)}")
print(f"  DEDICATED_OK: {len(dedicated_ok)} (already using 合集封面)")
print(f"  UPDATED: {len(updated)} (cover replaced)")
print(f"  NO MATCH: {len(nomatch)} (no file found in 缺失封面)")
print(f"  SAME: {len(same)}")

print(f"\n{'='*60}")
print("NO MATCH entries (need manual review):")
for r in nomatch:
    print(f"  Sheet=[{r[0]}] Cover=[{r[2][:80]}]")

sys.stdout.close()
print(f"Log written to {log_path}", flush=True)
