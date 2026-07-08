import sys; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl; wb=openpyxl.load_workbook('E:/workspace/github/film-tvs/res/data_new.xlsx')

total = 0
for sn in ['电视剧资源','电影资源','动漫资源']:
    ws=wb[sn]
    for r in range(2, ws.max_row+1):
        c8 = str(ws.cell(r,8).value or '')    # 目录路径
        c11 = str(ws.cell(r,11).value or '')   # 解压密码
        
        # If col11 has a valid dir key (quark_ or baidu_ or other key without special chars)
        # and col8 still has an old-style value like "其他系列" or a path
        if c11.startswith('quark_') or c11.startswith('baidu_') or (c11 and '/' not in c11 and c11 not in ['','无']):
            # Check if col8 has the wrong old value
            if c8.startswith('其他') or c8.startswith('../') or 'baidu-menus' in c8 or 'res/' in c8:
                ws.cell(r, 8).value = c11
                # Clear col11 (解压密码) if it was mistakenly set
                # But only clear if it was set to a dir key, not an actual password
                if c11.startswith('quark_') or c11.startswith('baidu_'):
                    ws.cell(r, 11).value = ''
                total += 1
                print(f'{sn} row{r}: Col8="{c8}" -> "{c11}" (col11 cleared)')

wb.save('E:/workspace/github/film-tvs/res/data_new.xlsx')
print(f'\nFixed {total} rows')
