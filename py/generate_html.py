#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 Excel 生成 HTML 页面脚本
功能：
1. 从 Sheet0 生成 index.html
2. 从 Sheet1-n 生成 section-01.html ~ section-n.html
"""

import os
import re
import random
import string
from datetime import datetime
from openpyxl import load_workbook
from jinja2 import Template
from color_log.clog import log
from theme_utils import is_heji, extract_theme

# ===========================
# 配置
# ===========================
CHASET = 'utf-8'
SPLIT_LINE = "=" * 60

# 网站数据文件
EXCEL_FILE = "./res/data_new.xlsx"
SHEET_INDEX = 'index'

# 网站配置文件
GLOBAL_FILE = "./res/README.md"
GLOBAL_CONFIG = {}

# 生成网页模板文件
TEMPLATE_SECTION = "./templates/section-tpl.html"  # section 模板文件
TEMPLATE_CARD = "./templates/card-tpl.html"  # card 模板文件
TEMPLATE_INDEX = "./templates/index-tpl.html"  # index 模板文件
TEMPLATE_COVER = "./templates/cover-tpl.html"  # cover 模板文件
TEMPLATE_SITEMAP = "./templates/sitemap-tpl.html"  # sitemap 模板文件
TEMPLATE_LINK = "./templates/link-tpl.html"  # sitemap link 模板文件

# 输出路径
OUTPUT_DIR = "./sections"  # 站点分类 sections 输出路径
OUTPUT_INDEX = "./index.html"  # 首页 index 输出文件
OUTPUT_SITEMAP = './sitemap.xml' # 搜索引擎 sitemap

# ===========================
# 主流程
# ===========================
def main():
    log.info(SPLIT_LINE)
    log.info("HTML 生成器 - 从 Excel 生成网站页面")
    log.info(SPLIT_LINE)
    
    # 加载配置变量
    load_config()
    
    # 加载 Excel 数据
    log.info(f"📖 读取 Excel 文件: {EXCEL_FILE}")
    if not os.path.exists(EXCEL_FILE):
        log.info(f"❌ 错误: 找不到 {EXCEL_FILE}")
        return
    sheets = load_excel_data(EXCEL_FILE)
    log.info(f"✓ 找到 {len(sheets)} 个 Sheet")
    
    # Sheet 分类（首页 + section）
    sheet_list = list(sheets.keys())
    index_sheet = SHEET_INDEX if SHEET_INDEX in sheets else (sheet_list[0] if sheet_list else None)
    section_sheets = [s for s in sheet_list if s != SHEET_INDEX]

    # 创建输出目录
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    log.info(f"✓ 输出目录: {OUTPUT_DIR}/")

    # 生成网页
    generate_sections(sheets, section_sheets)
    if index_sheet:
        generate_index(sheets, index_sheet, section_sheets)
    
    # 更新 JS 文件配置
    update_js_config(sheets, section_sheets)
    
    # 生成 sitemap.xml
    generate_sitemap(sheets, section_sheets)
    
    log.info(SPLIT_LINE)
    log.info("✅ 生成完成！")
    log.info(SPLIT_LINE)
    
    

# ===========================
# 工具函数 - 读取配置
# ===========================
def load_config():
    global GLOBAL_CONFIG
    
    if not os.path.exists(GLOBAL_FILE):
        log.warn(f"⚠️ 找不到 {GLOBAL_FILE}，使用默认配置")
        GLOBAL_CONFIG = {}
        return
    
    try:
        with open(GLOBAL_FILE, 'r', encoding=CHASET) as f:
            content = f.read()
        
        # 匹配所有代码块中的 KEY=VALUE 对
        # 匹配 ```\nKEY=VALUE\nKEY=VALUE\n...```
        pattern = r'```\n([\s\S]*?)\n```'
        matches = re.findall(pattern, content)
        
        for block in matches:
            # 从每个代码块中提取 KEY=VALUE
            lines = block.strip().split('\n')
            for line in lines:
                line = line.strip()
                if '=' in line and not line.startswith('#'):
                    key, value = line.split('=', 1)
                    GLOBAL_CONFIG[key.strip()] = value.strip()
        
        log.info(f"✓ 从 README 加载 {len(GLOBAL_CONFIG)} 个配置")
        
    except Exception as e:
        log.warn(f"⚠️ 读取 README 配置失败: {e}")
        GLOBAL_CONFIG = {}
        

# Sheet 1-n 生成 section HTML
def generate_sections(sheets, sheet_list) :
    for idx, sheet_name in enumerate(sheet_list):
        idx += 1
        html_name = f'section-{idx:02d}.html'
        log.info(f"📄 开始转换 Sheet {sheet_name} -> {html_name}")
        
        is_ok, header, lines = load_sheet_table(sheets, sheet_name)
        if not is_ok:
            log.warn(f"⚠️  {sheet_name} 没有数据，跳过")
            continue
        # log.debug(header)
        # log.debug(lines)
        
        
        # 生成 HTML
        html_content = generate_section(lines, sheet_name, idx)
        
        # 保存文件
        html_path = os.path.join(OUTPUT_DIR, html_name)
        with open(html_path, 'w', encoding=CHASET, newline='') as f:
            f.write(html_content)
        
        log.info(f"✓ 生成 {len(lines)} 个卡片")
        log.info(f"✓ 已保存到: {html_path}")


def generate_section(sheet_data, section_title, section_number):
    """生成 section HTML（扁平卡片网格 + 筛选导航）"""
    row_idx = 0
    for d in sheet_data:
        row_idx += 1
        d['_row_idx'] = row_idx

    # Step 1: assign each item a theme key
    for d in sheet_data:
        d['_theme'] = extract_theme(str(d.get('\u4e3b\u6807\u9898', '')))

    # Step 2: group by theme, compute max row per theme
    theme_groups = {}
    for d in sheet_data:
        th = d['_theme']
        theme_groups.setdefault(th, []).append(d)
    theme_maxrow = {th: max(d['_row_idx'] for d in members) for th, members in theme_groups.items()}

    # Step 3: sort themes by -maxrow (newest theme first)
    sorted_themes = sorted(theme_groups.keys(), key=lambda th: -theme_maxrow[th])

    # Step 4: within each theme:合集 first, then by title
    sorted_items = []
    for th in sorted_themes:
        group = theme_groups[th]
        group.sort(key=lambda x: (0 if is_heji(str(x.get('\u4e3b\u6807\u9898', ''))) else 1,
                                  str(x.get('\u4e3b\u6807\u9898', ''))))
        sorted_items.extend(group)

    cards_list = []
    for idx, d in enumerate(sorted_items, 1):
        cards_list.append(generate_card_html(d, idx))
    section_content = '\n\n'.join(cards_list)

    section_filename = f'section-{section_number:02d}.html'
    # 统计各系列条目数，仅多条目系列用于描述（避免单条目系列污染 description，如 007 各分集）
    from collections import Counter
    series_counter = Counter()
    for d in sorted_items:
        s = str(d.get('所属系列', ''))
        if s and s != '其他系列':
            series_counter[s] += 1
    multi_series = sorted((s for s, c in series_counter.items() if c >= 3), key=lambda s: -series_counter[s])
    if len(multi_series) < 3:
        multi_series = sorted(series_counter.keys(), key=lambda s: -series_counter[s])
    desc_series = ','.join(multi_series[:6])

    key_series = ','.join(sorted(set(
        str(d.get('所属系列', '')) for d in sorted_items
        if str(d.get('所属系列', '')) != '其他系列'
    ))[:12])

    if section_number == 1:
        section_desc = f"{section_title}合集导航，收录{desc_series}等经典剧集，覆盖日剧、港剧、国产剧与热门悬疑、青春、武侠作品。"
        section_keys = f"90后,经典电视剧,日剧,{key_series},高清下载,中文字幕,mp4,港剧,国产剧,悬疑剧,武侠剧,日剧下载"
        genres = ['古装', '悬疑', '喜剧', '犯罪', '奇幻', '剧情', '爱情', '惊悚', '科幻']
    elif section_number == 2:
        section_desc = f"{section_title}合集导航，收录{desc_series}等经典电影系列，覆盖好莱坞大片、华语电影、日影、恐怖、科幻与喜剧作品。"
        section_keys = f"90后,经典电影,好莱坞,{key_series},高清下载,中文字幕,mp4,恐怖,科幻,喜剧,经典电影下载"
        genres = ['喜剧', '动作', '科幻', '爱情', '悬疑', '奇幻', '剧情', '恐怖', '犯罪', '惊悚', '冒险']
    elif section_number == 3:
        section_desc = f"{section_title}合集导航，收录{desc_series}等经典动漫与番剧，覆盖热血、治愈、悬疑、科幻、异世界与童年怀旧题材。"
        section_keys = f"90后,经典动漫,番剧,动漫下载,日本动漫,{key_series},高清下载,中文字幕,mp4,动漫合集"
        genres = ['喜剧', '冒险', '动作', '运动', '剧情', '奇幻', '科幻', '爱情', '悬疑', '恐怖']
    else:
        section_desc = f"{section_title}合集"
        section_keys = GLOBAL_CONFIG.get('SEO_KEYWORDS', '')
        genres = ['古装', '悬疑', '喜剧', '犯罪', '动作', '奇幻', '剧情', '爱情', '惊悚']

    # 限制 description 长度（SEO 建议 50-160 字符）
    if len(section_desc) > 155:
        section_desc = section_desc[:152] + '...'
    # 限制 keywords 长度（避免过长稀释权重）
    if len(section_keys) > 300:
        parts = section_keys.split(',')
        # 保留前部核心词，从尾部丢弃多余
        while len(','.join(parts)) > 290 and len(parts) > 10:
            parts.pop()
        section_keys = ','.join(parts)

    heji_count = sum(1 for d in sorted_items if is_heji(str(d.get('主标题', ''))))

    context = {
        'section_title': section_title,
        'section_number': section_number,
        'section_filename': section_filename,
        'total_count': len(sorted_items),
        'section_content': section_content,
        'section_description': section_desc,
        'section_keywords': section_keys,
        'genres': genres,
        'heji_count': heji_count,
        **GLOBAL_CONFIG
    }

    tpl_text = load_template_section()
    tpl = Template(tpl_text)
    return tpl.render(**context)


def normalize_cover_path(path):
    """统一封面路径：确保以 ../res/ 开头（相对于 sections/ 目录）"""
    if not path:
        return '../res/placeholder.webp'
    path = str(path).replace('\\', '/')
    if path.startswith('../res/'):
        return path
    # 处理 res/covers/... 格式（缺少 ../ 前缀）
    if path.startswith('res/'):
        return f'../{path}'
    # 缺少 ../res/covers/ 前缀（如 "其他系列/xxx.webp"）
    return f'../res/covers/{path}'

def get_netdisk_name(download_link, fallback='百度网盘'):
    """根据下载链接识别网盘名称，字段缺失时使用 fallback。"""
    link = str(download_link or '').lower()
    if 'quark.cn' in link:
        return '夸克网盘'
    if 'baidu.com' in link:
        return '百度网盘'
    if 'aliyundrive.com' in link or 'alipan.com' in link:
        return '阿里云盘'
    if 'uc.cn' in link:
        return 'UC网盘'
    return fallback or '网盘下载'

def generate_card_html(card_data, card_idx=0):
    """生成单个电影卡片 HTML（从模板渲染）"""

    # 准备字段
    download_link = card_data.get('下载链接', 'FIXME')
    context = {
        'card_id': f'card-desc-{card_idx}',
        'poster_path': normalize_cover_path(card_data.get('封面图片路径', '')), 
        'title': card_data.get('主标题', ''),
        'subtitle': card_data.get('副标题', ''),
        'description': card_data.get('概要', ''),
        'language': card_data.get('语言', '日语'),
        'subtitle_text': card_data.get('字幕', '中文'),
        'menu_path': card_data.get('目录路径', ''), 
        'download_name': get_netdisk_name(download_link, card_data.get('网盘名称', '百度网盘')),
        'download_link': download_link,
        'password': card_data.get('解压密码', 'FIXME'),
        'formats': card_data.get('支持格式', 'mp4'),
        'card_foot': card_data.get('卡片页脚', ''),
        'genre': card_data.get('类型', ''),
        'region': card_data.get('地区', '')
    }

    # 生成格式徽章（仍保留原有解析逻辑）
    context['format_badges'] = parse_formats(context['formats'])
    context['is_heji'] = 'true' if is_heji(card_data.get('主标题', '')) else 'false'

    # 加载 card 模板并渲染
    tpl_text = load_template_card()
    tpl = Template(tpl_text)
    return tpl.render(**context)



# Sheet 0 生成 index.html 
def generate_index(sheets, sheet_name, section_sheets=None):
    log.info(f"📄 开始转换 Sheet 0 -> index.html")

    is_ok, header, lines = load_sheet_table(sheets, sheet_name)
    if not is_ok:
        log.warn(f"⚠️  {sheet_name} 没有数据，跳过")
        return
    
    # 生成 cover 卡片 HTML
    covers = []
    for idx, data in enumerate(lines, 1):
        cover = generate_cover_html(data, idx)
        covers.append(cover)
    covers_html = '\n\n'.join(covers)
    
    # 生成热点推荐 HTML
    hot_html = ''
    if section_sheets:
        hot_html = generate_hot_series_html(sheets, section_sheets)
        if hot_html:
            log.info(f"✓ 生成热点推荐卡片")
    
    # 收集轮播图片
    carousel_slides = collect_carousel_images()
    
    # 准备模板上下文，包含全局配置
    context = {
        'covers': covers_html,
        'hot_sections_html': hot_html,
        'carousel_slides': carousel_slides,
        # 添加全局配置变量
        **GLOBAL_CONFIG
    }
    
    # 加载 index 模板并渲染
    tpl_text = load_template_index()
    tpl = Template(tpl_text)
    html_content = tpl.render(**context)
    
    # 保存文件
    with open(OUTPUT_INDEX, 'w', encoding=CHASET, newline='') as f:
        f.write(html_content)
    
    log.info(f"✓ 生成 {len(lines)} 个 section 卡片")
    log.info(f"✓ 已保存到: {OUTPUT_INDEX}")


def generate_cover_html(cover_data, idx):
    """生成 section cover 卡片 HTML（从模板渲染）"""
    # 准备字段
    tags_str = cover_data.get('标签列表', '')
    tags = re.split(r'[,，;；]', tags_str)
    tags = [f'<i class="bi bi-star"></i> {t.strip()}<br/>' for t in tags if t.strip()]
    tags_list = '\n                                  '.join(tags)
    
    context = {
        'section_id': idx,
        'section_link': f"./sections/section-{idx:02d}.html",
        'poster_path': cover_data.get('封面图片路径', f'./res/covers/{idx:02d}.webp'),
        'section_name': cover_data.get('主标题', ''),
        'description': cover_data.get('概要', ''),
        'count': cover_data.get('卡片数量', 0),
        'tags': tags_list,
    }
    
    # 加载 cover 模板并渲染
    tpl_text = load_template_cover()
    tpl = Template(tpl_text)
    return tpl.render(**context)



def load_sheet_table(sheets, sheet_name) :
    table = sheets[sheet_name]
    header = table['headers']
    lines = table['data']
    if not lines:
        log.warn(f"⚠️  {sheet_name} 没有数据，跳过")
        return False, [], []
    # log.debug(header)
    # log.debug(lines)
    return (True, header, lines)


def generate_hot_series_html(sheets, section_sheets):
    """生成热点推荐系列卡片 HTML"""
    import os as _os
    # 按条目标题查找（用于电视剧/电影热点）
    hot_items_tv = [s.strip() for s in GLOBAL_CONFIG.get('HOT_ITEMS_TV', '').split(',') if s.strip()]
    hot_items_movie = [s.strip() for s in GLOBAL_CONFIG.get('HOT_ITEMS_MOVIE', '').split(',') if s.strip()]
    # 按系列名查找（用于动漫热点）
    hot_series_anime = [s.strip() for s in GLOBAL_CONFIG.get('HOT_SERIES_ANIME', '').split(',') if s.strip()]
    
    items_config = [
        (0, hot_items_tv, '热点电视剧资源'),
        (1, hot_items_movie, '热点电影资源'),
    ]
    series_config = [
        (2, hot_series_anime, '热点动漫资源'),
    ]
    
    def make_card(cover_path, title, badge_text, link_url):
        if cover_path.startswith('../'):
            cover_path = '.' + cover_path[2:]
        return f'''            <div class="col-md-4 col-sm-6 mb-3">
                <a href="{link_url}" class="text-decoration-none d-block hot-card-link">
                    <div class="hot-card position-relative overflow-hidden rounded-4" style="min-height:220px;">
                        <img src="{cover_path}" alt="{title}" class="hot-card-img" loading="lazy" style="position:absolute;top:0;left:0;width:100%;height:100%;object-fit:cover;">
                        <div class="hot-card-overlay" style="position:absolute;top:0;left:0;width:100%;height:100%;background:linear-gradient(0deg,rgba(0,0,0,0.9) 0%,rgba(0,0,0,0.3) 50%,rgba(0,0,0,0.1) 100%);z-index:1;"></div>
                        <div class="hot-card-content position-absolute bottom-0 start-0 w-100 p-3" style="z-index:2;">
                            <h4 class="text-white fw-bold mb-1" style="text-shadow:0 2px 8px rgba(0,0,0,0.8);font-size:1.2rem;">{title}</h4>
                            <span class="badge bg-warning text-dark rounded-pill px-3 py-1">{badge_text}</span>
                        </div>
                    </div>
                </a>
            </div>'''
    
    def section_link(sec_idx):
        return f'./sections/section-{sec_idx+1:02d}.html'
    
    result = ''
    
    # 按条目标题匹配
    for sec_idx, title_list, sec_title in items_config:
        if not title_list or sec_idx >= len(section_sheets):
            continue
        sheet_name = section_sheets[sec_idx]
        section_data = sheets.get(sheet_name, {}).get('data', [])
        cards = []
        for target_title in title_list:
            found = None
            for item in section_data:
                item_title = str(item.get('主标题', '') or '')
                if target_title in item_title:
                    found = item
                    break
            if not found:
                log.warn(f"⚠️  热点条目 '{target_title}' 在 {sheet_name} 中未找到，跳过")
                continue
            cover = str(found.get('封面图片路径', '') or '')
            series_name = str(found.get('所属系列', '') or '')
            if not cover:
                for sib in section_data:
                    sib_cover = str(sib.get('封面图片路径', '') or '')
                    if str(sib.get('所属系列', '') or '') == series_name and sib_cover:
                        cover = sib_cover
            # 优先使用热点专用封面 (res/hot-covers/{target_title}.jpg)
            hot_overlay = _os.path.join(_os.path.dirname(_os.path.dirname(__file__)), 'res', 'hot-covers', f'{target_title}.jpg')
            if _os.path.exists(hot_overlay):
                cover = f'../res/hot-covers/{target_title}.jpg'
            link = section_link(sec_idx)
            cards.append(make_card(cover, target_title, series_name, link))
        
        if cards:
            result += render_hot_section(sec_idx, sec_title, cards)
    
    # 按系列名匹配（动漫）
    for sec_idx, series_list, sec_title in series_config:
        if not series_list or sec_idx >= len(section_sheets):
            continue
        sheet_name = section_sheets[sec_idx]
        section_data = sheets.get(sheet_name, {}).get('data', [])
        groups = {}
        for item in section_data:
            group = str(item.get('所属系列', '') or '')
            groups.setdefault(group, []).append(item)
        
        cards = []
        for series_name in series_list:
            if series_name not in groups:
                log.warn(f"⚠️  热点系列 '{series_name}' 在 {sheet_name} 中未找到，跳过")
                continue
            items = groups[series_name]
            cover = items[0].get('封面图片路径', '')
            hot_overlay = _os.path.join(_os.path.dirname(_os.path.dirname(__file__)), 'res', 'hot-covers', f'{series_name}.jpg')
            if _os.path.exists(hot_overlay):
                cover = f'../res/hot-covers/{series_name}.jpg'
            anchor = 'g-' + re.sub(r'[^a-zA-Z0-9\u4e00-\u9fff]+', '-', series_name).strip('-').lower()
            link = f'{section_link(sec_idx)}#{anchor}'
            cards.append(make_card(cover, series_name, f'{len(items)} 部', link))
        
        if cards:
            result += render_hot_section(sec_idx, sec_title, cards)
    
    return result


def render_hot_section(sec_idx, sec_title, cards):
    link = f'./sections/section-{sec_idx+1:02d}.html'
    return f'''
        <!-- {sec_title} -->
        <div class="hot-section mt-4 mb-4">
            <div class="d-flex align-items-center justify-content-between mb-3">
                <h3 class="fw-bold mb-0" style="font-size:1.5rem;background:linear-gradient(135deg,#ffd700,#ffaa00);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;">{sec_title}</h3>
                <a href="{link}" class="text-decoration-none" style="font-size:0.85rem;color:rgba(255,255,255,0.4);transition:color .2s;white-space:nowrap;">更多 ›</a>
            </div>
            <div class="row g-3">
{chr(10).join(cards)}
            </div>
        </div>'''


def collect_carousel_images():
    """返回固定 8 张轮播图片"""
    slides = []
    alt_labels = [
        '经典动漫合集封面',
        '经典日剧合集封面',
        '经典电影合集封面',
        '怀旧动漫番剧精选',
        '热门影视剧集推荐',
        '经典日本电影精选',
        '热血动漫番剧合集',
        '90后经典影视回忆'
    ]
    for i in range(1, 9):
        alt_text = alt_labels[i-1] if i <= len(alt_labels) else '影视合集封面'
        slides.append({'path': f'./res/covers/slide-{i:02d}.jpg', 'alt': alt_text, 'type': 'slide'})
    log.info(f"✓ 使用 {len(slides)} 张固定轮播图片")
    return slides


# ===========================
# 工具函数
# ===========================
def generate_uuid(length=8):
    """生成随机 UUID (8 字符)"""
    return ''.join(random.choices(string.ascii_letters + string.digits, k=length))


def get_badge_html(format_name):
    """生成格式徽章 HTML"""
    badge_colors = {
        'mp4': ('blue', 'mp4'),
        'mkv': ('purple', 'mkv'),
        'rmvb': ('orange', 'rmvb'),
        'tv': ('green', 'tv'),
        'avi': ('red', 'avi'),
        'flv': ('yellow', 'flv'),
    }
    
    color, label = badge_colors.get(format_name.lower(), ('gray', format_name))
    return f'<img src="https://img.shields.io/badge/{label}-Yes-{color}.svg" data-bs-toggle="tooltip" title="包含 {label} 视频格式">'


def parse_formats(format_str):
    """解析格式列表，返回 HTML"""
    if not format_str or format_str.strip() == '':
        return ''
    
    formats = re.split(r'[,，;；]', format_str)
    formats = [f.strip() for f in formats if f.strip()]
    return '\n                      '.join([get_badge_html(f) for f in formats])


def load_excel_data(excel_file):
    """加载 Excel 数据"""
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"找不到文件: {excel_file}")
    
    wb = load_workbook(excel_file)
    sheets = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = []
        
        # 获取标题行
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # 读取数据行
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            row_data = {}
            for col_idx, cell in enumerate(row):
                header = headers[col_idx]
                row_data[header] = cell.value if cell.value is not None else ''
            data.append(row_data)
        
        sheets[sheet_name] = {
            'headers': headers,
            'data': data
        }
    
    return sheets


def load_template_section():
    """加载 section 模板"""
    with open(TEMPLATE_SECTION, 'r', encoding=CHASET) as f:
        return f.read()


def load_template_card():
    """加载 card 模板"""
    with open(TEMPLATE_CARD, 'r', encoding=CHASET) as f:
        return f.read()


def load_template_index():
    """加载 index 模板"""
    with open(TEMPLATE_INDEX, 'r', encoding=CHASET) as f:
        return f.read()


def load_template_cover():
    """加载 cover 模板"""
    with open(TEMPLATE_COVER, 'r', encoding=CHASET) as f:
        return f.read()


# ===========================
# JS 配置更新函数
# ===========================
def update_js_config(sheets, section_sheets):
    """更新所有 JS 配置文件"""
    log.info("📝 更新 JS 配置文件...")
    
    # 从 index sheet 获取数据
    index_data = sheets.get('index', {}).get('data', [])
    
    # 更新各个 JS 文件
    update_seo_meta_js(index_data, section_sheets)
    update_count_badges_js(section_sheets)
    update_global_search_js(section_sheets)
    
    log.info("✓ JS 配置更新完成")


def update_seo_meta_js(index_data, section_sheets):
    """更新 seo-meta.js 中的影视系列列表和 SEO 配置"""
    seo_file = './docs/js/seo-meta.js'
    
    if not os.path.exists(seo_file):
        log.warn(f"⚠️ 找不到 {seo_file}，跳过")
        return
    
    with open(seo_file, 'r', encoding=CHASET) as f:
        content = f.read()
    
    # 更新 SEO 基础配置（使用 json.dumps 正确处理转义）
    import json
    seo_keys = ['SEO_TITLE', 'SEO_DESCRIPTION', 'SEO_KEYWORDS', 'SEO_AUTHOR',
                'SEO_SITE_NAME', 'SEO_DOMAIN', 'SEO_IMAGE', 'SEO_LOCALE']
    js_keys = ['title', 'description', 'keywords', 'author',
               'siteName', 'domain', 'image', 'locale']
    for seo_key, js_key in zip(seo_keys, js_keys):
        val = GLOBAL_CONFIG.get(seo_key, '')
        replacement = f"        {js_key}: {json.dumps(val, ensure_ascii=False)},"
        # 只匹配 seoConfig 块内的 JS 属性行：8空格 + key: + 空格 + 引号内容 + 引号逗号 + 可选注释
        pattern = r'^        ' + re.escape(js_key) + r":\s*(['\"]).*?\1,"
        content = re.sub(pattern, replacement, content, count=1, flags=re.MULTILINE)
    
    # 生成影视系列列表 JavaScript（使用 json.dumps 确保安全转义）
    import json
    series_list = []
    for idx, data in enumerate(index_data[:len(section_sheets)]):
        name = data.get('主标题', '')
        description = data.get('概要', '')
        if name:
            series_list.append(f"            {{ name: {json.dumps(name, ensure_ascii=False)}, description: {json.dumps(description, ensure_ascii=False)} }}")
    
    series_js = ',\n'.join(series_list)
    
    # 替换 movieSeries 配置
    pattern = r"movieSeries:\s*\[[\s\S]*?\]"
    replacement = f"movieSeries: [\n{series_js}\n        ]"
    
    content = re.sub(pattern, replacement, content)
    
    with open(seo_file, 'w', encoding=CHASET, newline='') as f:
        f.write(content)
    
    log.info("✓ 已更新 docs/js/seo-meta.js")


def update_count_badges_js(section_sheets):
    """更新 count-badges.js 中的 SECTIONS 配置"""
    count_file = './docs/js/count-badges.js'
    
    if not os.path.exists(count_file):
        log.warn(f"⚠️ 找不到 {count_file}，跳过")
        return
    
    with open(count_file, 'r', encoding=CHASET) as f:
        content = f.read()
    
    # 生成 SECTIONS 配置
    sections_list = []
    for idx, sheet_name in enumerate(section_sheets):
        file_name = f'section-{idx+1:02d}.html'
        sections_list.append(f"        {{ file: '{file_name}', index: {idx} }}")
    
    sections_js = ',\n'.join(sections_list)
    
    # 替换 SECTIONS 配置
    pattern = r"const SECTIONS = \[[\s\S]*?\];"
    replacement = f"const SECTIONS = [\n{sections_js}\n    ];"
    
    content = re.sub(pattern, replacement, content)
    
    with open(count_file, 'w', encoding=CHASET, newline='') as f:
        f.write(content)
    
    log.info("✓ 已更新 docs/js/count-badges.js")


def update_global_search_js(section_sheets):
    """更新 global-search.js 中的 SECTIONS 配置"""
    search_file = './docs/js/global-search.js'
    
    if not os.path.exists(search_file):
        log.warn(f"⚠️ 找不到 {search_file}，跳过")
        return
    
    with open(search_file, 'r', encoding=CHASET) as f:
        content = f.read()
    
    # 生成 SECTIONS 配置（需要从 Excel 读取 section 名称）
    # 使用 sheet 名称作为 section 名称
    sections_list = []
    for idx, sheet_name in enumerate(section_sheets):
        file_name = f'section-{idx+1:02d}.html'
        sections_list.append(f"        {{ file: '{file_name}', name: '{sheet_name}' }}")
    
    sections_js = ',\n'.join(sections_list)
    
    # 替换 SECTIONS 配置
    pattern = r"const SECTIONS = \[[\s\S]*?\];"
    replacement = f"const SECTIONS = [\n{sections_js}\n    ];"
    
    content = re.sub(pattern, replacement, content)
    
    with open(search_file, 'w', encoding=CHASET, newline='') as f:
        f.write(content)
    
    log.info("✓ 已更新 docs/js/global-search.js")


def generate_sitemap(sheets, section_sheets):
    """生成 sitemap.xml 文件（使用模板）"""
    # 获取域名配置
    domain = GLOBAL_CONFIG.get('SEO_DOMAIN', 'https://yourdomain.com/').rstrip('/')
    
    # 获取当前日期
    today = datetime.now().strftime('%Y-%m-%d')
    
    # 生成所有 URL 条目
    links = []
    
    # 添加首页
    link_context = {
        'loc': f'{domain}/',
        'lastmod': today,
        'changefreq': 'weekly',
        'priority': '1.0'
    }
    links.append(render_template(TEMPLATE_LINK, link_context))
    
    # 添加各个 section 页面，从 Excel 获取标题
    for idx, sheet_name in enumerate(section_sheets, 1):
        # 从 index sheet 获取对应 section 的信息
        index_data = sheets.get('index', {}).get('data', [])
        section_info = index_data[idx - 1] if idx <= len(index_data) else {}
        section_title = section_info.get('主标题', f'Section {idx}')
        
        link_context = {
            'loc': f'{domain}/sections/section-{idx:02d}.html',
            'title': section_title,
            'lastmod': today,
            'changefreq': 'monthly',
            'priority': '0.8'
        }
        links.append(render_template(TEMPLATE_LINK, link_context))
    
    item_links = '\n'.join(links)
    
    # 使用 sitemap 模板渲染
    context = {
        'item_links': item_links
    }
    
    sitemap_content = render_template(TEMPLATE_SITEMAP, context)
    
    # 保存 sitemap.xml
    with open(OUTPUT_SITEMAP, 'w', encoding=CHASET, newline='') as f:
        f.write(sitemap_content)
    
    log.info(f"✓ 已生成 {OUTPUT_SITEMAP} ({len(section_sheets)} 个 section)")


def render_template(template_path, context):
    """渲染模板"""
    with open(template_path, 'r', encoding=CHASET) as f:
        tpl_text = f.read()
    tpl = Template(tpl_text)
    return tpl.render(**context)


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        log.error("❌ 发生未知异常")
