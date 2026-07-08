import os, shutil
from openpyxl import load_workbook

# Build index of all cover files in 缺失封面 folders
missing_cover_dir = r"F:\1、自媒体\3、网站\影视\缺失封面"
category_map = {}  # filename -> (category, full_path)

for cat in ["电影", "电视剧", "动漫"]:
    cat_dir = os.path.join(missing_cover_dir, cat)
    if not os.path.isdir(cat_dir):
        continue
    for f in os.listdir(cat_dir):
        fp = os.path.join(cat_dir, f)
        if os.path.isfile(fp):
            category_map[f] = (cat, fp)

print(f"Found {len(category_map)} cover files in 缺失封面 folder")

# Scan all sheets for 合集 entries
wb = load_workbook("res/data_new.xlsx")

heji_entries = []  # (sheet_name, title, current_cover, row_obj, cover_col_idx, title_col_idx)

for sn in wb.sheetnames:
    ws = wb[sn]
    h = [c.value for c in ws[1]]
    title_idx = next((i for i, hv in enumerate(h) if hv and "主标题" in str(hv)), None)
    cover_idx = next((i for i, hv in enumerate(h) if hv and "封面图片路径" in str(hv)), None)
    if title_idx is None or cover_idx is None:
        continue
    
    # Detect if 合集 via CJK brackets in title
    for row in ws.iter_rows(min_row=2):
        title = str(row[title_idx].value) if row[title_idx].value else ""
        if not title:
            continue
        # Check if it's a 合集 (has 【】 brackets)
        if "【" in title and "】" in title:
            cover = str(row[cover_idx].value) if row[cover_idx].value else ""
            heji_entries.append((sn, title, cover, row, cover_idx, title_idx))

print(f"Found {len(heji_entries)} 合集 entries across all sheets\n")

# For each 合集, check if a dedicated cover exists in 缺失封面
updated = 0
for sn, title, current_cover, row, cover_idx, title_idx in heji_entries:
    # Try to find a matching file in 缺失封面
    # Match by exact title + .webp or .jpg
    matched_file = None
    for ext in [".webp", ".jpg", ".jpeg", ".png"]:
        candidate = title + ext
        if candidate in category_map:
            matched_file = candidate
            break
    
    # Also try fuzzy: look for filename that contains the title
    if not matched_file:
        for fname, (cat, fp) in category_map.items():
            # Remove extension for comparison
            base = os.path.splitext(fname)[0]
            if base == title:
                matched_file = fname
                break
    
    if matched_file:
        cat, src_path = category_map[matched_file]
        
        # Determine target path
        if sn == "电影资源":
            # Use res/film/ default, or match to existing series folder
            target_dir = "res/film/合集封面"
        elif sn == "电视剧资源":
            target_dir = "res/tv/合集封面"
        elif sn == "动漫资源":
            target_dir = "res/anime/合集封面"
        else:
            target_dir = "res/covers"
        
        os.makedirs(target_dir, exist_ok=True)
        dst = os.path.join(target_dir, matched_file)
        shutil.copy2(src_path, dst)
        
        # Build relative path from sections/
        rel = os.path.relpath(dst, "res").replace("\\", "/")
        new_cover = f"../res/{rel}"
        
        old_cover = current_cover or "(empty)"
        if old_cover != new_cover:
            row[cover_idx].value = new_cover
            print(f"  [{sn:10s}] UPDATED [{title:45s}]  {old_cover}  →  {new_cover}")
            updated += 1
        else:
            print(f"  [{sn:10s}] SAME     [{title:45s}]  {old_cover}")
    else:
        current = current_cover or "(empty)"
        print(f"  [{sn:10s}] NO COVER [{title:45s}]  current={current}")

wb.save("res/data_new.xlsx")
wb.close()
print(f"\nDone. Updated {updated} entries.")
