#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
预部署排序验证脚本 v3

检查项：
1. 合集子集系列一致性：合集与子集应在同一 所属系列
2. 同名跨系列检测
3. 必填字段完整性
4. extract_theme 遗漏检测：检查孤立主题（不应独立存在但被分出来的单条主题）
5. HTML 输出排序验证：每组合集是否在子集之前

用法：python py/verify_sort.py
"""

import os, sys, re
from openpyxl import load_workbook
from theme_utils import extract_theme, is_heji

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_EXCEL = os.path.join(PROJECT_ROOT, 'res', 'data_new.xlsx')
SECTIONS_DIR = os.path.join(PROJECT_ROOT, 'sections')

def normalize(t):
    return re.sub(r'[\s\u3000\u200e\u200f]+', '', str(t or '')).strip()

def extract_pwd_id(link):
    m = re.search(r'/s/([a-zA-Z0-9]+)', str(link or ''))
    return m.group(1) if m else None


def check_orphan_themes(rows):
    """Detect orphan themes:
    A singleton theme whose extract_theme() result matches an existing larger theme.
    E.g. '9号秘事 万圣节' (from '9号秘事 万圣节特别篇')
         → extract_theme('9号秘事 万圣节特别篇') = '9号秘事'
         → '9号秘事' exists as another theme with more members → ORPHAN
    """
    issues = []

    # Build theme groups using extract_theme on original titles
    theme_groups = {}
    for rd in rows:
        if not rd['title']:
            continue
        th = extract_theme(rd['title'])
        theme_groups.setdefault(th, []).append(rd)

    # For singleton themes, check if extract_theme shortens the title to an existing theme
    for th, members in list(theme_groups.items()):
        if len(members) != 1:
            continue
        title = members[0]['title']
        base = extract_theme(title)
        if base != th and base in theme_groups and base != th:
            issues.append({
                'type': 'error',
                'msg': f"孤立主题: [{th}] 应归入 [{base}], 标题: '{title}'",
                'hint': 'extract_theme 未能完全剥离后缀，该条目被分成了独立主题'
            })

    return issues


def check_html_sort_order(rows):
    """Read generated HTML output and verify合集 appears before子集 within each theme."""
    issues = []

    if not os.path.isdir(SECTIONS_DIR):
        return issues

    # Build theme groups
    theme_groups = {}
    for rd in rows:
        if not rd['title']:
            continue
        th = extract_theme(rd['title'])
        theme_groups.setdefault(th, []).append(rd)

    html_files = sorted(f for f in os.listdir(SECTIONS_DIR)
                        if f.startswith('section-') and f.endswith('.html'))

    for hf in html_files:
        html_path = os.path.join(SECTIONS_DIR, hf)
        with open(html_path, 'r', encoding='utf-8') as f:
            content = f.read()

        card_titles = re.findall(r'class="card-title[^"]*"[^>]*>([^<]+)</', content)

        title_pos = {}
        for idx, t in enumerate(card_titles):
            t = t.strip()
            if t not in title_pos:
                title_pos[t] = idx

        for th, members in theme_groups.items():
            hejis = [m for m in members if is_heji(m['title'])]
            subsets = [m for m in members if not is_heji(m['title'])]
            if not hejis or not subsets:
                continue

            heji_pos = [title_pos.get(m['title']) for m in hejis if m['title'] in title_pos]
            sub_pos = [title_pos.get(m['title']) for m in subsets if m['title'] in title_pos]
            heji_pos = [p for p in heji_pos if p is not None]
            sub_pos = [p for p in sub_pos if p is not None]

            if not heji_pos or not sub_pos:
                continue

            first_heji = min(heji_pos)
            first_sub = min(sub_pos)

            if first_heji > first_sub:
                bad_sub = None
                for m in subsets:
                    p = title_pos.get(m['title'])
                    if p is not None and p < first_heji:
                        bad_sub = m['title']
                        break

                issues.append({
                    'type': 'error',
                    'msg': f"[{hf}] 主题 '{th}': 子集 '{bad_sub}' (pos {first_sub}) 出现在合集 (pos {first_heji}) 之前",
                    'hint': 'extract_theme 可能未正确归组，或该子集与合集不在同一主题组'
                })

    return issues


def verify():
    wb = load_workbook(DATA_EXCEL, data_only=True)
    all_issues = []

    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = [c.value for c in ws[1]]

        if '主标题' not in headers or '所属系列' not in headers:
            continue

        ti = headers.index('主标题')
        si = headers.index('所属系列')
        li = headers.index('下载链接') if '下载链接' in headers else None

        rows = []
        for r in range(2, ws.max_row + 1):
            title = str(ws.cell(r, ti + 1).value or '').strip()
            series = str(ws.cell(r, si + 1).value or '').strip()
            link = str(ws.cell(r, li + 1).value or '').strip() if li else ''
            rows.append({'row': r, 'title': title, 'series': series,
                         'link': link, 'pwd': extract_pwd_id(link), 'sheet': sn})

        # 1. Identical titles in multiple series
        title_series = {}
        for rd in rows:
            if not rd['title']:
                continue
            key = normalize(rd['title'])
            title_series.setdefault(key, []).append(rd)

        for title, entries in title_series.items():
            series_set = set(e['series'] for e in entries)
            if len(series_set) > 1:
                locs = [(e['sheet'], e['row'], e['series'], e['pwd']) for e in entries]
                pwds = set(e['pwd'] for e in entries if e['pwd'])
                if len(pwds) > 1 and all(e['pwd'] for e in entries if e['pwd']):
                    all_issues.append({
                        'type': 'info',
                        'msg': f"同名跨系列(不同链接): [{title}] → {locs}",
                        'hint': '不同链接属于不同合集，正常'
                    })
                else:
                    all_issues.append({
                        'type': 'warn',
                        'msg': f"同名跨系列: [{title}] → {locs}",
                        'hint': '检查是否为意外重复'
                    })

        # 2. 合集子集 series consistency (using extract_theme for prefix matching)
        heji_rows = [rd for rd in rows if is_heji(rd['title'])]
        for heji in heji_rows:
            prefix = extract_theme(heji['title'])
            if len(prefix) < 2:
                continue
            subs = [rd for rd in rows if rd['title'] != heji['title']
                    and normalize(rd['title']).startswith(prefix)]
            for sub in subs:
                if sub['series'] != heji['series']:
                    if heji['pwd'] and sub['pwd'] and heji['pwd'] == sub['pwd']:
                        err_type = 'error'
                        hint = f"子集链接与合集相同，应移到 {heji['series']}"
                    elif heji['pwd'] and sub['pwd'] and heji['pwd'] != sub['pwd']:
                        err_type = 'warn'
                        hint = '子集链接不同，可能属于其他合集的跨系列同名条目'
                    else:
                        err_type = 'warn'
                        hint = f"建议确认 {sub['title']} 是否应属于 {heji['series']}"
                    all_issues.append({
                        'type': err_type,
                        'msg': f"合集子集跨系列: [{heji['title']}]({heji['series']}) vs [{sub['title']}]({sub['series']})",
                        'hint': hint
                    })

        # 3. Missing fields
        for rd in rows:
            if not rd['title']:
                all_issues.append({'type': 'error', 'msg': f"Row {rd['row']}: 主标题为空"})
            if not rd['series']:
                all_issues.append({'type': 'error', 'msg': f"Row {rd['row']}: {rd['title']} 所属系列为空"})

        # 4. Orphan theme detection
        all_issues.extend(check_orphan_themes(rows))

    wb.close()

    # 5. HTML sort order validation across all sheets
    wb2 = load_workbook(DATA_EXCEL, data_only=True)
    all_rows = []
    for sn in wb2.sheetnames:
        ws = wb2[sn]
        headers = [c.value for c in ws[1]]
        if '主标题' not in headers:
            continue
        ti = headers.index('主标题')
        for r in range(2, ws.max_row + 1):
            title = str(ws.cell(r, ti + 1).value or '').strip()
            if title:
                all_rows.append({'title': title})
    wb2.close()

    all_issues.extend(check_html_sort_order(all_rows))

    return all_issues


if __name__ == '__main__':
    print(f"验证数据文件: {DATA_EXCEL}\n")
    issues = verify()

    errors = [i for i in issues if i['type'] == 'error']
    warns = [i for i in issues if i['type'] == 'warn']
    infos = [i for i in issues if i['type'] == 'info']

    print(f"{'='*60}")
    print(f"  结果: {len(errors)} 错误, {len(warns)} 警告, {len(infos)} 提示")
    print(f"{'='*60}\n")

    for issue in errors:
        print(f"  [错误] {issue['msg']}")
        print(f"          建议: {issue.get('hint', '')}\n")

    for issue in warns:
        print(f"  [警告] {issue['msg']}")
        print(f"          提示: {issue.get('hint', '')}\n")

    for issue in infos:
        print(f"  [信息] {issue['msg']}")
        print(f"          提示: {issue.get('hint', '')}\n")

    if errors:
        sys.exit(1)
    else:
        print("数据验证通过")
