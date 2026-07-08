import os, sys, json, urllib.parse, time, re, traceback
sys.stdout.reconfigure(encoding='utf-8')
import requests
import openpyxl

QUARK_COOKIE = os.environ.get('QUARK_COOKIE', '')
if not QUARK_COOKIE:
    print("QUARK_COOKIE not set")
    sys.exit(1)

s = requests.Session()
for part in QUARK_COOKIE.split(';'):
    if '=' in part.strip():
        k, v = part.strip().split('=', 1)
        s.cookies.set(k, v)

def get_stoken(pwd_id, passcode=''):
    r = s.post('https://drive-h.quark.cn/1/clouddrive/share/sharepage/token?pr=ucpro&fr=pc',
               json={'pwd_id': pwd_id, 'passcode': passcode, 'support_visit_limit_private_share': True},
               headers=HEADERS, timeout=15)
    d = r.json()
    if d.get('status') == 200:
        return d['data']['stoken']
    return None

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Referer': 'https://pan.quark.cn/',
    'Content-Type': 'application/json',
}

def list_folder_recursive(pwd_id, stoken, pdir_fid='0', depth=0):
    """递归列出文件夹内容，返回 (dir_entries, file_entries) 扁平列表"""
    dirs = []
    files = []
    page = 1
    while True:
        url = ('https://drive-h.quark.cn/1/clouddrive/share/sharepage/detail'
               '?pr=ucpro&fr=pc&ver=2&pwd_id=' + pwd_id +
               '&stoken=' + urllib.parse.quote(stoken) +
               '&pdir_fid=' + str(pdir_fid) +
               '&force=0&_page=' + str(page) + '&_size=200&_fetch_total=1')
        r = s.get(url, headers=HEADERS, timeout=20)
        d = r.json()
        if d.get('status') != 200:
            break
        items = d.get('data', {}).get('list', [])
        if not items:
            break
        for item in items:
            entry = {
                'fid': item.get('fid', ''),
                'name': item.get('file_name', ''),
                'is_dir': item.get('dir', False),
                'depth': depth,
                'size': item.get('size', 0),
            }
            if entry['is_dir']:
                dirs.append(entry)
                # Recurse into subdirectories
                sub_dirs, sub_files = list_folder_recursive(pwd_id, stoken, entry['fid'], depth + 1)
                dirs.extend(sub_dirs)
                files.extend(sub_files)
            else:
                files.append(entry)
        total = d.get('data', {}).get('total', 0)
        if page * 200 >= total:
            break
        page += 1
    return dirs, files

def generate_dir_html(title, pwd_id, stoken):
    """Generate multi-level nested dir HTML"""
    all_dirs, all_files = list_folder_recursive(pwd_id, stoken)
    entries = all_dirs + all_files
    entries.sort(key=lambda x: (0 if x['is_dir'] else 1, x['name'].lower()))

    lines = []
    indent_size = 24
    for e in entries:
        indent_px = e['depth'] * indent_size
        icon = "📁" if e['is_dir'] else "📄"
        color = "#ffd700" if e['is_dir'] else "#aaa"
        style = f"margin-left:{indent_px}px;color:{color}"
        lines.append(f'<div style="{style}">{icon} {e["name"]}</div>')
    content = '\n'.join(lines)

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><title>{title} - 目录</title>
<style>
body {{ font-family: 'Microsoft YaHei', sans-serif; background: #1a1a2e; padding: 20px; font-size: 14px; line-height: 1.8; }}
a {{ color: #4fc3f7; text-decoration: none; }}
a:hover {{ text-decoration: underline; }}
</style>
</head><body>
{content}
</body></html>"""
    return html

def extract_pwd(url):
    """Extract pwd_id and passcode from quark share URL"""
    # https://pan.quark.cn/s/3677bbe64645
    # https://pan.quark.cn/s/66555a6d5d69?pwd=GaQ7
    m = re.search(r'/s/([a-f0-9]+)', url)
    pwd_id = m.group(1) if m else ''
    m2 = re.search(r'[?&]pwd=([^&]+)', url)
    passcode = m2.group(1) if m2 else ''
    return pwd_id, passcode

# Read Excel
wb = openpyxl.load_workbook(r'E:\workspace\github\film-tvs\res\data_new.xlsx')
sheets = [('电视剧资源', '电视剧资源'), ('电影资源', '电影资源'), ('动漫资源', '动漫资源')]

# Track all rows that need updating
updates = []  # (sheet_name, row, dir_key, passcode)

for sheet_name, label in sheets:
    ws = wb[sheet_name]
    for row in range(2, ws.max_row + 1):
        link = str(ws.cell(row, 10).value or '')
        if 'quark' in link.lower():
            pwd_id, passcode = extract_pwd(link)
            if pwd_id:
                updates.append((sheet_name, row, pwd_id, passcode, link))

print(f"Found {len(updates)} Quark resource entries")

# Process each
success = 0
fail = 0
for sheet_name, row, pwd_id, passcode, link in updates:
    title_cell = wb[sheet_name].cell(row, 2).value or ''
    print(f"  [{sheet_name}] row {row}: {title_cell} ({pwd_id})", end='')
    try:
        stoken = get_stoken(pwd_id, passcode)
        if not stoken:
            print(" - STOKEN FAIL")
            fail += 1
            continue
        dirs, files = list_folder_recursive(pwd_id, stoken)
        count = len(dirs) + len(files)
        print(f" - {count} items", end='')
        # Generate HTML
        html_title = str(title_cell)[:30]
        html = generate_dir_html(html_title, pwd_id, stoken)
        # Save to res/dirs/{pwd_id}.html
        dir_path = f'E:/workspace/github/film-tvs/res/dirs/quark_{pwd_id}.html'
        with open(dir_path, 'w', encoding='utf-8') as f:
            f.write(html)
        # Update Excel
        wb[sheet_name].cell(row, 11).value = f'quark_{pwd_id}'
        print(f" -> res/dirs/quark_{pwd_id}.html")
        success += 1
    except Exception as e:
        print(f" - ERROR: {e}")
        fail += 1

# Save Excel
wb.save(r'E:\workspace\github\film-tvs\res\data_new.xlsx')
print(f"\nDone: {success} success, {fail} failed")
