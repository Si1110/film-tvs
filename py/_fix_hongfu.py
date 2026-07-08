import sys, os, re, urllib.parse
sys.stdout.reconfigure(encoding='utf-8')
import requests
import openpyxl

COOKIE_STR = r"b-user-id=47000940-681b-3d39-3ad3-ffd51e0bef3a; _UP_A4A_11_=wba2d199ca6d4156bbf065be9709fec0; _c_WBKFRo=geWYEMxjjBz1P0w0V5kjZMCOIKZTvcVAWZmcfw6a; _UP_335_2B_=1; __sdid=AASjLyqoKdwFTpUnYB15zKlS3acNl/8PW/P0YAW4kN8dfIgFKJhVplRuuO3OK1LOpOiG8RyVyKLGVP5N19E2DjX5rJk/Y6tnlcNmc+YuFhk1lw==; _UP_D_=pc; _UP_30C_6A_=sta2e6201d15vu1n67adq3lv00werwuk; _UP_TS_=sg1ace6b77f3ab536d7cbec3937c8f84f8e; _UP_E37_B7_=sg1ace6b77f3ab536d7cbec3937c8f84f8e; _UP_TG_=sta2e6201d15vu1n67adq3lv00werwuk; tfstk=g26ivNjmF1Rs76gUEqJ_tqnJD7Fpfd9XRZHvkKL4Te8BklsA0KkVkip9gPsA-E8ebl6T5h8miaxNBV8aSwqcWiMv7FBvnI7JQI19kFLDni_uvze8eGs6hQz8y8BZ1lMJhEyv3WpU4gCt1ze8eGPwfKIUysK21uxv0K-wghuUxnKqgK-2ukYem3Gq7ZJ4xkx2cAk2_fyULhtM3E723kjeRnJ2bZJ4xM8B0jFb_F6VL9zSfXp9Rs1cKhAMzGVS3xXneCYPbeD4z9xghUSw-xk2PQmzrMjTSYspCt7k2NeEIajVVsRhn2zkl9jFQsSISkJlx6CJs64ZneB9mQJ2tmkVxpYBw9pzU2xCTGCVCNmm3h6OeIYWtokXMpWJa_7icuseQn7XNtUSCFSVVT1JER0Wj6SkKgRiTb-jIxtUD9ljGC-BxUnM7RQnxS6UdkqnaoOwAhT8xkcjGC-BxUE3xbyX_Ht6y; isg=BJ6eJiY2eP8mPqyBlRZBSGaP7zTgX2LZsPI_N0gnCuHcaz5FsO-y6cQIY3fn01rx; __pus=cfef8172d4b5c92d7d734315fd53ee47AAQxWjik0hG2opNiwmrnKLuL47LP6ewPS/HbECKHyA2QMswU/GcB81es/VV8rSN+K5ezgGipMCnRILw7RqLHx6Kx; __kp=571f8620-6c60-11f1-9146-55a822775a5f; __kps=AASj3UTlfAF8TNsc+2TeHr0r; __ktd=rdKtNeSsloEmLNyAxt/hVQ==; __uid=AASj3UTlfAF8TNsc+2TeHr0r; Video-Auth=Qsm/8v6XH0LY1llR1w94JDjXETtFT5Yb91LlXTFhu0U1g73EE5gIpCX9KFhEriYAxH/DNBCuvjKZ/9NWc6RaWqzIFGFEEipHavgsEa+AlJhMQXrLRBUD/4KffSw04p7xvD/9QepLa0cjitR4BBOviA==; __puus=21d2a6f8f99315e2a1ea9521fd809f8eAAQ271TZHk/+1L76Ex31OOfLj3KZKtUgJMlLiCpobHGg73k7AughQFofRl8LLS3amjWnzZz64CA22Zt6U38AIIBq0zEVGilSR1Kk31Ki894HTYwNcCkYgW4edu09CTf8m2CGtHSZbRXmUOhaFGoUnuP4qBw6r88eiXgCZvVPfv2vUuDdS7jylOPnOXcFeNcWLE7/W3u6bGRofRQTAhbEIvNh"

s = requests.Session()
for part in COOKIE_STR.split(';'):
    if '=' in part.strip():
        k, v = part.strip().split('=', 1)
        s.cookies.set(k, v)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Referer': 'https://pan.quark.cn/',
    'Content-Type': 'application/json',
}

pwd_id = '34c8cd1d9904'
print(f'Fetching {pwd_id}...')
r = s.post('https://drive-h.quark.cn/1/clouddrive/share/sharepage/token?pr=ucpro&fr=pc',
           json={'pwd_id': pwd_id, 'passcode': '', 'support_visit_limit_private_share': True},
           headers=HEADERS, timeout=10)
d = r.json()
if d.get('status') == 200:
    stoken = d['data']['stoken']
    url = f'https://drive-h.quark.cn/1/clouddrive/share/sharepage/detail?pr=ucpro&fr=pc&ver=2&pwd_id={pwd_id}&stoken={urllib.parse.quote(stoken)}&pdir_fid=0&force=0&_page=1&_size=200&_fetch_total=1'
    r2 = s.get(url, headers=HEADERS, timeout=15)
    d2 = r2.json()
    items = d2.get('data', {}).get('list', [])
    print(f'Items: {len(items)}')
    if items:
        # Save dir file
        lines = ['<div class="dir-list">', '  <div class="dir-section mb-3">',
                 '    <h6 style="color:#ffd700;border-bottom:1px solid rgba(255,215,0,0.2);padding-bottom:8px;margin-bottom:12px;">',
                 '      <i class="bi bi-folder2-open"></i> 夸克网盘目录', '    </h6>']
        for item in items:
            name = item.get('file_name', '?')
            icon = "📁" if item.get('dir') else "📄"
            color = "#ffd700" if item.get('dir') else "#aaa"
            lines.append(f'    <div style="margin-left:0px;color:{color};font-size:0.9rem;" class="mb-1">{icon} {name}</div>')
        lines.append('  </div>\n</div>')
        html = '\n'.join(lines)
        
        with open('res/dirs/quark_34c8cd1d9904.html', 'w', encoding='utf-8') as f:
            f.write(html)
        print('Saved res/dirs/quark_34c8cd1d9904.html')
        
        # Update Excel: change download link and dir path
        wb = openpyxl.load_workbook('res/data_new.xlsx')
        ws = wb['电影资源']
        for row in range(2, ws.max_row+1):
            title = str(ws.cell(row, 3).value or '')
            if '洪福齐天' in title:
                ws.cell(row, 10).value = 'https://pan.quark.cn/s/34c8cd1d9904'
                ws.cell(row, 8).value = 'quark_34c8cd1d9904'
                wb.save('res/data_new.xlsx')
                print(f'Updated Row {row}: {title} → quark_34c8cd1d9904')
                break
        wb.close()
else:
    print(f'API FAIL: status={d.get("status")}, msg={d.get("message","")}')
