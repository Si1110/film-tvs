import os, shutil, re
from openpyxl import load_workbook

# Build index of all cover files in 缺失封面
missing_dir = r"F:\1、自媒体\3、网站\影视\缺失封面"
all_files = []  # (category, filename, fullpath)
for cat in ["电影", "电视剧", "动漫"]:
    d = os.path.join(missing_dir, cat)
    if not os.path.isdir(d): continue
    for f in os.listdir(d):
        fp = os.path.join(d, f)
        if os.path.isfile(fp):
            all_files.append((cat, f, fp))

print(f"Total files in 缺失封面: {len(all_files)}")

# Heuristic: extract the core Chinese name from title (before 【)
def extract_core(title):
    m = re.match(r'^([^【]+)', title)
    return m.group(1).strip() if m else title.strip()

wb = load_workbook("res/data_new.xlsx")

for sn in wb.sheetnames:
    ws = wb[sn]
    h = [c.value for c in ws[1]]
    ti = next((i for i, hv in enumerate(h) if hv and "主标题" in str(hv)), None)
    ci = next((i for i, hv in enumerate(h) if hv and "封面图片路径" in str(hv)), None)
    if ti is None or ci is None: continue
    
    for row in ws.iter_rows(min_row=2):
        title = str(row[ti].value) if row[ti].value else ""
        if not title or "【" not in title or "】" not in title:
            continue
        
        core = extract_core(title)
        current = str(row[ci].value) if row[ci].value else ""
        
        # Check if current cover looks like a sub-item cover (not a dedicated合集 cover)
        # A dedicated cover usually has 【】 in the filename itself
        current_is_subitem = False
        if current and "【" not in current and "】" not in current:
            current_is_subitem = True
        
        # Try to find matching file in 缺失封面
        match = None
        # Strategy 1: exact match on full title
        for ext in [".webp", ".jpg", ".jpeg", ".png"]:
            cand = title + ext
            for cat, fname, fp in all_files:
                if fname == cand:
                    match = (cat, fname, fp)
                    break
            if match: break
        
        # Strategy 2: filename starts with core name and contains 【
        if not match:
            for cat, fname, fp in all_files:
                base = os.path.splitext(fname)[0]
                if base.startswith(core) and "【" in base and "】" in base:
                    match = (cat, fname, fp)
                    break
        
        # Strategy 3: core name is substring of filename and file has 【】
        if not match:
            for cat, fname, fp in all_files:
                base = os.path.splitext(fname)[0]
                if core in base and "【" in base and "】" in base:
                    match = (cat, fname, fp)
                    break
        
        if match:
            cat, fname, src_path = match
            # Determine target dir based on sheet
            target_map = {"电影资源": "res/film/合集封面", "电视剧资源": "res/tv/合集封面", "动漫资源": "res/anime/合集封面"}
            target_dir = target_map.get(sn, "res/covers")
            os.makedirs(target_dir, exist_ok=True)
            
            dst = os.path.join(target_dir, fname)
            shutil.copy2(src_path, dst)
            rel = os.path.relpath(dst, "res").replace("\\", "/")
            new_cover = f"../res/{rel}"
            
            if current != new_cover:
                row[ci].value = new_cover
                flag = "SUBITEM→DEDICATED" if current_is_subitem else "UPDATED"
                print(f"  [{sn:8s}] {flag:18s} [{title:40s}] → {new_cover}")
            else:
                print(f"  [{sn:8s}] ALREADY OK    [{title:40s}] {current}")
        else:
            print(f"  [{sn:8s}] NO MATCH      [{title:40s}] current=[{current[:60]}]")

wb.save("res/data_new.xlsx")
wb.close()
print("\nDone")
