#!/usr/bin/env python3
"""Import Baidu 0707 film/TV resources into res/data_new.xlsx.

Existing cards are updated with a second Baidu download button. New resources are
added as new rows with covers, Baidu directory HTML, language/subtitle metadata,
and short descriptions built from Douban public metadata.
"""

from __future__ import annotations

import html
import json
import os
import re
import shutil
import sys
import time
import urllib.parse
from pathlib import Path

import requests
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _batch_baidu_via_browser import call_api, generate_html  # noqa: E402
from batch_list import find_series_for_new  # noqa: E402
from theme_utils import extract_theme  # noqa: E402

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

PROJECT = Path(__file__).resolve().parents[1]
DATA_EXCEL = PROJECT / "res" / "data_new.xlsx"
SOURCE_EXCEL = Path(r"F:\1、自媒体\3、网站\影视\百度网盘资源（0707）.xlsx")
COVER_SOURCE_DIR = Path(r"F:\1、自媒体\3、网站\影视\影视封面\缺失封面\电视剧")
COVER_TARGET_DIR = PROJECT / "res" / "covers" / "百度网盘0707"
DIRS_DIR = PROJECT / "res" / "dirs"
DOUBAN_CACHE = PROJECT / "py" / ".douban_0707_cache.json"

CATEGORY_TO_SHEET = {
    "剧集": "电视剧资源",
    "电影": "电影资源",
    "动漫": "动漫资源",
}

EXTRA_HEADERS = ["备用网盘名称", "备用下载链接", "备用目录路径", "备用解压密码"]


def norm_link(value: str) -> str:
    link = str(value or "").strip()
    link = re.sub(r"([?&])pwd=[^&]+&?", r"\1", link).rstrip("?&")
    return link.rstrip("/")


def norm_title(value: str) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s+", "", text)
    text = text.replace("（", "(").replace("）", ")")
    return text


def safe_key(value: str, max_len: int = 80) -> str:
    text = re.sub(r"[\\/:*?\"<>|&\s#]+", "_", str(value or "")).strip("_ .")
    text = re.sub(r"_+", "_", text)
    return text[:max_len] or "baidu_resource"


def clean_display_title(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def strip_query_title(value: str) -> str:
    text = re.sub(r"[【\[（(].*?[】\]）)]", "", str(value or ""))
    text = re.sub(r"\s+第[一二三四五六七八九十0-9]+季", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text or clean_display_title(value)


def extract_shorturl(link: str) -> tuple[str, str | None]:
    m = re.search(r"pan\.baidu\.com/s/([^?&/\s]+)", link)
    pwd = re.search(r"[?&]pwd=([^&\s]+)", link)
    return (m.group(1) if m else "", pwd.group(1) if pwd else None)


def ensure_headers(ws) -> dict[str, int]:
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    for name in EXTRA_HEADERS:
        if name not in headers:
            col = ws.max_column + 1
            ws.cell(1, col).value = name
            headers[name] = col
    return headers


def load_cache() -> dict:
    if DOUBAN_CACHE.exists():
        return json.loads(DOUBAN_CACHE.read_text(encoding="utf-8"))
    return {}


def save_cache(cache: dict) -> None:
    DOUBAN_CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def get_douban_meta(title: str, category: str, cache: dict) -> dict:
    query = strip_query_title(title)
    if query in cache:
        return cache[query]

    meta: dict = {}
    headers = {"User-Agent": "Mozilla/5.0", "Referer": "https://movie.douban.com/"}
    try:
        suggest_url = "https://movie.douban.com/j/subject_suggest?q=" + urllib.parse.quote(query)
        r = requests.get(suggest_url, headers=headers, timeout=15)
        arr = r.json() if r.ok and r.text.startswith("[") else []
        if arr:
            item = arr[0]
            meta.update(
                {
                    "title": item.get("title") or query,
                    "sub_title": item.get("sub_title") or "",
                    "year": item.get("year") or "",
                    "episode": item.get("episode") or "",
                    "id": item.get("id") or "",
                    "img": item.get("img") or "",
                }
            )
            if meta.get("id"):
                api = f'https://m.douban.com/rexxar/api/v2/movie/{meta["id"]}'
                rr = requests.get(
                    api,
                    headers={"User-Agent": "Mozilla/5.0", "Referer": f'https://m.douban.com/movie/subject/{meta["id"]}/'},
                    timeout=15,
                )
                if rr.ok and rr.text.startswith("{"):
                    data = rr.json()
                    meta["intro"] = data.get("intro") or data.get("card_subtitle") or ""
                    meta["genres"] = [g.get("name") for g in data.get("genres", []) if g.get("name")]
                    meta["actors"] = [a.get("name") for a in data.get("actors", [])[:4] if a.get("name")]
                    meta["directors"] = [d.get("name") for d in data.get("directors", [])[:2] if d.get("name")]
    except Exception as exc:
        meta["error"] = str(exc)

    cache[query] = meta
    time.sleep(0.15)
    return meta


def build_description(title: str, source_genre: str, region: str, category: str, meta: dict) -> str:
    name = clean_display_title(title)
    year = meta.get("year") or ""
    genres = source_genre or "/".join(meta.get("genres") or [])
    actors = "、".join(meta.get("actors") or [])
    directors = "、".join(meta.get("directors") or [])
    intro = re.sub(r"\s+", "", meta.get("intro") or "")
    intro = re.sub(r"©豆瓣", "", intro)

    parts = []
    head = name
    if year:
        head += f"（{year}）"
    head += f"是一部{region or ''}{genres or category}作品"
    if directors:
        head += f"，由{directors}执导"
    if actors:
        head += f"，{actors}等出演"
    parts.append(head + "。")
    if intro:
        parts.append(intro[:130])
    else:
        parts.append(
            f"作品围绕核心人物的情感选择、身份处境与主要冲突展开，兼具{genres or '剧情'}元素，适合喜欢{region or '经典'}{category}的观众收藏观看。"
        )
    parts.append("本站整理百度网盘资源，便于快速查找剧集目录、清晰度版本与字幕信息。")
    desc = "".join(parts)
    return desc[:230]


def flatten_names(items: list[dict]) -> list[str]:
    names = []
    for item in items:
        name = item.get("server_filename") or ""
        if name:
            names.append(name)
        names.extend(flatten_names(item.get("_children") or []))
    return names


def names_from_existing_dir(dir_key: str) -> list[str]:
    path = DIRS_DIR / f"{dir_key}.html"
    if not path.exists():
        return []
    text = path.read_text(encoding="utf-8", errors="ignore")
    text = re.sub(r"<[^>]+>", " ", text)
    text = html.unescape(text)
    return [p.strip() for p in re.split(r"[\r\n]+", text) if p.strip()]


def infer_media_info(title: str, region: str, category: str, names: list[str]) -> tuple[str, str, str, str]:
    text = " ".join([title, region or "", *names]).lower()
    if "粤" in text or "cantonese" in text:
        lang = "粤语"
    elif "国语" in text or "普通话" in text or region == "大陆":
        lang = "汉语普通话"
    elif region == "韩国" or re.search(r"[\uac00-\ud7af]", text):
        lang = "韩语"
    elif region == "日本" or re.search(r"[\u3040-\u30ff]", text):
        lang = "日语"
    elif region == "泰国":
        lang = "泰语"
    elif region == "印度":
        lang = "印地语"
    else:
        lang = "英语" if region == "欧美" else "汉语普通话"

    sub = "中文字幕" if any(k in text for k in ["中字", "中文", "字幕", "简繁", "内封"]) else "中文字幕"
    exts = []
    for ext in ["mp4", "mkv", "iso", "avi", "ts", "mov"]:
        if re.search(rf"\.{ext}\b", text):
            exts.append(ext)
    formats = "/".join(exts[:4]) or "mp4/mkv"
    foot = ""
    episode = re.search(r"(?:全|ep|e)(\d{1,3})", text, re.I)
    if episode and category == "剧集":
        foot = f"约{episode.group(1)}集"
    return lang, sub, formats, foot


def copy_cover(title: str, meta: dict) -> str:
    COVER_TARGET_DIR.mkdir(parents=True, exist_ok=True)
    candidates = []
    title_norm = norm_title(title)
    for path in COVER_SOURCE_DIR.glob("*"):
        if not path.is_file():
            continue
        stem_norm = norm_title(path.stem)
        if stem_norm == title_norm or stem_norm in title_norm or title_norm in stem_norm:
            candidates.append(path)
    if candidates:
        src = sorted(candidates, key=lambda p: len(p.stem))[0]
        dest = COVER_TARGET_DIR / src.name
        if not dest.exists():
            shutil.copy2(src, dest)
        return "../" + dest.relative_to(PROJECT).as_posix()

    if meta.get("img"):
        try:
            r = requests.get(meta["img"].replace("/s_ratio_poster/", "/l_ratio_poster/"), headers={"User-Agent": "Mozilla/5.0"}, timeout=15)
            if r.ok and len(r.content) > 5000:
                dest = COVER_TARGET_DIR / f"{safe_key(title)}.jpg"
                dest.write_bytes(r.content)
                return "../" + dest.relative_to(PROJECT).as_posix()
        except Exception:
            pass
    return "../res/placeholder.webp"


def build_existing_indexes(wb) -> tuple[dict[str, list[dict]], set[str]]:
    by_sheet = {}
    all_links = set()
    for sn in [s for s in wb.sheetnames if s != "index"]:
        ws = wb[sn]
        headers = ensure_headers(ws)
        rows = []
        for r in range(2, ws.max_row + 1):
            row = {name: ws.cell(r, col).value for name, col in headers.items()}
            row["_row"] = r
            row["_norm_title"] = norm_title(row.get("主标题"))
            row["_theme"] = extract_theme(str(row.get("主标题") or ""))
            rows.append(row)
            all_links.add(norm_link(row.get("下载链接")))
            all_links.add(norm_link(row.get("备用下载链接")))
        by_sheet[sn] = {"ws": ws, "headers": headers, "rows": rows}
    return by_sheet, all_links


def link_fits_row(row: dict, link: str | None) -> bool:
    if not link:
        return True
    target = norm_link(link)
    main = norm_link(row.get("下载链接"))
    backup = norm_link(row.get("备用下载链接"))
    return not backup or backup == target or main == target


def find_existing(sheet_rows: list[dict], title: str, link: str | None = None) -> dict | None:
    nt = norm_title(title)
    for row in sheet_rows:
        if row["_norm_title"] == nt and link_fits_row(row, link):
            return row
    if "【" not in title:
        return None
    theme = extract_theme(title)
    exact_theme = [
        r
        for r in sheet_rows
        if r["_theme"] == theme
        and ("【" in title) == ("【" in str(r.get("主标题") or ""))
        and link_fits_row(r, link)
    ]
    if len(exact_theme) == 1:
        return exact_theme[0]
    return None


def row_matches_source(row: dict, source_title: str) -> bool:
    if row.get("_norm_title") == norm_title(source_title):
        return True
    row_title = str(row.get("主标题") or "")
    return "【" in source_title and "【" in row_title and row.get("_theme") == extract_theme(source_title)


def recompute_links(existing: dict[str, dict]) -> set[str]:
    links = set()
    for bundle in existing.values():
        for row in bundle["rows"]:
            links.add(norm_link(row.get("下载链接")))
            links.add(norm_link(row.get("备用下载链接")))
    return links


def clean_misplaced_backups(existing: dict[str, dict], source_by_link: dict[str, dict]) -> int:
    fixed = 0
    for bundle in existing.values():
        ws = bundle["ws"]
        headers = bundle["headers"]
        for row in bundle["rows"]:
            backup_link = norm_link(row.get("备用下载链接"))
            source = source_by_link.get(backup_link)
            if not source or row_matches_source(row, source["title"]):
                continue
            row_idx = row["_row"]
            for name in EXTRA_HEADERS:
                ws.cell(row_idx, headers[name]).value = ""
                row[name] = ""
            fixed += 1
    return fixed


def expand_description(title: str, source: dict, old_desc: str) -> str:
    region = source.get("region") or ""
    genre = source.get("genre") or "剧情"
    category = source.get("category") or "影视"
    title = clean_display_title(title or source.get("title") or "")
    old = re.sub(r"\s+", "", str(old_desc or ""))
    old = old.rstrip("。")
    if old and len(old) > 90:
        core = old[:130]
    elif old:
        core = old
    else:
        core = f"{title}以人物关系、关键事件和情绪冲突推进叙事，兼具{genre}等类型看点。"
    desc = (
        f"{title}是一部{region}{category}资源，题材涵盖{genre}。{core}。"
        f"本卡片整理百度网盘版本，结合网盘目录中的文件名、清晰度与字幕标记补充语言/字幕信息，方便快速判断是否适合收藏。"
        f"适合喜欢{region or '经典'}{category}、想集中查找高清版本和中文字幕资源的观众浏览下载。"
        f"简介已按公开资料和资源目录重新整理，尽量突出故事气质、类型标签与下载版本特点。"
    )
    return desc[:230]


def expand_source_descriptions(existing: dict[str, dict], source_by_link: dict[str, dict]) -> int:
    changed = 0
    for bundle in existing.values():
        ws = bundle["ws"]
        headers = bundle["headers"]
        for row in bundle["rows"]:
            source = None
            for key in ["下载链接", "备用下载链接"]:
                source = source_by_link.get(norm_link(row.get(key)))
                if source:
                    break
            if not source:
                continue
            current = str(row.get("概要") or "")
            if len(current) >= 170:
                continue
            desc = expand_description(str(row.get("主标题") or source["title"]), source, current)
            ws.cell(row["_row"], headers["概要"]).value = desc
            row["概要"] = desc
            changed += 1
    return changed


def main() -> None:
    if not SOURCE_EXCEL.exists():
        raise FileNotFoundError(SOURCE_EXCEL)

    cache = load_cache()
    source_wb = load_workbook(SOURCE_EXCEL, read_only=True, data_only=True)
    source_ws = source_wb.active
    source_rows = []
    for vals in source_ws.iter_rows(min_row=2, values_only=True):
        title, source, link, genre, region, category = vals[:6]
        if not title or not link or "pan.baidu.com" not in str(link):
            continue
        sheet = CATEGORY_TO_SHEET.get(str(category or "").strip())
        if not sheet:
            continue
        source_rows.append(
            {
                "title": clean_display_title(title),
                "source": source or "百度网盘",
                "link": str(link).strip(),
                "genre": str(genre or "").replace(" / ", "/"),
                "region": str(region or "").strip(),
                "category": str(category or "").strip(),
                "sheet": sheet,
            }
        )
    source_wb.close()
    print(f"Loaded {len(source_rows)} source rows", flush=True)

    wb = load_workbook(DATA_EXCEL)
    print("Workbook loaded", flush=True)
    existing, all_links = build_existing_indexes(wb)
    print("Existing indexes built", flush=True)
    source_by_link = {norm_link(item["link"]): item for item in source_rows}
    misplaced = clean_misplaced_backups(existing, source_by_link)
    if misplaced:
        all_links = recompute_links(existing)
        print(f"Cleaned misplaced backup links: {misplaced}", flush=True)

    updated = 0
    added = 0
    skipped = 0
    dir_ok = 0
    dir_fail = 0

    DIRS_DIR.mkdir(parents=True, exist_ok=True)
    for idx, item in enumerate(source_rows, 1):
        print(f"[{idx}/{len(source_rows)}] {item['title']}", flush=True)
        base_link = norm_link(item["link"])
        if base_link in all_links:
            skipped += 1
            continue

        shorturl, pwd = extract_shorturl(item["link"])
        dir_key = f"baidu_{shorturl}" if shorturl else ""
        items = []
        existing_names = names_from_existing_dir(dir_key) if dir_key else []
        if shorturl:
            if existing_names:
                dir_ok += 1
            else:
                data = call_api(shorturl, pwd)
                if data and data.get("errno") == 0:
                    # Root-level listing is enough for language/subtitle inference and
                    # avoids one huge share blocking the whole batch.
                    items = data.get("list") or []
                    dir_ok += 1
                else:
                    dir_fail += 1
                (DIRS_DIR / f"{dir_key}.html").write_text(
                    generate_html(shorturl, item["title"], items), encoding="utf-8"
                )

        names = existing_names or flatten_names(items)
        lang, sub, formats, foot = infer_media_info(item["title"], item["region"], item["category"], names)
        meta = get_douban_meta(item["title"], item["category"], cache)
        desc = build_description(item["title"], item["genre"], item["region"], item["category"], meta)
        cover_path = copy_cover(item["title"], meta)

        bundle = existing[item["sheet"]]
        ws = bundle["ws"]
        headers = bundle["headers"]
        rows = bundle["rows"]
        match = find_existing(rows, item["title"], item["link"])
        if match:
            row_idx = match["_row"]
            ws.cell(row_idx, headers["备用网盘名称"]).value = "百度网盘"
            ws.cell(row_idx, headers["备用下载链接"]).value = item["link"]
            ws.cell(row_idx, headers["备用目录路径"]).value = dir_key
            ws.cell(row_idx, headers["备用解压密码"]).value = ""
            if not ws.cell(row_idx, headers["类型"]).value and item["genre"]:
                ws.cell(row_idx, headers["类型"]).value = item["genre"]
            if "地区" in headers and not ws.cell(row_idx, headers["地区"]).value and item["region"]:
                ws.cell(row_idx, headers["地区"]).value = item["region"]
            updated += 1
        else:
            old_rows_for_match = [{k: v for k, v in r.items() if not k.startswith("_")} for r in rows]
            series = find_series_for_new(old_rows_for_match, item["title"])
            if not series:
                series = safe_key(extract_theme(item["title"]), 40).replace("_", "") + "系列"
            new_row = ws.max_row + 1
            ws.cell(new_row, headers["所属系列"]).value = series
            ws.cell(new_row, headers["封面图片路径"]).value = cover_path
            ws.cell(new_row, headers["主标题"]).value = item["title"]
            ws.cell(new_row, headers["副标题"]).value = ""
            ws.cell(new_row, headers["概要"]).value = desc
            ws.cell(new_row, headers["语言"]).value = lang
            ws.cell(new_row, headers["字幕"]).value = sub
            ws.cell(new_row, headers["目录路径"]).value = dir_key
            ws.cell(new_row, headers["网盘名称"]).value = "百度网盘"
            ws.cell(new_row, headers["下载链接"]).value = item["link"]
            ws.cell(new_row, headers["解压密码"]).value = ""
            ws.cell(new_row, headers["支持格式"]).value = formats
            ws.cell(new_row, headers["卡片页脚"]).value = foot
            ws.cell(new_row, headers["类型"]).value = item["genre"] or "/".join(meta.get("genres") or [])
            if "地区" in headers:
                ws.cell(new_row, headers["地区"]).value = item["region"]
            row_obj = {name: ws.cell(new_row, col).value for name, col in headers.items()}
            row_obj["_row"] = new_row
            row_obj["_norm_title"] = norm_title(item["title"])
            row_obj["_theme"] = extract_theme(item["title"])
            rows.append(row_obj)
            added += 1

        all_links.add(base_link)
        if idx % 25 == 0:
            print(f"[{idx}/{len(source_rows)}] updated={updated} added={added} dir_ok={dir_ok} dir_fail={dir_fail}")
            wb.save(DATA_EXCEL)
            save_cache(cache)

    desc_changed = expand_source_descriptions(existing, source_by_link)
    wb.save(DATA_EXCEL)
    save_cache(cache)
    print(
        f"Done. source={len(source_rows)} updated={updated} added={added} skipped={skipped} "
        f"dir_ok={dir_ok} dir_fail={dir_fail} desc_changed={desc_changed}",
        flush=True,
    )


if __name__ == "__main__":
    main()
