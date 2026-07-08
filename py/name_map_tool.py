"""
name_map_tool.py — 夸克脱敏名称映射管理工具

支持"大写字母=该字拼音首字母"的脱敏规则自动解码。

用法:
  python py/name_map_tool.py fetch <url>           # 获取目录结构并显示 fid
  python py/name_map_tool.py list                   # 列出所有已存映射
  python py/name_map_tool.py show <pwd_id>          # 查看某个链接的映射
  python py/name_map_tool.py map <pwd_id> <fid> <正确名称>  # 设置/修改单个映射
  python py/name_map_tool.py batch <pwd_id>         # 对某个链接应用全部映射并生成目录HTML
  python py/name_map_tool.py rebuild <pwd_id>       # 重新获取 + 应用映射 + 生成目录HTML
  python py/name_map_tool.py suggest <pwd_id>       # 显示脱敏名并提示每个字母可能的汉字
"""
import sys, os, json, urllib.parse, re, requests
sys.stdout.reconfigure(encoding='utf-8')

# Add parent to path
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from py.name_map import NameMapper

BASE = os.path.dirname(os.path.dirname(__file__))
DIRS_DIR = os.path.join(BASE, 'res', 'dirs')
EXCEL_PATH = os.path.join(BASE, 'res', 'data_new.xlsx')

# 拼音首字母 → 常见汉字映射（用于自动解码）
PINYIN_MAP = {
    'A': '啊阿爱安暗按案奥',
    'B': '八把白百办半帮包保报暴爆悲北备背本比笔必闭边编变便标表别冰并病不布步部',
    'C': '才材财采彩参残惨藏操草测层插查差拆柴产长常厂场唱超车彻沉晨成城程吃持尺充冲虫抽丑出初除础楚处川穿传船窗床创吹春词此刺从凑粗促村存寸错',
    'D': '大达打答代带待担单但蛋当党到道得的等底弟第点店电殿调顶定丢东冬动洞都斗毒独读度端短段断对队顿多夺',
    'E': '恶恩而儿童尔二',
    'F': '发法番翻反返饭犯范方房放飞非费分份丰风封疯佛夫服福府辅付负妇复父',
    'G': '该盖干敢感刚钢港高搞哥歌格个各根跟更工公功共够姑古故固顾挂怪关观官管馆光归规鬼贵国果过',
    'H': '哈海害含韩寒喊汉汗航毫好号喝合何和河黑很红洪后候呼忽湖互护户花华化画话还坏欢环换皇黄回会活火或惑',
    'J': '击机机积基激及即极急集几己计记纪技季既继寄加家价甲架假尖间件建健将讲奖降交教叫接阶节结解今金紧尽进近京经惊精警净静境镜究九久酒旧救就居举剧据聚觉决绝军',
    'K': '卡开看考靠科可刻客课肯空口哭苦库快款况困扩',
    'L': '拉拉来蓝览老乐了雷累类冷离里理力历立利例连联脸练炼恋凉两亮量疗料列烈林零灵领令流留龙楼露录旅律率乱落',
    'M': '马吗买满慢毛么没每美门们猛梦迷米密秘免眠面民明名命模魔末莫默某母目',
    'N': '拿那哪难男南脑内能你年念娘鸟牛农弄努女暖',
    'P': '怕排派旁跑泡陪配朋片票平评凭瓶颇破普',
    'Q': '七期其奇起气汽器千前钱强墙抢切亲青清情秋求区取去全权劝确',
    'R': '然让热人认任日容融肉如软若',
    'S': '三散色杀山伤商上烧少设社射申身深神审生声省失师十时识史使始世示事势是视试室室收手首受兽书术数树双水睡顺说思死四送速算虽随岁孙损所索锁',
    'T': '他她它太谈特提体天条听同通统痛头突图团推退托脱',
    'W': '外完玩晚碗万王往危为围唯维伟伪尾委未位温文闻问我屋无五午武舞务物误',
    'X': '西吸希息习席洗喜系下吓先显险现线限相香箱想向象像消销小效校笑些协写谢心新信兴星行形型性修秀须需许序续宣选学血寻训',
    'Y': '压牙雅烟言严研眼演验央羊阳养样要药也业叶夜一医依已以意因阴音银引印应英迎影映永用由油游有友又于余鱼雨语预原圆远院愿月云运',
    'Z': '在再早造责怎增展战站张章长招找照真阵争整正证政之支知执职只指至志制质治中终钟众重周州主住注抓转传装追准自字总走组嘴最尊左作坐座',
}

# 夸克API工具
def get_stoken(s, pwd_id, passcode=''):
    try:
        r = s.post('https://drive-h.quark.cn/1/clouddrive/share/sharepage/token?pr=ucpro&fr=pc',
                   json={'pwd_id':pwd_id,'passcode':passcode,'support_visit_limit_private_share':True},
                   headers={'User-Agent':'Mozilla/5.0','Content-Type':'application/json','Referer':'https://pan.quark.cn/'},
                   timeout=10)
        d = r.json()
        if d.get('status')==200: return d['data']['stoken']
        return None
    except: return None

def list_all(s, pwd_id, stoken, pdir_fid='0', depth=0):
    dirs, files = [], []
    page = 1
    while True:
        try:
            url = ('https://drive-h.quark.cn/1/clouddrive/share/sharepage/detail?pr=ucpro&fr=pc&ver=2'
                   '&pwd_id='+pwd_id+'&stoken='+urllib.parse.quote(stoken)+
                   '&pdir_fid='+str(pdir_fid)+'&force=0&_page='+str(page)+'&_size=200&_fetch_total=1')
            r = s.get(url, headers={'User-Agent':'Mozilla/5.0','Referer':'https://pan.quark.cn/'}, timeout=15)
            d = r.json()
            if d.get('status')!=200: break
            items = d.get('data',{}).get('list',[])
            if not items: break
            for item in items:
                entry = {'fid':item.get('fid',''),'name':item.get('file_name',''),
                         'is_dir':item.get('dir',False),'depth':depth}
                if entry['is_dir']:
                    dirs.append(entry)
                    sd, sf = list_all(s, pwd_id, stoken, entry['fid'], depth+1)
                    dirs.extend(sd); files.extend(sf)
                else:
                    files.append(entry)
            total = d.get('data',{}).get('total',0)
            if page*200 >= total: break
            page += 1
        except: break
    return dirs, files

def extract_pwd(url):
    m = re.search(r'/s/([a-f0-9]+)', url)
    pwd_id = m.group(1) if m else ''
    m2 = re.search(r'[?&]pwd=([^&]+)', url)
    passcode = m2.group(1) if m2 else ''
    return pwd_id, passcode

def get_cookie():
    return os.environ.get('QUARK_COOKIE', '')

def make_session():
    cookie = get_cookie()
    s = requests.Session()
    for part in cookie.split(';'):
        if '=' in part.strip():
            k,v=part.strip().split('=',1)
            s.cookies.set(k,v)
    return s

def generate_dir_html(title, pwd_id, stoken, nm):
    """获取目录，应用映射，生成HTML"""
    s = make_session()
    all_dirs, all_files = list_all(s, pwd_id, stoken)
    entries = all_dirs + all_files
    # 应用映射
    mapped = nm.apply(pwd_id, entries)
    # 排序
    entries.sort(key=lambda x: (0 if x['is_dir'] else 1, x['name'].lower()))
    lines = []
    for e in entries:
        indent_px = e['depth'] * 24
        icon = '📁' if e['is_dir'] else '📄'
        color = '#ffd700' if e['is_dir'] else '#aaa'
        lines.append(f'<div style="margin-left:{indent_px}px;color:{color}">{icon} {e["name"]}</div>')
    content = '\n'.join(lines)
    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><title>{title} - 目录</title>
<style>
body {{ font-family: 'Microsoft YaHei', sans-serif; background: #1a1a2e; padding: 20px; font-size: 14px; line-height: 1.8; }}
a {{ color: #4fc3f7; text-decoration: none; }}
a:hover {{ text-decoration: underline; }}
</style>
</head><body>
{content}
</body></html>'''
    return html, mapped

# ===== Commands =====

def cmd_fetch(url):
    pwd_id, passcode = extract_pwd(url)
    if not pwd_id:
        print('Invalid URL')
        return
    print(f'Fetching {pwd_id}...')
    s = make_session()
    stoken = get_stoken(s, pwd_id, passcode)
    if not stoken:
        print('STOKEN FAIL')
        return
    dirs, files = list_all(s, pwd_id, stoken)
    entries = dirs + files
    entries.sort(key=lambda x: (0 if x['is_dir'] else 1, x['name'].lower()))
    print(f'\n  Total: {len(dirs)} folders, {len(files)} files\n')
    print(f'  {"FID":<20} {"DP":>2} {"TYPE":<4} NAME')
    print(f'  {"-"*20} {"--":>2} {"----":<4} {"----"}')
    for e in entries:
        typ = '📁' if e['is_dir'] else '📄'
        print(f'  {e["fid"]:<20} {e["depth"]:>2} {typ:<4} {e["name"][:70]}')
    print(f'\n  To set mappings, use:')
    print(f'    python py/name_map_tool.py map {pwd_id} <fid> <正确名称>')

def cmd_list():
    nm = NameMapper()
    pwd_ids = nm.list_pwd_ids()
    if not pwd_ids:
        print('No mappings saved yet')
        return
    print(f'  Saved mappings:')
    for pk in pwd_ids:
        mapping = nm.get_all(pk)
        t = nm.get_title(pk)
        title = f' ({t})' if t else ''
        print(f'  {pk}{title}: {len(mapping)} 条映射')

def cmd_show(pwd_id):
    nm = NameMapper()
    mapping = nm.get_all(pwd_id)
    if not mapping:
        print(f'No mappings for {pwd_id}')
        return
    t = nm.get_title(pwd_id)
    print(f'Mappings for {pwd_id}{" ("+t+")" if t else ""}:')
    for fid, name in sorted(mapping.items()):
        print(f'  {fid[:20]:<22} → {name}')

def cmd_map(pwd_id, fid, *name_parts):
    nm = NameMapper()
    name = ' '.join(name_parts)
    nm.set(pwd_id, fid, name)
    nm.save()
    print(f'Saved: {fid[:20]} → {name}')

def cmd_batch(pwd_id):
    """Apply all mappings for a pwd_id and regenerate dir HTML"""
    nm = NameMapper()
    mapping = nm.get_all(pwd_id)
    if not mapping:
        print(f'No mappings for {pwd_id}. Add mappings first.')
        return
    
    # Find the link from Excel
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH)
    link = ''
    title = ''
    for sn in [n for n in wb.sheetnames if n != 'index']:
        ws = wb[sn]
        for r in range(2, ws.max_row+1):
            lk = str(ws.cell(r, 10).value or '')
            if pwd_id in lk:
                link = lk
                title = str(ws.cell(r, 3).value or '')
                # Save title to mapping
                nm.set_title(pwd_id, title)
                break
        if link: break
    
    if not link:
        print(f'Cannot find link for {pwd_id} in Excel.')
        # Use mapping title as fallback
        title = nm.get_title(pwd_id) or pwd_id
    
    _, passcode = extract_pwd(link)
    s = make_session()
    stoken = get_stoken(s, pwd_id, passcode)
    if not stoken:
        print('STOKEN FAIL')
        return
    
    html, mapped = generate_dir_html(title, pwd_id, stoken, nm)
    
    # Write dir file
    dir_path = os.path.join(DIRS_DIR, f'quark_{pwd_id}.html')
    with open(dir_path, 'w', encoding='utf-8') as f:
        f.write(html)
    nm.save()
    print(f'  ✅ Regenerated: {dir_path}')
    print(f'  ✅ Applied {mapped} name mappings')

def cmd_suggest(pwd_id):
    """Show garbled names with pinyin decode hints"""
    # Find link from Excel
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH)
    link = ''
    for sn in [n for n in wb.sheetnames if n != 'index']:
        ws = wb[sn]
        for r in range(2, ws.max_row+1):
            lk = str(ws.cell(r, 10).value or '')
            if pwd_id in lk:
                link = lk
                break
        if link: break

    if not link:
        print(f'Cannot find link for {pwd_id} in Excel')
        return

    _, passcode = extract_pwd(link)
    s = make_session()
    stoken = get_stoken(s, pwd_id, passcode)
    if not stoken:
        print('STOKEN FAIL')
        return

    dirs, files = list_all(s, pwd_id, stoken)
    entries = dirs + files
    entries.sort(key=lambda x: (0 if x['is_dir'] else 1, x['name'].lower()))

    nm = NameMapper()
    mapping = nm.get_all(pwd_id)
    total = len(entries)
    mapped_count = 0

    # Pattern: Chinese[-]Letter[-]Chinese (the garbled mark, optional hyphens)
    garbled_re = re.compile(r'[\u4e00-\u9fff]-?([A-Z])-?[\u4e00-\u9fff]')

    print(f'\n  {"="*70}')
    print(f'  📋 共 {total} 项 | 本地映射 {len(mapping)} 条')
    print(f'  {"="*70}')
    print(f'  {"#":>2} {"FID (32-char)":<34} {"DP":>2} NAME 及 解码建议')
    print(f'  {"-"*80}')

    idx = 0
    for e in entries:
        idx += 1
        fid = e['fid']
        name = e['name']
        dp = e['depth']
        icon = '📁' if e['is_dir'] else '📄'

        if fid in mapping:
            mapped_count += 1
            print(f'  {idx:>2} {fid:<34} {dp:>2} {icon} ✅ {name[:50]}')
            print(f'     {"":34} ↳ 已映射: "{mapping[fid]}"')
        else:
            # Check for garbled patterns
            matches = list(garbled_re.finditer(name))
            if matches:
                print(f'  {idx:>2} {fid:<34} {dp:>2} {icon} ⚠️ {name[:60]}')
                for m in matches:
                    letter = m.group(1)
                    chars = PINYIN_MAP.get(letter, '?')
                    print(f'     {"":34} ↳ 字母 "{letter}" 可能是: {chars[:30]}...')
            else:
                print(f'  {idx:>2} {fid:<34} {dp:>2} {icon}   {name[:60]}')

    print(f'  {"-"*80}')
    print(f'  API匹配: {mapped_count}/{total} | 本地映射表: {len(mapping)} 条')
    if mapped_count < total:
        print(f'\n  用法: python py/name_map_tool.py map {pwd_id} <fid> <正确名称>')

def cmd_rebuild(pwd_id):
    """Fetch fresh + apply mappings + regenerate"""
    cmd_suggest(pwd_id)
    print()
    cmd_batch(pwd_id)

# ===== Main =====

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return
    
    cmd = sys.argv[1]
    
    if cmd == 'fetch' and len(sys.argv) >= 3:
        cmd_fetch(sys.argv[2])
    elif cmd == 'list':
        cmd_list()
    elif cmd == 'show' and len(sys.argv) >= 3:
        cmd_show(sys.argv[2])
    elif cmd == 'map' and len(sys.argv) >= 5:
        cmd_map(sys.argv[2], sys.argv[3], *sys.argv[4:])
    elif cmd == 'batch' and len(sys.argv) >= 3:
        cmd_batch(sys.argv[2])
    elif cmd == 'suggest' and len(sys.argv) >= 3:
        cmd_suggest(sys.argv[2])
    elif cmd == 'rebuild' and len(sys.argv) >= 3:
        cmd_rebuild(sys.argv[2])
    else:
        print(f'Unknown command or missing args')
        print(__doc__)

if __name__ == '__main__':
    main()
