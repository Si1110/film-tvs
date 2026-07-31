#!/usr/bin/env python3
"""Import 0731 source resources into res/data_new.xlsx.

EXIST rows: add backup link (or replace main for 无耻之徒/西部世界).
NEW rows: create one card per title (main + backup), copy cover, placeholder dirs.
"""
# -*- coding: utf-8 -*-
from __future__ import annotations

import html
import json
import os
import re
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from batch_list import find_series_for_new  # noqa: E402
from theme_utils import extract_theme  # noqa: E402

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

PROJECT = Path(__file__).resolve().parents[1]
DATA_EXCEL = PROJECT / "res" / "data_new.xlsx"
SOURCE_EXCEL = Path(r"F:\1、自媒体\3、网站\影视\影视资源上架（0731）.xlsx")
COVER_SOURCE_DIRS = {
    "电视剧": Path(r"F:\1、自媒体\3、网站\影视\影视封面\缺失封面\电视剧"),
    "电影": Path(r"F:\1、自媒体\3、网站\影视\影视封面\缺失封面\电影"),
    "动漫": Path(r"F:\1、自媒体\3、网站\影视\影视封面\缺失封面\动漫"),
}
COVERS_DIR = PROJECT / "res" / "covers"
DIRS_DIR = PROJECT / "res" / "dirs"
DESC_FILE = PROJECT / "tmp" / "desc_0731.json"

CATEGORY_TO_SHEET = {
    "电视剧": "电视剧资源",
    "电影": "电影资源",
    "动漫": "动漫资源",
}

EXTRA_HEADERS = ["备用网盘名称", "备用下载链接", "备用目录路径", "备用解压密码"]

# Titles whose main+backup are both filled -> replace main, keep backup
REPLACE_MAIN = {"无耻之徒【11季合集】", "西部世界【4季合集】"}


def norm(s):
    s = re.sub(r"[（(]\d{4}[）)]", "", s)
    s = re.sub(r"[\s\u3000\u200b]+", "", s or "")
    s = s.replace("【", "").replace("】", "")
    return s.lower()


def safe_key(value, max_len=80):
    text = re.sub(r'[\\/:*?"<>|&\s#]+', "_", str(value or "")).strip("_ .")
    text = re.sub(r"_+", "_", text)
    return text[:max_len] or "resource"


def dir_key_for(link):
    link = str(link or "").strip()
    if "quark.cn" in link:
        m = re.search(r"pan\.quark\.cn/s/([a-zA-Z0-9]+)", link)
        return f"quark_{m.group(1)}" if m else ""
    if "baidu.com" in link:
        m = re.search(r"pan\.baidu\.com/s/([^?&\s/]+)", link)
        return f"baidu_{m.group(1)}" if m else ""
    return ""


def pwd_for(link):
    m = re.search(r"[?&]pwd=([^&\s]+)", str(link or ""))
    return m.group(1) if m else ""


def net_name_for(link):
    link = str(link or "")
    if "quark.cn" in link:
        return "夸克网盘"
    if "baidu.com" in link:
        return "百度网盘"
    return ""


def placeholder_dir_html(link, title):
    name = title or "目录"
    if "quark.cn" in str(link):
        label, btn = "该资源目录请前往夸克网盘查看", str(link)
    else:
        label, btn = "该资源目录请前往百度网盘查看", str(link)
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><title>{html.escape(name)} - 目录</title>
<style>
body {{ font-family: 'Microsoft YaHei', sans-serif; background: #1a1a2e; padding: 20px; font-size: 14px; line-height: 1.8; }}
a {{ color: #4fc3f7; text-decoration: none; }}
a:hover {{ text-decoration: underline; }}
.placeholder {{ color: #888; text-align: center; margin-top: 40px; }}
</style>
</head><body>
<div class="placeholder">
  <div style="font-size: 18px; color: #ffd700;">📁 {html.escape(name)}</div>
  <div style="margin-top: 10px; color: #666;">{label}</div>
  <div style="margin-top: 15px;">
    <a href="{html.escape(btn)}" target="_blank" style="display: inline-block; padding: 10px 24px; background: #4fc3f7; color: #fff; border-radius: 6px; text-decoration: none;">打开网盘</a>
  </div>
</div>
</body></html>"""


def ensure_headers(ws):
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    for name in EXTRA_HEADERS:
        if name not in headers:
            col = ws.max_column + 1
            ws.cell(1, col).value = name
            headers[name] = col
    return headers


def normalize_region(region):
    m = {
        "大陆": "大陆",
        "香港": "香港",
        "台湾": "台湾",
        "欧美": "欧美",
        "韩国": "韩国",
        "日本": "日本",
        "泰国": "泰国",
        "印度": "印度",
    }
    return m.get(str(region or "").strip(), str(region or "").strip())


def infer_lang(region, title):
    text = str(title or "")
    if region == "大陆" or region == "台湾":
        return "汉语普通话"
    if region == "香港":
        return "粤语" if re.search(r"[\u4e00-\u9fff]", text) else "粤语"
    if region == "韩国" or re.search(r"[\uac00-\ud7af]", text):
        return "韩语"
    if region == "日本" or re.search(r"[\u3040-\u30ff]", text):
        return "日语"
    if region == "泰国":
        return "泰语"
    if region == "印度":
        return "印地语"
    return "英语"


def copy_cover(title, category, series):
    src_dir = COVER_SOURCE_DIRS.get(category)
    if not src_dir or not src_dir.is_dir():
        return None
    src = src_dir / f"{title}.webp"
    if not src.exists():
        return None
    series_dir = COVERS_DIR / series
    series_dir.mkdir(parents=True, exist_ok=True)
    dest = series_dir / f"{title}.webp"
    if not dest.exists():
        shutil.copy2(src, dest)
    return f"../res/covers/{series}/{title}.webp"


def main():
    if not SOURCE_EXCEL.exists():
        raise FileNotFoundError(SOURCE_EXCEL)

    desc_map = {}
    if DESC_FILE.exists():
        desc_map = json.loads(DESC_FILE.read_text(encoding="utf-8"))
    print(f"Desc entries: {len(desc_map)}", flush=True)

    wb = load_workbook(DATA_EXCEL)
    index = {}
    for sn in [s for s in wb.sheetnames if s != "index"]:
        ws = wb[sn]
        headers = ensure_headers(ws)
        rows = []
        for r in range(2, ws.max_row + 1):
            rows.append({"row": r, "norm": norm(str(ws.cell(r, headers["主标题"]).value or ""))})
        index[sn] = {"ws": ws, "headers": headers, "rows": rows}

    swb = load_workbook(SOURCE_EXCEL, read_only=True, data_only=True)
    sws = swb["Sheet1"]
    source_rows = []
    for r in range(2, sws.max_row + 1):
        name = str(sws.cell(r, 1).value or "").strip()
        link = str(sws.cell(r, 3).value or "").strip()
        genre = str(sws.cell(r, 4).value or "").strip()
        region = str(sws.cell(r, 5).value or "").strip()
        cat = str(sws.cell(r, 6).value or "").strip()
        if not name or not link:
            continue
        sheet = CATEGORY_TO_SHEET.get(cat)
        if not sheet:
            continue
        source_rows.append(
            {
                "src_row": r,
                "title": name,
                "link": link,
                "genre": re.sub(r"\s*/\s*", "/", genre),
                "region": normalize_region(region),
                "category": cat,
                "sheet": sheet,
                "dir_key": dir_key_for(link),
                "pwd": pwd_for(link),
                "net": net_name_for(link),
            }
        )
    swb.close()
    print(f"Source rows: {len(source_rows)}", flush=True)

    # Group by (sheet, norm title)
    from collections import OrderedDict

    groups = OrderedDict()
    for item in source_rows:
        key = (item["sheet"], norm(item["title"]))
        groups.setdefault(key, []).append(item)

    updated = 0
    added = 0
    no_desc = []
    for (sheet, ntitle), items in groups.items():
        bundle = index[sheet]
        ws, headers = bundle["ws"], bundle["headers"]
        # Find existing row by norm title
        existing_row = None
        for r in bundle["rows"]:
            if r["norm"] == ntitle:
                existing_row = r["row"]
                break

        title = items[0]["title"]
        if existing_row:
            is_replace = title in REPLACE_MAIN
            if is_replace:
                # Replace main link with first source link, keep backup
                item = items[0]
                ws.cell(existing_row, headers["网盘名称"]).value = item["net"]
                ws.cell(existing_row, headers["下载链接"]).value = item["link"]
                ws.cell(existing_row, headers["目录路径"]).value = item["dir_key"]
                ws.cell(existing_row, headers["解压密码"]).value = item["pwd"]
                DIRS_DIR.mkdir(parents=True, exist_ok=True)
                (DIRS_DIR / f"{item['dir_key']}.html").write_text(
                    placeholder_dir_html(item["link"], title), encoding="utf-8"
                )
                print(f"[REPLACE-MAIN] {sheet}/{title} -> {item['link'][:45]}", flush=True)
                updated += 1
            else:
                # Add backup link
                item = items[0]
                bak_link = item["link"]
                ws.cell(existing_row, headers["备用网盘名称"]).value = item["net"]
                ws.cell(existing_row, headers["备用下载链接"]).value = bak_link
                ws.cell(existing_row, headers["备用目录路径"]).value = item["dir_key"]
                ws.cell(existing_row, headers["备用解压密码"]).value = ""
                DIRS_DIR.mkdir(parents=True, exist_ok=True)
                (DIRS_DIR / f"{item['dir_key']}.html").write_text(
                    placeholder_dir_html(bak_link, title), encoding="utf-8"
                )
                print(f"[ADD-BACKUP] {sheet}/{title} <- {bak_link[:45]}", flush=True)
                updated += 1
            continue

        # NEW card
        main = items[0]
        backup = items[1] if len(items) > 1 else None

        # Determine series
        old_rows = [
            {
                "主标题": str(ws.cell(r["row"], headers["主标题"]).value or ""),
                "所属系列": str(ws.cell(r["row"], headers["所属系列"]).value or ""),
            }
            for r in bundle["rows"]
        ]
        series = find_series_for_new(old_rows, title)
        if not series:
            series = safe_key(extract_theme(title), 40).replace("_", "") + "系列"

        desc = desc_map.get(title, "")
        if not desc:
            no_desc.append(title)
            desc = f"{title}是一部{main['genre'] or '剧情'}题材作品，高清全集资源提供下载收藏。"
        desc = desc[:230]

        lang = infer_lang(main["region"], title)
        cover = copy_cover(title, main["category"], series)
        if not cover:
            cover = "../res/placeholder.webp"

        DIRS_DIR.mkdir(parents=True, exist_ok=True)
        main_key = main["dir_key"]
        (DIRS_DIR / f"{main_key}.html").write_text(
            placeholder_dir_html(main["link"], title), encoding="utf-8"
        )
        bak_key = ""
        if backup:
            bak_key = backup["dir_key"]
            (DIRS_DIR / f"{bak_key}.html").write_text(
                placeholder_dir_html(backup["link"], title), encoding="utf-8"
            )

        new_row = ws.max_row + 1
        ws.cell(new_row, headers["所属系列"]).value = series
        ws.cell(new_row, headers["封面图片路径"]).value = cover
        ws.cell(new_row, headers["主标题"]).value = title
        ws.cell(new_row, headers["副标题"]).value = ""
        ws.cell(new_row, headers["概要"]).value = desc
        ws.cell(new_row, headers["语言"]).value = lang
        ws.cell(new_row, headers["字幕"]).value = "中文字幕"
        ws.cell(new_row, headers["目录路径"]).value = main_key
        ws.cell(new_row, headers["网盘名称"]).value = main["net"]
        ws.cell(new_row, headers["下载链接"]).value = main["link"]
        ws.cell(new_row, headers["解压密码"]).value = main["pwd"]
        ws.cell(new_row, headers["支持格式"]).value = "mp4/mkv"
        ws.cell(new_row, headers["卡片页脚"]).value = ""
        ws.cell(new_row, headers["类型"]).value = main["genre"]
        if "地区" in headers:
            ws.cell(new_row, headers["地区"]).value = main["region"]
        if "备用网盘名称" in headers:
            ws.cell(new_row, headers["备用网盘名称"]).value = backup["net"] if backup else ""
            ws.cell(new_row, headers["备用下载链接"]).value = backup["link"] if backup else ""
            ws.cell(new_row, headers["备用目录路径"]).value = bak_key
            ws.cell(new_row, headers["备用解压密码"]).value = backup["pwd"] if backup else ""
        bundle["rows"].append({"row": new_row, "norm": ntitle})
        print(f"[ADD] {sheet}/{title} | {series} | {main['net']}" + (f" + {backup['net']}" if backup else ""), flush=True)
        added += 1

    wb.save(DATA_EXCEL)
    print(f"\nDone. updated={updated} added={added}", flush=True)
    if no_desc:
        print(f"Missing desc for {len(no_desc)} titles:", flush=True)
        for t in no_desc:
            print("  ", t, flush=True)


if __name__ == "__main__":
    main()
