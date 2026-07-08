#!/usr/bin/env python3
"""批量修复全站问题：字幕为空 + 目录路径指向baidu-menus路径"""
import os, re, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

PROJECT = r'E:\workspace\github\film-tvs'
EXCEL = os.path.join(PROJECT, 'res', 'data_new.xlsx')
DIRS_DIR = os.path.join(PROJECT, 'res', 'dirs')
BAIDU_MENUS_TV = os.path.join(PROJECT, 'res', 'baidu-menus', 'tv')
BAIDU_MENUS_MOVIE = os.path.join(PROJECT, 'res', 'baidu-menus', 'movie')

# ====== Helper: parse baidu menu txt to dir HTML ======
def txt_to_dir_html(txt_path, title='目录', source='百度网盘'):
    lines = open(txt_path, 'r', encoding='utf-8').read().strip().split('\n')
    folders = []
    files = []
    in_folder = False
    in_file = False
    for line in lines:
        line = line.strip().rstrip('/')
        if not line or line.startswith('---') or line.startswith('  '):
            continue
        if line.startswith('📁') or '文件夹' in line:
            in_folder = True
            in_file = False
            continue
        if line.startswith('📄') or '文件' in line:
            in_folder = False
            in_file = True
            continue
        if in_folder and line:
            folders.append(line.lstrip('  ').lstrip('📁').strip())
        if in_file and line:
            files.append(line.lstrip('  ').lstrip('📄').strip())
    
    html = '<div class="dir-list">\n  <div class="dir-section mb-3">\n'
    html += f'    <h6 style="color:#ffd700;border-bottom:1px solid rgba(255,215,0,0.2);padding-bottom:8px;margin-bottom:12px;">\n'
    html += f'      <i class="bi bi-folder2-open"></i> {source}目录\n    </h6>\n'
    html += '    <div class="mb-2">\n'
    for f in folders:
        html += f'      <div class="d-flex align-items-center mb-1">\n        <span style="color:#ffd700;">&#x1f4c1; {f}</span>\n      </div>\n'
    for f in files:
        html += f'      <div class="mb-1" style="color:#aaa;font-size:0.9rem;">&#x1f4c4; {f}</div>\n'
    html += '    </div>\n  </div>\n</div>'
    return html

# ====== Step 1: Fix empty subtitles ======
print('=' * 60)
print('Step 1: 修复字幕字段')
print('=' * 60)
wb = load_workbook(EXCEL)
sub_fixed = 0
for sn in wb.sheetnames:
    if sn == 'index': continue
    ws = wb[sn]
    for r in range(2, ws.max_row + 1):
        sub = str(ws.cell(r, 7).value or '').strip()
        link = str(ws.cell(r, 10).value or '').strip()
        title = str(ws.cell(r, 3).value or '').strip()
        if not sub and link:
            ws.cell(r, 7).value = '中文字幕'
            sub_fixed += 1
            print(f'  {sn}行{r} [{title}] 字幕空 → 中文字幕')
print(f'✓ 已修复 {sub_fixed} 条字幕')

# ====== Step 2: Collect unique baidu menu paths and generate dir HTMLs ======
print()
print('=' * 60)
print('Step 2: 转换百度菜单 → dirs HTML')
print('=' * 60)

# Collect all broken dir entries grouped by their menu file path
menu_to_entries = {}
for sn in wb.sheetnames:
    if sn == 'index': continue
    ws = wb[sn]
    for r in range(2, ws.max_row + 1):
        dp = str(ws.cell(r, 8).value or '').strip()
        title = str(ws.cell(r, 3).value or '').strip()
        series = str(ws.cell(r, 1).value or '').strip()
        if dp and (dp.startswith('../') or 'baidu-menus' in dp or 'res/tv' in dp or 'res/film' in dp):
            # Extract the actual file path from the relative path
            # It could be: ../res/baidu-menus/tv/xxx_menu.txt
            # Or: ../res/tv/世にも奇妙な物語/...
            menu_to_entries.setdefault(dp, []).append({'sn': sn, 'r': r, 'title': title, 'series': series})

# Map baidu menu paths to dir keys
# Generate HTML files
fixed_dir = 0
for menu_path, entries in sorted(menu_to_entries.items()):
    # Determine the actual file path
    # The menu_path is stored as ../res/baidu-menus/tv/xxx or ../res/tv/世にも奇妙な物語/xxx
    rel_path = menu_path.replace('../', '', 1) if menu_path.startswith('../') else menu_path
    abs_path = os.path.join(PROJECT, rel_path)
    
    # Determine dir key - use the first entry's clean title or series
    entry = entries[0]
    title = entry['title']
    series = entry['series']
    sn = entry['sn']
    
    # Create a clean key from the title
    clean = re.sub(r'[【\[\(（\].*?[】\]\)）]', '', title).strip()
    clean = re.sub(r'[\\/:*?"<>|&\s#]', '_', clean).strip('_ ')
    clean = re.sub(r'_{2,}', '_', clean)[:40]
    
    # If it's a series where multiple items share same menu, use a group key
    # For 世奇: each SP has its own menu file, use the title as key
    # For 国产精品 items: use the title as key
    
    # Read existing baidu menu file if it exists
    dir_key = clean
    dir_html = None
    
    if os.path.exists(abs_path):
        source = '百度网盘'
        if 'movie' in rel_path:
            source = '百度网盘'
        elif 'tv' in rel_path:
            source = '百度网盘'
        try:
            dir_html = txt_to_dir_html(abs_path, title, source)
        except Exception as e:
            print(f'  ⚠️ 读取菜单文件失败 {abs_path}: {e}')
    elif os.path.exists(abs_path + '.txt'):
        try:
            dir_html = txt_to_dir_html(abs_path + '.txt', title)
        except:
            pass
    
    if dir_html is None:
        # Generate a simple fallback HTML
        dir_html = '<div class="dir-list">\n  <div class="dir-section mb-3">\n'
        dir_html += f'    <h6 style="color:#ffd700;border-bottom:1px solid rgba(255,215,0,0.2);padding-bottom:8px;margin-bottom:12px;">\n'
        dir_html += f'      <i class="bi bi-folder2-open"></i> 百度网盘目录\n    </h6>\n'
        dir_html += f'    <div class="mb-2">\n      <div class="mb-1" style="color:#aaa;">{title}</div>\n    </div>\n  </div>\n</div>'
    
    # Save HTML file
    os.makedirs(DIRS_DIR, exist_ok=True)
    html_path = os.path.join(DIRS_DIR, f'{dir_key}.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(dir_html)
    
    # Update all entries pointing to this menu path
    for e in entries:
        ws = wb[e['sn']]
        old = str(ws.cell(e['r'], 8).value or '')
        ws.cell(e['r'], 8).value = dir_key
        fixed_dir += 1
        if len(entries) <= 3:
            print(f'  {e["sn"]}行{e["r"]} [{e["title"]}] {old[:40]} → {dir_key}')
    
    if len(entries) > 3:
        print(f'  [{series}] {title} 等 {len(entries)} 条 → {dir_key}')

# Save
wb.save(EXCEL)
print(f'✓ 已修复 {fixed_dir} 条目录路径，生成 HTML 文件到 res/dirs/')

# ====== Summary ======
print()
print('=' * 60)
print('修复完成统计')
print('=' * 60)
print(f'  字幕修复: {sub_fixed} 条')
print(f'  目录修复: {fixed_dir} 条')
print(f'  总计: {sub_fixed + fixed_dir} 条')
print()
print('下一步: 运行 python py/generate_html.py 重新生成网站')
