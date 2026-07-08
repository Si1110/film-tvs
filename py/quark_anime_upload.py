#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
动漫资源批量上架脚本 - 处理 0630 源表中所有夸克条目
"""
import os, sys, re, json, time, urllib.parse, subprocess, io, shutil, glob
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_EXCEL = r'F:\1、自媒体\3、网站\影视\动漫资源（0630）.xlsx'
DATA_EXCEL = os.path.join(PROJECT_ROOT, 'res', 'data_new.xlsx')
COVERS_SRC = r'F:\1、自媒体\3、网站\影视\影视封面\缺失封面\动漫'
COVERS_DIR = os.path.join(PROJECT_ROOT, 'res', 'covers')
DIRS_DIR = os.path.join(PROJECT_ROOT, 'res', 'dirs')
GENERATE_SCRIPT = os.path.join(PROJECT_ROOT, 'py', 'generate_html.py')
TARGET_SHEET = '动漫资源'

import requests
from openpyxl import load_workbook

# ========== Quark API ==========
class QuarkAPI:
    API_TOKEN = 'https://drive-h.quark.cn/1/clouddrive/share/sharepage/token?pr=ucpro&fr=pc'
    API_DETAIL = 'https://drive-h.quark.cn/1/clouddrive/share/sharepage/detail'

    def __init__(self, cookie=None):
        self.cookie = cookie or os.environ.get('QUARK_COOKIE', '')
        self._s = None
        self._stoken_cache = {}

    @property
    def session(self):
        if self._s is None:
            s = requests.Session()
            if self.cookie:
                for part in self.cookie.split(';'):
                    part = part.strip()
                    if '=' in part:
                        k, v = part.split('=', 1)
                        s.cookies[k.strip()] = v.strip()
            self._s = s
        return self._s

    def _h(self):
        return {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Referer': 'https://pan.quark.cn/',
            'Origin': 'https://pan.quark.cn',
            'Content-Type': 'application/json',
        }

    def get_stoken(self, pwd_id):
        if pwd_id in self._stoken_cache:
            return self._stoken_cache[pwd_id]
        r = self.session.post(self.API_TOKEN, json={
            'pwd_id': pwd_id, 'passcode': '', 'support_visit_limit_private_share': True,
        }, headers=self._h(), timeout=15)
        stoken = r.json()['data']['stoken']
        self._stoken_cache[pwd_id] = stoken
        return stoken

    def list_folder(self, pwd_id, pdir_fid='0'):
        stoken = self.get_stoken(pwd_id)
        url = (f'{self.API_DETAIL}?pr=ucpro&fr=pc&ver=2'
               f'&pwd_id={pwd_id}&stoken={urllib.parse.quote(stoken)}'
               f'&pdir_fid={pdir_fid}&force=0&_page=1&_size=200&_fetch_total=1')
        r = self.session.get(url, headers=self._h(), timeout=15)
        d = r.json()
        return d.get('data', {}).get('list', []) if d.get('status') == 200 else []

    def build_tree(self, pwd_id, pdir_fid='0', depth=0):
        files = self.list_folder(pwd_id, pdir_fid)
        tree = []
        for f in files:
            entry = {
                'name': f['file_name'],
                'dir': f.get('dir', False),
                'fid': f['fid'],
                'size': f.get('size', 0),
            }
            if entry['dir']:
                entry['children'] = self.build_tree(pwd_id, entry['fid'], depth + 1)
                entry['count'] = f.get('include_items', len(entry.get('children', [])))
            tree.append(entry)
        tree.sort(key=lambda x: (not x['dir'], x['name']))
        return tree

# ========== Dir HTML ==========
def generate_dir_html(tree, title='目录'):
    """从夸克目录树生成目录 HTML"""
    def render_items(items, depth=0):
        html = ''
        ml = depth * 24
        for item in items:
            if item['dir']:
                count = item.get('count', len(item.get('children', [])))
                badge = f'<span class="badge bg-secondary ms-2" style="font-size:0.7rem;">{count} 项</span>'
                html += f'<li style="margin-left:{ml}px"><strong>{item["name"]}</strong>{badge}<ul>'
                html += render_items(item.get('children', []), depth + 1)
                html += '</ul></li>'
            else:
                size_mb = item.get('size', 0) / (1024*1024)
                if size_mb > 1024:
                    size_str = f'{size_mb/1024:.1f} GB'
                else:
                    size_str = f'{size_mb:.0f} MB'
                html += f'<li style="margin-left:{ml}px">{item["name"]} <span class="text-muted" style="font-size:0.8rem;">({size_str})</span></li>'
        return html

    content = render_items(tree)
    return f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <title>{title} - 目录</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
  <style>
    body {{ background: #1a1a2e; color: #e0e0e0; padding: 20px; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; }}
    ul {{ list-style: none; padding-left: 0; }}
    li {{ padding: 3px 0; }}
    .badge {{ background: #4a4a6a; }}
  </style>
</head>
<body>
  <div class="container">
    <h4 class="mb-3">{title}</h4>
    <div class="card" style="background:#16213e;border:1px solid #0f3460;">
      <div class="card-body">
        <ul>{content}</ul>
      </div>
    </div>
    <div class="text-muted small mt-2" style="border-top:1px solid rgba(255,255,255,0.06);padding-top:10px;">
      <i class="bi bi-info-circle"></i> 以上目录仅供参考，具体以夸克网盘实际内容为准
    </div>
  </div>
</body>
</html>'''

# ========== Data helpers ==========
def read_source():
    """读取用户源表"""
    wb = load_workbook(SOURCE_EXCEL)
    ws = wb['Sheet1']
    entries = []
    for r in range(2, ws.max_row + 1):
        original_name = str(ws.cell(r, 1).value or '').strip()
        final_name = str(ws.cell(r, 2).value or '').strip()
        source = str(ws.cell(r, 3).value or '').strip()
        link = str(ws.cell(r, 4).value or '').strip()
        genre = str(ws.cell(r, 5).value or '').strip()
        region = str(ws.cell(r, 6).value or '').strip()

        if final_name and link:
            pwd_match = re.search(r'/s/([a-zA-Z0-9]+)', link)
            pwd_id = pwd_match.group(1) if pwd_match else ''
            entries.append({
                'original_name': original_name,
                'final_name': final_name,
                'source': source,
                'link': link,
                'pwd_id': pwd_id,
                'genre': genre,
                'region': region,
                'row': r,
            })
    wb.close()
    return entries

def read_existing_data():
    """读取 data_new.xlsx 动漫资源所有现有条目"""
    wb = load_workbook(DATA_EXCEL)
    ws = wb[TARGET_SHEET]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for k, c in headers.items():
            row[k] = str(ws.cell(r, c).value or '').strip()
        rows.append(row)
    wb.close()
    return {'headers': headers, 'rows': rows}

def find_existing_by_link(rows, link):
    """通过链接精确匹配已有条目"""
    pwd_match = re.search(r'/s/([a-zA-Z0-9]+)', link)
    if not pwd_match:
        return None
    target_pwd = pwd_match.group(1)
    for row in rows:
        elink = row.get('下载链接', '')
        exist_pwd = re.search(r'/s/([a-zA-Z0-9]+)', elink)
        if exist_pwd and exist_pwd.group(1) == target_pwd:
            return row
    return None

def find_existing_by_title(rows, title):
    """通过标题匹配已有条目"""
    clean = re.sub(r'[\s\u3000【】\[\]（）()]+', '', title).lower()
    for row in rows:
        etitle = row.get('主标题', '')
        eclean = re.sub(r'[\s\u3000【】\[\]（）()]+', '', etitle).lower()
        if clean and (clean == eclean or clean in eclean or eclean in clean):
            return row
    return None

def find_series_for_new(rows, title):
    """为新条目推断所属系列"""
    clean = re.sub(r'[\s\u3000]+', '', title).lower()
    prefix = re.sub(r'[【\[\(（\[]\d+.*?[】\]\)）\]]', '', title).strip()
    prefix = re.sub(r'【.*?】', '', prefix).strip()
    prefix = re.sub(r'[\s\u3000]+', '', prefix).lower()

    best_match = None
    best_len = 0

    for row in rows:
        etitle = row.get('主标题', '')
        eseries = row.get('所属系列', '')
        if not etitle or not eseries:
            continue
        etitle_clean = re.sub(r'[\s\u3000【】\[\]（）()]+', '', etitle).lower()

        if prefix and (etitle_clean.startswith(prefix) or prefix.startswith(etitle_clean) or
                        prefix in etitle_clean or etitle_clean in prefix):
            common = len(os.path.commonprefix([prefix, etitle_clean]))
            if common > best_len and common >= 2:
                best_len = common
                best_match = eseries

    if best_match:
        return best_match

    # 动漫默认系列
    return '精选动画系列'

def find_cover(final_name, covers_src):
    """查找封面图"""
    for ext in ['.webp', '.jpg', '.jpeg', '.png']:
        src = os.path.join(covers_src, final_name + ext)
        if os.path.exists(src):
            return src, ext
    return None, None

def extract_info_from_tree(tree, depth=0):
    """从目录树中提取语言/字幕/格式信息"""
    all_names = []
    formats = set()
    langs = set()
    subs = set()

    def walk(items):
        for item in items:
            name = item['name']
            all_names.append(name)
            if not item['dir']:
                ext = os.path.splitext(name)[1].lower()
                if ext in ['.mp4', '.mkv', '.avi', '.rmvb', '.ts', '.flv', '.wmv']:
                    formats.add(ext[1:])
            if item.get('children'):
                walk(item['children'])

    walk(tree)

    full_text = ' '.join(all_names).lower()

    # 语言检测
    if '日语' in full_text or '日文' in full_text or 'japanese' in full_text:
        langs.add('日语')
    if '国语' in full_text or '普通话' in full_text or '中文配音' in full_text:
        langs.add('国语')
    if '粤语' in full_text or '广东话' in full_text:
        langs.add('粤语')
    if '英语' in full_text or 'english' in full_text:
        langs.add('英语')
    if '韩语' in full_text or '韩文' in full_text or 'korean' in full_text:
        langs.add('韩语')

    # 字幕检测
    if '中字' in full_text or '中文字幕' in full_text or '双语' in full_text:
        subs.add('中文')
    if '英字' in full_text or '英文字幕' in full_text:
        subs.add('英文字幕')
    if '内封' in full_text:
        subs.add('内封字幕')
    if '简繁' in full_text:
        subs.add('简繁字幕')
    if '简体' in full_text:
        subs.add('简体字幕')
    if '繁体' in full_text:
        subs.add('繁体字幕')

    # 如果没检测到，默认值
    if not langs:
        if '日本' in str(all_names):
            langs.add('日语')
        else:
            langs.add('日语')  # 动漫默认日语
    if not subs:
        subs.add('中文')

    lang_str = '/'.join(sorted(langs)) if langs else '日语'
    sub_str = '/'.join(sorted(subs)) if subs else '中文'
    fmt_str = '/'.join(sorted(formats)) if formats else 'mp4'

    return lang_str, sub_str, fmt_str

DESC_FILE = os.path.join(PROJECT_ROOT, 'py', 'anime_descriptions.json')

# ========== Main ==========
def main():
    QUARK_COOKIE = 'b-user-id=47000940-681b-3d39-3ad3-ffd51e0bef3a; _UP_A4A_11_=wba2d199ca6d4156bbf065be9709fec0; _c_WBKFRo=geWYEMxjjBz1P0w0V5kjZMCOIKZTvcVAWZmcfw6a; _UP_335_2B_=1; __sdid=AASjLyqoKdwFTpUnYB15zKlS3acNl/8PW/P0YAW4kN8dfIgFKJhVplRuuO3OK1LOpOiG8RyVyKLGVP5N19E2DjX5rJk/Y6tnlcNmc+YuFhk1lw==; _UP_D_=pc; _UP_30C_6A_=sta2e6201d15vu1n67adq3lv00werwuk; _UP_TS_=sg1ace6b77f3ab536d7cbec3937c8f84f8e; _UP_E37_B7_=sg1ace6b77f3ab536d7cbec3937c8f84f8e; _UP_TG_=sta2e6201d15vu1n67adq3lv00werwuk; xlly_s=1; __pus=c8dabec9a56b73b76a9a17a5d5556484AARg1ZBbAdQ50a/l7n6K0rdTMIZja9q6Bbr4Jm5FFyQR7AhGTB+ANApk6+uLPJTwCJU0FC7foDEsF2jIIJyz+0Zc; __kp=d0fc4810-76a2-11f1-9926-dfd29c6e79a4; __kps=AASVveG4ImeBJ6LsHBekuXpv; __ktd=Syw0VaFC+r7411Krkk5Ofw==; __uid=AASVveG4ImeBJ6LsHBekuXpv; isg=BBISyZ9nrOXsQND94VoVrIoLY9j0Ixa9GU0V6txrPkWw77LpxLNmzRiMX0tTn45V; tfstk=gVbZxOXxzPUaH72la_Y4LFVmyfT9kEyWiZ9XisfD1dvgDsYmot1X5n2OG98VMs06IInxOiWOjhpbBFxc3_1lfVTgIp0hyCXADnIs6MB5ZV__XEg2WEL0F8a7PC19kEjLChlJB6fpiWOGktFhEEL0FJa7P1C9kO6xz2LGtWRXGccGnVVexIJnSI0MivVeMpYmSqvi-2AWwKciotf3TIpHnEYcneVeMpvDoEc41KvS8Q-g6iB9jEm0YxpPspumow4vbL5icq0c8C-No1vE_f7e_hJkc9NUYNXCiN_9vloMk6sP3iXudc9G4QWeVGyr7TW9isRcQzhyKN5V8hIIIcAy7d-lShlSr1KcYw81S-nRO6JM4esQvRtX7OSJehqLpt5eCtbwxAyBhg1CSHWud2Wv0ib6-N23zgoxHBxSmZIZnmxM9BJ7TWyf8UduZlP5WmnvXMdeF5FtDmKM9BJ7TWoxDhLpTLNT6; __puus=033395a8f2587bc08601d7067e33b116AAQeiTZSRJIWXaasXH+OiBTRqVsDAkmlU7YhJROE5s7yzp5I7rldGwFxFn348AVEIcmexBnBa+2/oINbeI4cajL2Y+yf+s3mVA7Qg1oiT40rpySZSuKU5JaiOkM3qHkSjNuhT9He9KMckHdSSGX6Ntb+qpuagIbWCfMtldgmhl4Kk2BhkWCWyKfkeXv+I3i2dnNvC456mKCDrNVv0aX4BGE6'

    print('=' * 60)
    print('动漫资源批量上架脚本 (0630)')
    print('=' * 60)

    # Load descriptions
    descriptions = {}
    if os.path.exists(DESC_FILE):
        with open(DESC_FILE, 'r', encoding='utf-8') as f:
            descriptions = json.load(f)
        print(f'  已加载 {len(descriptions)} 个简介')

    # Step 1: 读取源表
    print('\n[1/7] 读取源表...')
    entries = read_source()
    print(f'  找到 {len(entries)} 个资源')
    
    # Assign descriptions
    for e in entries:
        e['description'] = descriptions.get(e['final_name'], '')

    # Step 2: 读取已有数据
    print('\n[2/7] 读取已有数据...')
    existing = read_existing_data()
    existing_rows = existing['rows']
    headers = existing['headers']
    print(f'  已有 {len(existing_rows)} 条动漫资源')

    # Step 3: 检查哪些是新增、哪些是更新
    new_entries = []
    update_entries = []
    for e in entries:
        matched = find_existing_by_link(existing_rows, e['link'])
        if not matched:
            matched = find_existing_by_title(existing_rows, e['final_name'])
        if matched:
            e['existing_row'] = matched
            update_entries.append(e)
        else:
            new_entries.append(e)

    print(f'  新增: {len(new_entries)}, 更新: {len(update_entries)}')

    # Step 4: 获取夸克目录信息
    print('\n[3/7] 获取夸克网盘目录信息...')
    api = QuarkAPI(cookie=QUARK_COOKIE)
    dir_results = {}
    for i, e in enumerate(entries):
        pwd_id = e['pwd_id']
        if not pwd_id:
            print(f'  [{i+1}/{len(entries)}] {e["final_name"]}: 无有效链接，跳过')
            continue
        try:
            tree = api.build_tree(pwd_id)
            dir_results[pwd_id] = tree
            lang, sub, fmt = extract_info_from_tree(tree)
            e['lang'] = lang
            e['sub'] = sub
            e['fmt'] = fmt
            e['dir_tree'] = tree

            # 生成目录HTML
            dir_html = generate_dir_html(tree, e['final_name'])
            dir_path = os.path.join(DIRS_DIR, f'quark_{pwd_id}.html')
            with open(dir_path, 'w', encoding='utf-8') as f:
                f.write(dir_html)
            e['dir_path'] = f'quark_{pwd_id}'

            print(f'  [{i+1}/{len(entries)}] {e["final_name"]}: OK (语言:{lang}, 字幕:{sub}, 格式:{fmt})')
            time.sleep(0.5)
        except Exception as ex:
            print(f'  [{i+1}/{len(entries)}] {e["final_name"]}: ERROR - {ex}')
            e['lang'] = '日语'
            e['sub'] = '中文'
            e['fmt'] = 'mp4'
            e['dir_path'] = ''

    # Step 5: 匹配封面图
    print('\n[4/7] 匹配封面图...')
    for e in entries:
        src, ext = find_cover(e['final_name'], COVERS_SRC)
        if src:
            # 复制到项目封面目录
            series_dir = os.path.join(COVERS_DIR, '动漫资源')
            os.makedirs(series_dir, exist_ok=True)
            dst = os.path.join(series_dir, e['final_name'] + ext)
            if not os.path.exists(dst):
                shutil.copy2(src, dst)
            e['cover_path'] = f'../res/covers/动漫资源/{e["final_name"]}{ext}'
            print(f'  {e["final_name"]}: 找到封面')
        else:
            e['cover_path'] = ''
            print(f'  {e["final_name"]}: 未找到封面')

    # Step 6: 推断系列
    print('\n[5/7] 推断所属系列...')
    for e in new_entries:
        series = find_series_for_new(existing_rows, e['final_name'])
        e['series'] = series
        print(f'  {e["final_name"]} -> {series}')

    for e in update_entries:
        e['series'] = e['existing_row'].get('所属系列', '精选动画系列')

    # Step 7: 写入 data_new.xlsx
    print('\n[6/7] 写入 data_new.xlsx...')
    wb = load_workbook(DATA_EXCEL)
    ws = wb[TARGET_SHEET]

    # 找到最后一个非空行
    last_row = ws.max_row
    while last_row > 1 and not ws.cell(last_row, 1).value:
        last_row -= 1

    next_row = last_row + 1
    print(f'  从第 {next_row} 行开始写入')

    for e in new_entries:
        row = next_row
        ws.cell(row, headers['所属系列']).value = e.get('series', '精选动画系列')
        ws.cell(row, headers['封面图片路径']).value = e.get('cover_path', '')
        ws.cell(row, headers['主标题']).value = e['final_name']
        ws.cell(row, headers['副标题']).value = e.get('original_name', '')
        ws.cell(row, headers['概要']).value = e.get('description', '')
        ws.cell(row, headers['语言']).value = e.get('lang', '日语')
        ws.cell(row, headers['字幕']).value = e.get('sub', '中文')
        ws.cell(row, headers['目录路径']).value = e.get('dir_path', '')
        ws.cell(row, headers['网盘名称']).value = '夸克网盘'
        ws.cell(row, headers['下载链接']).value = e['link']
        ws.cell(row, headers['解压密码']).value = ''
        ws.cell(row, headers['支持格式']).value = e.get('fmt', 'mp4')
        ws.cell(row, headers['卡片页脚']).value = e.get('genre', '')
        ws.cell(row, headers['类型']).value = e.get('genre', '')
        next_row += 1

    wb.save(DATA_EXCEL)
    print(f'  写入 {len(new_entries)} 条新记录')

    # Step 8: 生成 HTML
    print('\n[7/7] 生成 HTML...')
    result = subprocess.run([sys.executable, GENERATE_SCRIPT], capture_output=True, text=True, encoding='utf-8')
    if result.returncode == 0:
        print('  HTML 生成成功')
    else:
        print(f'  HTML 生成失败: {result.stderr[:500]}')

    print('\n' + '=' * 60)
    print(f'完成！新增 {len(new_entries)} 个资源')
    print('=' * 60)

if __name__ == '__main__':
    main()
