import sys; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl; wb=openpyxl.load_workbook('E:/workspace/github/film-tvs/res/data_new.xlsx')
for sn in ['电视剧资源','电影资源','动漫资源']:
    ws=wb[sn]
    for r in range(2, ws.max_row+1):
        t=str(ws.cell(r,3).value or '')[:25]
        c8=str(ws.cell(r,8).value or '')
        c11=str(ws.cell(r,11).value or '')
        if '杰森' in t or '录像带' in t or '斯坦森' in t:
            print(f'{sn} row{r}: {t} | Col8(目录路径)=\"{c8}\" | Col11(解压密码)=\"{c11}\"')
