import sys, re, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl import load_workbook

EXCEL = r'E:\workspace\github\film-tvs\res\data_new.xlsx'
DIRS_DIR = r'E:\workspace\github\film-tvs\res\dirs'

wb = load_workbook(EXCEL)

def extract_surl(url):
    """Extract surl from Baidu share URL"""
    m = re.search(r'/s/([a-zA-Z0-9_-]+)', url)
    if m:
        return m.group(1)
    return None

def generate_placeholder_html(title, surl, source='百度网盘'):
    return f'''<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><title>{title} - 目录</title>
<style>
body {{ font-family: 'Microsoft YaHei', sans-serif; background: #1a1a2e; padding: 20px; font-size: 14px; line-height: 1.8; }}
a {{ color: #4fc3f7; text-decoration: none; }}
a:hover {{ text-decoration: underline; }}
.placeholder {{ color: #888; text-align: center; margin-top: 40px; }}
</style>
</head><body>
<div class="placeholder">
  <div style="font-size: 18px; color: #ffd700;">📁 {title}</div>
  <div style="margin-top: 10px; color: #666;">资源目录暂未获取，请直接点击下方按钮访问{source}查看</div>
  <div style="margin-top: 15px;">
    <a href="https://pan.baidu.com/s/{surl}" target="_blank" style="display: inline-block; padding: 10px 24px; background: #4fc3f7; color: #fff; border-radius: 6px; text-decoration: none;">打开{source}</a>
  </div>
</div>
</body></html>'''

updates = []
for sheet_name in [n for n in wb.sheetnames if n != 'index']:
    ws = wb[sheet_name]
    for row in range(2, ws.max_row + 1):
        dir_key = str(ws.cell(row, 11).value or '')
        link = str(ws.cell(row, 10).value or '')
        title = str(ws.cell(row, 3).value or '')[:30]
        
        # If dir_key is empty and link is Baidu
        if not dir_key and 'pan.baidu.com' in link:
            surl = extract_surl(link)
            if surl:
                key = f'baidu_{surl}'
                html_path = os.path.join(DIRS_DIR, f'{key}.html')
                
                # Generate placeholder HTML
                html = generate_placeholder_html(title, surl)
                with open(html_path, 'w', encoding='utf-8') as f:
                    f.write(html)
                
                # Update Excel
                ws.cell(row, 11).value = key
                updates.append(f'{sheet_name} row{row}: {title} -> {key}')
                print(f'  [{sheet_name}] row{row}: {title} -> {key}')

# Save Excel
wb.save(EXCEL)
print(f'\nDone: {len(updates)} Baidu placeholder HTML files created')
