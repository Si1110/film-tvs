#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
批量上架脚本 - 处理 0607 源表中所有夸克条目
"""
import os, sys, re, json, time, urllib.parse, subprocess, io
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_EXCEL = os.environ.get('FILM_TVS_SOURCE', r'F:\1、自媒体\3、网站\影视\影视动漫新增（0607）.xlsx')
DATA_EXCEL = os.path.join(PROJECT_ROOT, 'res', 'data_new.xlsx')
COVERS_DIR = os.path.join(PROJECT_ROOT, 'res', 'covers')
DIRS_DIR = os.path.join(PROJECT_ROOT, 'res', 'dirs')
GENERATE_SCRIPT = os.path.join(PROJECT_ROOT, 'py', 'generate_html.py')

SHEET_MAP = {'动漫': '动漫资源', '电视剧': '影视资源', '电影': '影视资源'}

import requests
from PIL import Image, ImageDraw
from openpyxl import load_workbook

# ========== Quark API ==========
class QuarkAPI:
    API_TOKEN = 'https://drive-h.quark.cn/1/clouddrive/share/sharepage/token?pr=ucpro&fr=pc'
    API_DETAIL = 'https://drive-h.quark.cn/1/clouddrive/share/sharepage/detail'

    def __init__(self, cookie=None):
        self.cookie = cookie or os.environ.get('QUARK_COOKIE', '')
        self._s = None
        self._stoken_cache = {}

    @property
    def session(self):
        if self._s is None:
            s = requests.Session()
            if self.cookie:
                for part in self.cookie.split(';'):
                    part = part.strip()
                    if '=' in part:
                        k, v = part.split('=', 1)
                        s.cookies[k.strip()] = v.strip()
            self._s = s
        return self._s

    def _h(self):
        return {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Referer': 'https://pan.quark.cn/',
            'Origin': 'https://pan.quark.cn',
            'Content-Type': 'application/json',
        }

    def get_stoken(self, pwd_id):
        if pwd_id in self._stoken_cache:
            return self._stoken_cache[pwd_id]
        r = self.session.post(self.API_TOKEN, json={
            'pwd_id': pwd_id, 'passcode': '', 'support_visit_limit_private_share': True,
        }, headers=self._h(), timeout=15)
        stoken = r.json()['data']['stoken']
        self._stoken_cache[pwd_id] = stoken
        return stoken

    def list_folder(self, pwd_id, pdir_fid='0'):
        stoken = self.get_stoken(pwd_id)
        url = (f'{self.API_DETAIL}?pr=ucpro&fr=pc&ver=2'
               f'&pwd_id={pwd_id}&stoken={urllib.parse.quote(stoken)}'
               f'&pdir_fid={pdir_fid}&force=0&_page=1&_size=200&_fetch_total=1')
        r = self.session.get(url, headers=self._h(), timeout=15)
        d = r.json()
        return d.get('data', {}).get('list', []) if d.get('status') == 200 else []

    def build_tree(self, pwd_id, pdir_fid='0', depth=0):
        files = self.list_folder(pwd_id, pdir_fid)
        tree = []
        for f in files:
            entry = {
                'name': f['file_name'],
                'dir': f.get('dir', False),
                'fid': f['fid'],
                'size': f.get('size', 0),
            }
            if entry['dir']:
                entry['children'] = self.build_tree(pwd_id, entry['fid'], depth + 1)
                entry['count'] = f.get('include_items', len(entry.get('children', [])))
            tree.append(entry)
        tree.sort(key=lambda x: (not x['dir'], x['name']))
        return tree

# ========== Dir HTML ==========
def generate_dir_html(tree, title='目录'):
    """从夸克目录树生成目录 HTML（支持多级文件夹嵌套）"""
    def render_items(items, depth=0):
        html = ''
        ml = depth * 24
        for item in items:
            if item['dir']:
                count = item.get('count', len(item.get('children', [])))
                badge = f'<span class="badge bg-secondary ms-2" style="font-size:0.7rem;">{count} 项</span>'
                html += f'<div style="margin-left:{ml}px" class="mb-1">\n'
                html += f'  <div class="d-flex align-items-center mb-1">\n'
                html += f'    <span style="color:#ffd700;">&#x1f4c1; {item["name"]}</span>\n'
                html += f'    {badge}\n'
                html += f'  </div>\n'
                if item.get('children'):
                    html += render_items(item['children'], depth + 1)
                html += f'</div>\n'
            else:
                html += f'<div style="margin-left:{ml}px;color:#aaa;font-size:0.9rem;" class="mb-1">&#x1f4c4; {item["name"]}</div>\n'
        return html

    body = render_items(tree)
    return f'''<div class="dir-list">
  <div class="dir-section mb-3">
    <h6 style="color:#ffd700;border-bottom:1px solid rgba(255,215,0,0.2);padding-bottom:8px;margin-bottom:12px;">
      <i class="bi bi-folder2-open"></i> 夸克网盘目录
    </h6>
    {body}
  </div>
  <div class="text-muted small mt-2" style="border-top:1px solid rgba(255,255,255,0.06);padding-top:10px;">
    <i class="bi bi-info-circle"></i> 以上目录仅供参考，具体以夸克网盘实际内容为准
  </div>
</div>'''

# ========== Data helpers ==========
def read_source():
    wb = load_workbook(SOURCE_EXCEL)
    quark_entries = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(2, ws.max_row + 1):
            name = str(ws.cell(r, 1).value or '').strip()
            source = str(ws.cell(r, 2).value or '').strip()
            link = str(ws.cell(r, 3).value or '').strip()
            if name and '夸克' in source and link:
                quark_entries.append({
                    'sheet': sn, 'target_sheet': SHEET_MAP.get(sn, '影视资源'),
                    'name': name, 'source': source, 'link': link, 'row': r
                })
    wb.close()
    return quark_entries

def read_existing_data():
    """读取 data_new.xlsx 所有现有条目"""
    wb = load_workbook(DATA_EXCEL)
    result = {}
    for sn in ['影视资源', '动漫资源']:
        ws = wb[sn]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        rows = []
        for r in range(2, ws.max_row + 1):
            row = {}
            for k, c in headers.items():
                row[k] = str(ws.cell(r, c).value or '').strip()
            rows.append(row)
        result[sn] = {'headers': headers, 'rows': rows, 'ws': ws}
    wb.close()
    return result

def clean_title(name):
    """从源表名称中提取干净标题（去除前缀字母）"""
    name = re.sub(r'^[A-Z]\s+', '', name)
    return name.strip()

def normalize(s):
    """去掉空格和特殊字符"""
    return re.sub(r'[\s\u3000]+', '', s).lower()

def find_matching_row(rows, source_name, link):
    """在现有行中查找匹配项
    
    匹配优先级：链接精确匹配 > 标题+链接联合匹配 > 系列名匹配
    注意：仅有标题相同但链接不同时，视为新条目（不同合集下的同名资源）
    """
    clean = clean_title(source_name)
    nclean = normalize(clean)
    src_pwd = re.search(r'/s/([a-zA-Z0-9]+)', link)

    for row in rows:
        etitle = row.get('主标题', '')
        elink = row.get('下载链接', '')
        eseries = row.get('所属系列', '')
        ntitle = normalize(etitle)
        nseries = normalize(eseries)
        exist_pwd = re.search(r'/s/([a-zA-Z0-9]+)', elink)

        # 1. Link exact match (same pwd_id)
        if src_pwd and exist_pwd and src_pwd.group(1) == exist_pwd.group(1):
            return row, 'link'

        # 2. Title match + same link pattern → existing item with different series
        if nclean and (nclean == ntitle or nclean in ntitle or ntitle in nclean):
            # Same title but different link → different resource, add as NEW
            if src_pwd and exist_pwd and src_pwd.group(1) != exist_pwd.group(1):
                continue  # treat as new
            return row, 'title'

        # 3. Series name match
        if nclean and nseries.replace('系列', '') in nclean:
            return row, 'series'

    return None, None

def find_series_for_new(rows, source_name):
    """对于新条目，决定放在哪个系列下
    
    策略：如果标题前缀匹配某个已有条目，继承它的系列。
    例如 唐朝诡事录之西行 → 与 唐朝诡事录 同系列。
    否则返回 None（创建独立系列）。
    """
    clean = clean_title(source_name)
    nclean = normalize(clean)

    # 1. Extract key prefix: remove year, version info
    prefix = re.sub(r'[（(]\d{4}[）)].*', '', clean).strip()
    prefix = re.sub(r'[【\[\(（\[]\d+.*?[】\]\)）\]]', '', prefix).strip()
    prefix = re.sub(r'\s+', '', prefix)
    nprefix = normalize(prefix)

    best_match = None
    best_len = 0

    for row in rows:
        etitle = row.get('主标题', '')
        eseries = row.get('所属系列', '')
        if not etitle or not eseries:
            continue

        etitle_clean = normalize(etitle)

        # Check if existing title starts with or contains the prefix
        # e.g. 神探狄仁杰 is prefix of 神探狄仁杰之西行
        if nprefix and (etitle_clean.startswith(nprefix) or nprefix.startswith(etitle_clean) or
                        nprefix in etitle_clean or etitle_clean in nprefix):
            # Longer match = higher confidence
            common = len(os.path.commonprefix([nprefix, etitle_clean]))
            if common > best_len and common >= 3:  # at least 3 chars to avoid false match
                best_len = common
                best_match = eseries

    if best_match:
        return best_match

    # 2. For anime: fallback to 精选动画系列
    for row in rows:
        eseries = row.get('所属系列', '')
        if '精选动画' in eseries:
            etitle = normalize(row.get('主标题', ''))
            if nclean in etitle or etitle in nclean:
                return eseries

    return None

def batch_process():
    quark_entries = read_source()
    existing = read_existing_data()

    print(f'找到 {len(quark_entries)} 个夸克条目\n')

    new_entries = []
    update_entries = []

    for e in quark_entries:
        ts = e['target_sheet']
        rows = existing[ts]['rows']
        matched_row, match_type = find_matching_row(rows, e['name'], e['link'])

        if matched_row:
            update_entries.append({**e, 'existing_row': matched_row})
            print(f'  [已有] {e["sheet"]}/{e["name"]} → 匹配方式: {match_type}')
        else:
            new_entries.append(e)
            print(f'  [新增] {e["sheet"]}/{e["name"]}')

    print(f'\n待更新: {len(update_entries)} 个')
    print(f'待新增: {len(new_entries)} 个')

    if not (update_entries or new_entries):
        print('❌ 没有需要处理的内容')
        return

    # Auto-continue (non-interactive mode)
    print('\n⏳ 开始批量处理...')

    # Open workbook for editing
    wb = load_workbook(DATA_EXCEL)
    quark = QuarkAPI()

    # Process updates first
    print('\n' + '=' * 60)
    print('  处理已有条目（更新链接 + 补充目录）')
    print('=' * 60)
    updates_done = 0
    for e in update_entries:
        old_row = e['existing_row']
        ts = e['target_sheet']
        ws = wb[ts]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

        # Find actual row index by matching the link
        row_idx = None
        for r in range(2, ws.max_row + 1):
            elink = str(ws.cell(r, headers['下载链接']).value or '').strip()
            old_link = old_row.get('下载链接', '')
            if re.sub(r'\?pwd=.*', '', elink) == re.sub(r'\?pwd=.*', '', old_link):
                row_idx = r
                break

        if row_idx is None:
            print(f'  ❌ [{e["name"]}] 找不到对应行')
            continue

        print(f'\n  [{e["sheet"]}] {e["name"]}')

        # Update download link
        ws.cell(row_idx, headers['下载链接']).value = e['link']
        ws.cell(row_idx, headers['网盘名称']).value = '夸克网盘'

        # Fetch Quark dir
        pwd_id = re.search(r'/s/([a-zA-Z0-9]+)', e['link'])
        if pwd_id:
            pwd_id = pwd_id.group(1)
            try:
                root_files = quark.list_folder(pwd_id)
                folder_fid = '0'
                for f in root_files:
                    if f.get('dir'):
                        folder_fid = f['fid']
                        break
                if folder_fid != '0':
                    dir_tree = quark.build_tree(pwd_id, folder_fid)
                    if dir_tree:
                        os.makedirs(DIRS_DIR, exist_ok=True)
                        # Generate series key based on existing series name
                        series_name_full = old_row.get('所属系列', '')
                        # For broad series like 精选动画系列, use title as key
                        dir_key = series_name_full
                        if '精选动画' in series_name_full or '其他' in series_name_full or '国产' in series_name_full:
                            dir_key = series_name_full  # Keep series-level dir key

                        dir_html = generate_dir_html(dir_tree, clean_title(e['name']))
                        dir_file = os.path.join(DIRS_DIR, f'{dir_key}.html')
                        with open(dir_file, 'w', encoding='utf-8') as f:
                            f.write(dir_html)
                        print(f'    ✅ 目录已生成: {dir_key}')

                        # Update 目录路径 column
                        ws.cell(row_idx, headers['目录路径']).value = dir_key

                        # Generate description from dir tree
                        version_parts = [item['name'] for item in dir_tree if item['dir']]
                        if version_parts:
                            total = sum(len(item.get('children', [])) for item in dir_tree if item['dir'])
                            desc = f'收录了{", ".join(version_parts[:5])}'
                            if len(version_parts) > 5:
                                desc += f'等{len(version_parts)}个版本'
                            desc += f'的{clean_title(e["name"])}全集。'
                            ws.cell(row_idx, headers['概要']).value = desc
                            if total:
                                ws.cell(row_idx, headers['卡片页脚']).value = f'{len(version_parts)}个版本 共约{total}集'
                        print(f'    ✅ 目录已添加')
            except Exception as ex:
                print(f'    ⚠️ 夸克目录读取失败: {ex}')

        updates_done += 1

    # Save intermediate
    wb.save(DATA_EXCEL)
    print(f'\n✅ 已完成 {updates_done} 个条目的更新')

    # Process new entries
    print('\n' + '=' * 60)
    print('  处理新增条目')
    print('=' * 60)

    new_done = 0
    for e in new_entries:
        ts = e['target_sheet']
        ws = wb[ts]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

        print(f'\n  [{e["sheet"]}] {e["name"]}')
        print(f'    链接: {e["link"]}')

        clean = clean_title(e['name'])
        link = e['link']

        # base = stripped title for series name & cover key generation
        # (aggressive removal: year, version, brackets, etc.)
        base = clean
        # Remove version/size/quality info from brackets (for series name & cover key)
        base = re.sub(r'[【\[\(（\[]\d+个版本.*?[】\)）\]]', '', base).strip()
        base = re.sub(r'[【\[\(（\[]\d+.*?[】\)）\]]', '', base).strip()
        # Remove extra details
        base = re.sub(r'\s*[（(]\d{4}[）)]', '', base)
        base = re.sub(r'\s*\d+K.*', '', base)
        base = re.sub(r'\s*收藏版.*', '', base)
        base = re.sub(r'\s*修复版.*', '', base)
        base = re.sub(r'全\d+.*', '', base)
        base = re.sub(r'\s+', '', base).strip()

        series_name = base + '系列'
        # Remove bad chars from series name
        series_name = re.sub(r'[\\/:*?"<>|]', '', series_name)

        # Determine what series this new entry belongs to
        put_under_series = series_name
        
        # 1. Check if title prefix matches existing entry → inherit its series
        #    例如 唐朝诡事录之西行 → 继承 唐朝诡事录 的系列
        matched_series = find_series_for_new(existing[ts]['rows'], e['name'])
        if matched_series:
            put_under_series = matched_series
            print(f'    📎 继承已有条目系列: {matched_series}')
        
        # 2. For anime: put in 精选动画系列 if applicable
        if ts == '动漫资源' and not matched_series:
            existing_series = set()
            for r in range(2, ws.max_row + 1):
                v = str(ws.cell(r, headers['所属系列']).value or '').strip()
                if v:
                    existing_series.add(v)
            if '精选动画系列' in existing_series:
                put_under_series = '精选动画系列'

        # Auto-generate fields
        pwd_id = re.search(r'/s/([a-zA-Z0-9]+)', link)
        pwd_id = pwd_id.group(1) if pwd_id else ''

        if ts == '动漫资源':
            lang, sub = '日语', '中文字幕'
        else:
            # Detect Cantonese
            if '粤语' in clean:
                lang = '粤语'
                sub = '中文'
            else:
                lang = '汉语普通话'
                sub = '中文字幕'

        cover_key = re.sub(r'[^\w\u4e00-\u9fff]+', '_', base).strip('_')

        # Description (use full title for SEO)
        desc = f'{clean}全集，高清资源下载。'

        # Fetch Quark dir
        dir_tree = []
        if pwd_id:
            try:
                root_files = quark.list_folder(pwd_id)
                folder_fid = '0'
                for f in root_files:
                    if f.get('dir'):
                        folder_fid = f['fid']
                        break
                if folder_fid != '0':
                    dir_tree = quark.build_tree(pwd_id, folder_fid)
            except Exception as ex:
                print(f'    ⚠️ 夸克目录读取失败: {ex}')

        # Generate dir HTML (use series name as key)
        if dir_tree:
            os.makedirs(DIRS_DIR, exist_ok=True)
            dir_key = series_name
            dir_html = generate_dir_html(dir_tree, base)
            dir_file = os.path.join(DIRS_DIR, f'{dir_key}.html')
            with open(dir_file, 'w', encoding='utf-8') as f:
                f.write(dir_html)
            print(f'    ✅ 目录已生成: res/dirs/{dir_key}.html')

            version_parts = [item['name'] for item in dir_tree if item['dir']]
            if version_parts:
                total = sum(len(item.get('children', [])) for item in dir_tree if item['dir'])
                desc = f'收录了{", ".join(version_parts[:5])}'
                if len(version_parts) > 5:
                    desc += f'等{len(version_parts)}个版本'
                desc += f'的{base}全集。'
                if total:
                    card_foot = f'{len(version_parts)}个版本 共约{total}集'
                else:
                    card_foot = ''
            else:
                card_foot = ''
        else:
            dir_key = series_name
            card_foot = ''

        # Download cover
        cover_dir = os.path.join(COVERS_DIR, series_name)
        os.makedirs(cover_dir, exist_ok=True)
        cover_found = False

        for query in [base, clean_title(e['name'])]:
            if not query.strip():
                continue
            try:
                search_url = f'https://movie.douban.com/j/subject_suggest?q={urllib.parse.quote(query)}'
                r = requests.get(search_url, headers={
                    'User-Agent': 'Mozilla/5.0', 'Referer': 'https://movie.douban.com/',
                }, timeout=15)
                if r.status_code == 200:
                    results = r.json()
                    if results:
                        best = results[0]
                        for res in results:
                            if res.get('subtype') == 'tv':
                                best = res
                                break
                        poster = best.get('img', '')
                        if poster:
                            for size in ['l', 'm']:
                                url = re.sub(r'/[a-z]/', f'/{size}/', poster)
                                img_r = requests.get(url, headers={
                                    'User-Agent': 'Mozilla/5.0', 'Referer': 'https://movie.douban.com/',
                                }, timeout=15)
                                if len(img_r.content) > 5000:
                                    img = Image.open(io.BytesIO(img_r.content))
                                    webp_path = os.path.join(cover_dir, f'{cover_key}.webp')
                                    img.save(webp_path, 'WEBP', quality=90)
                                    jpg_path = webp_path.replace('.webp', '.jpg')
                                    img.save(jpg_path, 'JPEG', quality=90)
                                    print(f'    ✅ 封面已下载: {cover_key}')
                                    cover_found = True
                                    break
                            if cover_found:
                                break
            except Exception as ex:
                print(f'    ⚠️ 豆瓣搜索失败({query}): {ex}')

        if not cover_found:
            # Generate placeholder
            img = Image.new('RGB', (500, 700), (30, 30, 50))
            draw = ImageDraw.Draw(img)
            try:
                from PIL import ImageFont
                font = ImageFont.truetype('C:/Windows/Fonts/msyhbd.ttc', 36)
            except:
                font = ImageFont.load_default()
            draw.text((50, 300), base, fill='#ffd700', font=font)
            webp_path = os.path.join(cover_dir, f'{cover_key}.webp')
            img.save(webp_path, 'WEBP', quality=85)
            print(f'    ⚠️ 占位封面已生成')

        # Cover path
        cover_path = f'../res/covers/{series_name}/{cover_key}.webp'

        # Write to Excel
        new_row = ws.max_row + 1
        ws.cell(new_row, headers['所属系列']).value = put_under_series
        ws.cell(new_row, headers['封面图片路径']).value = cover_path
        ws.cell(new_row, headers['主标题']).value = clean
        ws.cell(new_row, headers['副标题']).value = ''
        ws.cell(new_row, headers['概要']).value = desc
        ws.cell(new_row, headers['语言']).value = lang
        ws.cell(new_row, headers['字幕']).value = sub
        ws.cell(new_row, headers['目录路径']).value = dir_key
        ws.cell(new_row, headers['网盘名称']).value = '夸克网盘'
        ws.cell(new_row, headers['下载链接']).value = link
        ws.cell(new_row, headers['解压密码']).value = ''
        ws.cell(new_row, headers['支持格式']).value = 'mp4/mkv'
        ws.cell(new_row, headers['卡片页脚']).value = card_foot
        print(f'    ✅ 已写入 {ts} 第 {new_row} 行')

        new_done += 1

    # Save final
    wb.save(DATA_EXCEL)
    print(f'\n✅ 新增完成: {new_done} 个')

    # Generate HTML
    print('\n⚙️  正在重新生成网站 HTML...')
    result = subprocess.run([sys.executable, GENERATE_SCRIPT],
                          cwd=PROJECT_ROOT, capture_output=True, text=True)
    if result.returncode == 0:
        print('✅ HTML 生成完成')
    else:
        print(f'❌ HTML 生成失败:\n{result.stderr[-1000:]}')
        sys.exit(1)

    # Commit
    print('\n📤 正在推送...')
    try:
        subprocess.run(['git', 'add', '-A'], cwd=PROJECT_ROOT, check=True, capture_output=True)
        subprocess.run(['git', 'commit', '-m', f'批量上架/更新夸克条目 ({updates_done}更新+{new_done}新增)'],
                      cwd=PROJECT_ROOT, check=True, capture_output=True)
        subprocess.run(['git', 'push'], cwd=PROJECT_ROOT, check=True, capture_output=True)
        print('✅ 已自动 commit + push')
    except subprocess.CalledProcessError as e:
        stderr = e.stderr.decode() if e.stderr else ''
        print(f'⚠️ Git 操作失败: {stderr[:500]}')

    print('\n' + '=' * 60)
    print(f'  全部完成! 更新 {updates_done} 个 + 新增 {new_done} 个')
    print('=' * 60)

if __name__ == '__main__':
    batch_process()
