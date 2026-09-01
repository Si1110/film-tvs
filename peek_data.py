from openpyxl import load_workbook
wb = load_workbook('res/data_new.xlsx')
print('Sheet names:', wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    headers = [c.value for c in ws[1]]
    print(f'\nSheet: {sn}')
    print('Headers:', headers)
    print('Row count:', ws.max_row - 1)
    # Print first 3 rows
    for r in range(2, min(5, ws.max_row + 1)):
        row_data = {}
        for c_idx, cell in enumerate(ws[r]):
            h = headers[c_idx] if c_idx < len(headers) else f'col_{c_idx}'
            row_data[h] = cell.value
        print(f'  Row {r}: {row_data}')